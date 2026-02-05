# =============================================================================
# STOCK ENGINE - Excel Exporter Service
# =============================================================================
# Versione: 1.0.0
# Data: 28 gennaio 2026
#
# Servizio per generazione file Excel output.
# Genera file simili a quelli prodotti dal sistema bash attuale.
# =============================================================================

from pathlib import Path
from datetime import date
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from flask import current_app

from app.models.veicolo import Veicolo


class ExcelExporter:
    """
    Servizio export Excel
    
    Genera:
    - File stock giornaliero per noleggiatore
    - Fogli: Estrazione, Riepilogo, Selezioni
    """
    
    # Stili
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    MATCHED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    NO_MATCH_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    def genera_excel(self, noleggiatore: str, data: date = None) -> Path:
        """
        Genera file Excel per noleggiatore
        
        Args:
            noleggiatore: Nome noleggiatore
            data: Data (default oggi)
            
        Returns:
            Path: Percorso file generato
        """
        if data is None:
            data = date.today()
        
        # Query veicoli
        veicoli = Veicolo.get_by_noleggiatore_data(noleggiatore, data).all()
        
        if not veicoli:
            raise ValueError(f"Nessun veicolo trovato per {noleggiatore} data {data}")
        
        # Crea workbook
        wb = Workbook()
        
        # Foglio Estrazione (tutti i dati)
        self._crea_foglio_estrazione(wb, veicoli)
        
        # Foglio Riepilogo (aggregato per modello)
        self._crea_foglio_riepilogo(wb, veicoli)
        
        # Foglio Selezioni (solo matched)
        self._crea_foglio_selezioni(wb, veicoli)
        
        # Salva file
        output_dir = Path(current_app.config['DIR_OUTPUT'])
        output_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"{noleggiatore.lower()}_stock_{data.strftime('%d-%m-%Y')}.xlsx"
        filepath = output_dir / filename
        
        wb.save(filepath)
        
        return filepath
    
    def _crea_foglio_estrazione(self, wb: Workbook, veicoli: List[Veicolo]):
        """Crea foglio Estrazione con tutti i dati"""
        ws = wb.active
        ws.title = "Estrazione"
        
        # Colonne
        columns = [
            'VIN', 'Marca', 'Modello', 'Description', 'Alimentazione',
            'KW', 'HP', 'CO2', 'Prezzo Listino', 'Prezzo Totale',
            'Location', 'Data Arrivo', 'Colore', 'Neopatentati',
            'Jato Code', 'Product ID', 'Omologazione',
            'Match Status', 'Match Score', 'Stato'
        ]
        
        # Header
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        
        # Dati
        for row_idx, veicolo in enumerate(veicoli, 2):
            data = veicolo.to_excel_row()
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=data.get(col_name))
                
                # Colorazione per match status
                if col_name == 'Match Status':
                    if data.get(col_name) == 'MATCHED':
                        cell.fill = self.MATCHED_FILL
                    elif data.get(col_name) == 'NO_MATCH':
                        cell.fill = self.NO_MATCH_FILL
        
        # Auto-width colonne
        self._auto_width(ws)
    
    def _crea_foglio_riepilogo(self, wb: Workbook, veicoli: List[Veicolo]):
        """Crea foglio Riepilogo aggregato per modello"""
        ws = wb.create_sheet("Riepilogo")
        
        # Aggrega per marca + modello
        aggregato = {}
        
        for v in veicoli:
            key = (v.marca, v.jato_product_description or v.description)
            if key not in aggregato:
                aggregato[key] = {
                    'marca': v.marca,
                    'modello': v.jato_product_description or v.description,
                    'alimentazione': v.alimentazione,
                    'kw': v.kw,
                    'neopatentati': v.neopatentati,
                    'count': 0,
                    'prezzo_min': float('inf'),
                    'prezzo_max': 0,
                }
            
            aggregato[key]['count'] += 1
            if v.prezzo_totale:
                aggregato[key]['prezzo_min'] = min(aggregato[key]['prezzo_min'], v.prezzo_totale)
                aggregato[key]['prezzo_max'] = max(aggregato[key]['prezzo_max'], v.prezzo_totale)
        
        # Colonne
        columns = ['Marca', 'Modello', 'Alimentazione', 'KW', 'Neopatentati', 
                   'Quantità', 'Prezzo Min', 'Prezzo Max']
        
        # Header
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # Dati
        for row_idx, data in enumerate(aggregato.values(), 2):
            ws.cell(row=row_idx, column=1, value=data['marca'])
            ws.cell(row=row_idx, column=2, value=data['modello'])
            ws.cell(row=row_idx, column=3, value=data['alimentazione'])
            ws.cell(row=row_idx, column=4, value=data['kw'])
            ws.cell(row=row_idx, column=5, value=data['neopatentati'])
            ws.cell(row=row_idx, column=6, value=data['count'])
            ws.cell(row=row_idx, column=7, value=data['prezzo_min'] if data['prezzo_min'] != float('inf') else None)
            ws.cell(row=row_idx, column=8, value=data['prezzo_max'] if data['prezzo_max'] > 0 else None)
        
        self._auto_width(ws)
    
    def _crea_foglio_selezioni(self, wb: Workbook, veicoli: List[Veicolo]):
        """Crea foglio Selezioni (solo veicoli matched)"""
        ws = wb.create_sheet("Selezioni")
        
        # Solo matched
        matched = [v for v in veicoli if v.match_status == 'MATCHED']
        
        columns = ['Marca', 'Modello', 'Alimentazione', 'KW', 'Neopatentati',
                   'Prezzo', 'Location', 'Jato Code']
        
        # Header
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # Dati
        for row_idx, v in enumerate(matched, 2):
            ws.cell(row=row_idx, column=1, value=v.marca)
            ws.cell(row=row_idx, column=2, value=v.jato_product_description or v.description)
            ws.cell(row=row_idx, column=3, value=v.alimentazione)
            ws.cell(row=row_idx, column=4, value=v.kw)
            ws.cell(row=row_idx, column=5, value=v.neopatentati)
            ws.cell(row=row_idx, column=6, value=v.prezzo_totale)
            ws.cell(row=row_idx, column=7, value=v.location)
            ws.cell(row=row_idx, column=8, value=v.jato_code)
        
        self._auto_width(ws)
    
    def _auto_width(self, ws):
        """Adatta larghezza colonne al contenuto"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
