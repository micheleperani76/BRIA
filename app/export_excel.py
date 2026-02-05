#!/usr/bin/env python3
# ==============================================================================
# GESTIONE FLOTTA - Modulo Export Excel
# ==============================================================================
# Versione: 2.0.0
# Data: 2026-01-30
# Descrizione: Sistema export Excel/CSV/Stampa configurabile con storico
#              - Tab Clienti: configuratore campi
#              - Tab Top Prospect: confermati con campi fissi
#              - Tab Trattative: filtri multipli
# ==============================================================================

import os
import csv
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import per visibilita gerarchica
try:
    from app.database_utenti import get_subordinati
except ImportError:
    def get_subordinati(conn, user_id):
        return [user_id]

# ==============================================================================
# CONFIGURAZIONE CAMPI EXPORT
# ==============================================================================

# Campi principali (uso frequente) - ordinati come richiesto
CAMPI_PRINCIPALI = [
    {'id': 'nome_cliente', 'label': 'Nome Azienda', 'tipo': 'text'},
    {'id': 'p_iva', 'label': 'P.IVA', 'tipo': 'text'},
    {'id': 'commerciale', 'label': 'Commerciale', 'tipo': 'text'},
    {'id': 'score', 'label': 'Score', 'tipo': 'text'},
    {'id': 'punteggio_rischio', 'label': 'Punteggio Rischio', 'tipo': 'number'},
    {'id': 'valore_produzione', 'label': 'Fatturato', 'tipo': 'euro'},
    {'id': 'capitale_sociale', 'label': 'Capitale Sociale', 'tipo': 'euro'},
    {'id': 'patrimonio_netto', 'label': 'Patrimonio Netto', 'tipo': 'euro'},
    {'id': 'utile', 'label': 'Utile/Perdita', 'tipo': 'euro'},
    {'id': 'dipendenti', 'label': 'Dipendenti', 'tipo': 'number'},
    {'id': 'num_veicoli', 'label': 'Flotta', 'tipo': 'number', 'calcolato': True},
    {'id': 'provincia', 'label': 'Provincia', 'tipo': 'text', 'calcolato': True},
    {'id': 'forma_giuridica', 'label': 'Forma Giuridica', 'tipo': 'text'},
]

# Campi secondari (ricerche elaborate)
CAMPI_SECONDARI = [
    {'id': 'credito', 'label': 'Fido Consigliato', 'tipo': 'euro'},
    {'id': 'indirizzo', 'label': 'Indirizzo', 'tipo': 'text'},
    {'id': 'telefono', 'label': 'Telefono', 'tipo': 'text'},
    {'id': 'pec', 'label': 'PEC', 'tipo': 'text'},
    {'id': 'stato', 'label': 'Stato Azienda', 'tipo': 'text'},
    {'id': 'cod_fiscale', 'label': 'Codice Fiscale', 'tipo': 'text'},
    {'id': 'numero_registrazione', 'label': 'N. Reg. Imprese', 'tipo': 'text'},
    {'id': 'ragione_sociale', 'label': 'Ragione Sociale', 'tipo': 'text'},
    {'id': 'data_costituzione', 'label': 'Data Costituzione', 'tipo': 'date'},
    {'id': 'codice_ateco', 'label': 'Codice ATECO', 'tipo': 'text'},
    {'id': 'desc_ateco', 'label': 'Descrizione ATECO', 'tipo': 'text'},
    {'id': 'desc_attivita', 'label': 'Attivita', 'tipo': 'text'},
    {'id': 'legale_rappresentante', 'label': 'Legale Rappresentante', 'tipo': 'text'},
    {'id': 'capogruppo_nome', 'label': 'Capogruppo', 'tipo': 'text'},
    {'id': 'protesti', 'label': 'Protesti', 'tipo': 'text'},
    {'id': 'importo_protesti', 'label': 'Importo Protesti', 'tipo': 'euro'},
    {'id': 'debiti', 'label': 'Debiti', 'tipo': 'euro'},
    {'id': 'anno_bilancio', 'label': 'Anno Bilancio', 'tipo': 'number'},
    {'id': 'valore_produzione_prec', 'label': 'Fatturato Anno Prec.', 'tipo': 'euro'},
    {'id': 'patrimonio_netto_prec', 'label': 'Patrimonio Netto Prec.', 'tipo': 'euro'},
    {'id': 'utile_prec', 'label': 'Utile/Perdita Prec.', 'tipo': 'euro'},
    {'id': 'debiti_prec', 'label': 'Debiti Prec.', 'tipo': 'euro'},
    {'id': 'data_import_creditsafe', 'label': 'Data Report CS', 'tipo': 'date'},
    {'id': 'canone_totale', 'label': 'Canone Totale Flotta', 'tipo': 'euro', 'calcolato': True},
]

# Dizionario completo per lookup rapido
TUTTI_CAMPI = {c['id']: c for c in CAMPI_PRINCIPALI + CAMPI_SECONDARI}

# Retention storico export (giorni)
EXPORT_RETENTION_DAYS = 90


# ==============================================================================
# FUNZIONI ESTRAZIONE DATI
# ==============================================================================

def estrai_provincia_da_indirizzo(indirizzo):
    """Estrae la sigla provincia da un indirizzo."""
    if not indirizzo:
        return None
    
    import re
    indirizzo_upper = str(indirizzo).upper()
    
    # Pattern: CAP + citta + provincia
    m = re.search(r'\b\d{5}\s+\w+\s+([A-Z]{2})\b', indirizzo_upper)
    if m:
        return m.group(1)
    
    # Pattern: 2 lettere alla fine
    m = re.search(r'\b([A-Z]{2})\s*$', indirizzo_upper)
    if m:
        return m.group(1)
    
    # Pattern: tra parentesi
    m = re.search(r'\(([A-Z]{2})\)', indirizzo_upper)
    if m:
        return m.group(1)
    
    return None


def get_dati_export(db_path, campi_selezionati, ordinamento=None):
    """
    Estrae i dati dal database per l'export.
    
    Args:
        db_path: Path al database SQLite
        campi_selezionati: Lista di dict {'id': campo_id, 'order': 'asc'/'desc'/None}
        ordinamento: Lista di tuple (campo, direzione) per ORDER BY
    
    Returns:
        Lista di dict con i dati
    """
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Determina quali campi sono dal DB e quali calcolati
    campi_db = []
    campi_calcolati = []
    
    for campo_sel in campi_selezionati:
        campo_id = campo_sel['id']
        campo_info = TUTTI_CAMPI.get(campo_id)
        if campo_info:
            if campo_info.get('calcolato'):
                campi_calcolati.append(campo_id)
            else:
                campi_db.append(campo_id)
    
    # Query base - sempre prendi id, indirizzo e provincia
    campi_query = list(set(campi_db + ['id', 'indirizzo', 'provincia']))
    
    # Costruisci ORDER BY
    order_clause = ""
    if ordinamento:
        order_parts = []
        for campo, direzione in ordinamento:
            if campo in campi_db:
                order_parts.append(f"{campo} {direzione.upper()}")
        if order_parts:
            order_clause = " ORDER BY " + ", ".join(order_parts)
    
    query = f"SELECT {', '.join(campi_query)} FROM clienti WHERE 1=1 {order_clause}"
    cursor.execute(query)
    rows = cursor.fetchall()
    
    # Pre-calcola conteggi veicoli se necessario
    veicoli_count = {}
    canoni_totali = {}
    if 'num_veicoli' in campi_calcolati or 'canone_totale' in campi_calcolati:
        cursor.execute("""
            SELECT c.id, COUNT(v.id) as num_veicoli, COALESCE(SUM(v.canone), 0) as canone_totale
            FROM clienti c
            LEFT JOIN veicoli v ON v.cliente_id = c.id
            GROUP BY c.id
        """)
        for row in cursor.fetchall():
            veicoli_count[row['id']] = row['num_veicoli']
            canoni_totali[row['id']] = row['canone_totale']
    
    conn.close()
    
    # Costruisci risultato
    risultati = []
    for row in rows:
        record = {}
        for campo_sel in campi_selezionati:
            campo_id = campo_sel['id']
            
            if campo_id == 'num_veicoli':
                record[campo_id] = veicoli_count.get(row['id'], 0)
            elif campo_id == 'canone_totale':
                record[campo_id] = canoni_totali.get(row['id'], 0)
            elif campo_id == 'provincia':
                # Usa provincia dal DB se presente, altrimenti calcola da indirizzo
                if row['provincia']:
                    record[campo_id] = row['provincia']
                else:
                    record[campo_id] = estrai_provincia_da_indirizzo(row['indirizzo'])
            elif campo_id in row.keys():
                record[campo_id] = row[campo_id]
            else:
                record[campo_id] = None
        
        risultati.append(record)
    
    return risultati


# ==============================================================================
# GENERAZIONE EXCEL
# ==============================================================================

def genera_excel(db_path, campi_selezionati, output_path, ordinamento=None):
    """
    Genera file Excel con i dati configurati.
    
    Args:
        db_path: Path al database
        campi_selezionati: Lista di dict {'id': campo_id, 'order': 'asc'/'desc'/None}
        output_path: Path file output
        ordinamento: Lista di tuple (campo, direzione)
    
    Returns:
        Path del file generato
    """
    # Estrai dati
    dati = get_dati_export(db_path, campi_selezionati, ordinamento)
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Clienti"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Intestazioni
    for col_idx, campo_sel in enumerate(campi_selezionati, 1):
        campo_info = TUTTI_CAMPI.get(campo_sel['id'], {})
        cell = ws.cell(row=1, column=col_idx)
        cell.value = campo_info.get('label', campo_sel['id'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Dati
    for row_idx, record in enumerate(dati, 2):
        for col_idx, campo_sel in enumerate(campi_selezionati, 1):
            campo_id = campo_sel['id']
            campo_info = TUTTI_CAMPI.get(campo_id, {})
            valore = record.get(campo_id)
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            # Formattazione per tipo
            if valore is None:
                cell.value = ""
            elif campo_info.get('tipo') == 'euro':
                cell.value = valore if valore else 0
                cell.number_format = '#,##0.00'
            elif campo_info.get('tipo') == 'number':
                cell.value = valore if valore else 0
                cell.number_format = '#,##0'
            else:
                cell.value = str(valore) if valore else ""
    
    # Applica filtro PRIMA del ridimensionamento
    if dati:
        last_col = get_column_letter(len(campi_selezionati))
        last_row = len(dati) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    
    # Ridimensiona colonne DOPO il filtro
    for col_idx, campo_sel in enumerate(campi_selezionati, 1):
        campo_info = TUTTI_CAMPI.get(campo_sel['id'], {})
        col_letter = get_column_letter(col_idx)
        
        # Calcola larghezza ottimale
        max_length = len(campo_info.get('label', campo_sel['id']))
        for row_idx in range(2, min(len(dati) + 2, 102)):  # Max 100 righe per calcolo
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Aggiungi spazio per icona filtro (+3)
        ws.column_dimensions[col_letter].width = min(max_length + 5, 50)
    
    # Blocca prima riga
    ws.freeze_panes = 'A2'
    
    # Salva
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    
    return output_path


# ==============================================================================
# GESTIONE STORICO EXPORT
# ==============================================================================

def get_storico_export(exports_dir):
    """
    Ritorna lista export salvati nello storico.
    
    Returns:
        Lista di dict con info file (nome, data, dimensione, path)
    """
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        return []
    
    files = []
    for f in sorted(exports_dir.glob("*.xlsx"), reverse=True):
        stat = f.stat()
        files.append({
            'nome': f.name,
            'path': str(f),
            'data': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
            'dimensione': f"{stat.st_size / 1024:.1f} KB",
            'timestamp': stat.st_mtime
        })
    
    return files


def pulisci_export_vecchi(exports_dir, retention_days=EXPORT_RETENTION_DAYS):
    """
    Rimuove export piu vecchi della retention.
    
    Returns:
        Numero file rimossi
    """
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        return 0
    
    data_limite = datetime.now() - timedelta(days=retention_days)
    rimossi = 0
    
    for f in exports_dir.glob("*.xlsx"):
        if datetime.fromtimestamp(f.stat().st_mtime) < data_limite:
            f.unlink()
            rimossi += 1
    
    return rimossi


def genera_nome_export():
    """Genera nome file export con timestamp."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"export_clienti_{timestamp}.xlsx"


# ==============================================================================
# SALVATAGGIO/CARICAMENTO CONFIGURAZIONI
# ==============================================================================

def salva_configurazione(config_path, nome, campi_selezionati):
    """
    Salva una configurazione export per riutilizzo futuro.
    
    Args:
        config_path: Path cartella configurazioni
        nome: Nome configurazione
        campi_selezionati: Lista campi
    """
    import json
    
    config_path = Path(config_path)
    config_path.mkdir(parents=True, exist_ok=True)
    
    config_file = config_path / f"{nome}.json"
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump({
            'nome': nome,
            'data_creazione': datetime.now().isoformat(),
            'campi': campi_selezionati
        }, f, indent=2, ensure_ascii=False)
    
    return config_file


def carica_configurazione(config_file):
    """Carica una configurazione salvata."""
    import json
    
    with open(config_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def lista_configurazioni(config_path):
    """Lista configurazioni salvate."""
    config_path = Path(config_path)
    if not config_path.exists():
        return []
    
    configs = []
    for f in sorted(config_path.glob("*.json")):
        try:
            config = carica_configurazione(f)
            configs.append({
                'nome': config.get('nome', f.stem),
                'file': str(f),
                'data': config.get('data_creazione', ''),
                'num_campi': len(config.get('campi', []))
            })
        except:
            pass
    
    return configs


# ==============================================================================
# CONFIGURAZIONE CAMPI TOP PROSPECT
# ==============================================================================

CAMPI_TOP_PROSPECT = [
    {'id': 'nome_cliente', 'label': 'Nome Azienda', 'tipo': 'text'},
    {'id': 'p_iva', 'label': 'P.IVA', 'tipo': 'text'},
    {'id': 'commerciale', 'label': 'Commerciale', 'tipo': 'text'},
    {'id': 'flotta', 'label': 'Flotta', 'tipo': 'number'},
    {'id': 'dipendenti', 'label': 'Dipendenti', 'tipo': 'number'},
    {'id': 'provincia', 'label': 'Provincia', 'tipo': 'text'},
    {'id': 'ultimo_appuntamento', 'label': 'Ultimo Appuntamento', 'tipo': 'date'},
    {'id': 'prossimo_appuntamento', 'label': 'Prossimo Appuntamento', 'tipo': 'date'},
    {'id': 'car_policy', 'label': 'Car Policy', 'tipo': 'text'},
    {'id': 'ultima_nota', 'label': 'Ultima Nota', 'tipo': 'text'},
    {'id': 'nota_fissata', 'label': 'Nota Fissata', 'tipo': 'text'},
]

CAMPI_TOP_PROSPECT_DICT = {c['id']: c for c in CAMPI_TOP_PROSPECT}


# ==============================================================================
# CONFIGURAZIONE CAMPI TRATTATIVE
# ==============================================================================

CAMPI_TRATTATIVE = [
    {'id': 'ragione_sociale', 'label': 'Cliente', 'tipo': 'text'},
    {'id': 'p_iva', 'label': 'P.IVA', 'tipo': 'text'},
    {'id': 'commerciale_nome', 'label': 'Commerciale', 'tipo': 'text'},
    {'id': 'noleggiatore', 'label': 'Noleggiatore', 'tipo': 'text'},
    {'id': 'marca', 'label': 'Marca', 'tipo': 'text'},
    {'id': 'descrizione_veicolo', 'label': 'Veicolo', 'tipo': 'text'},
    {'id': 'tipologia_veicolo', 'label': 'Tipologia', 'tipo': 'text'},
    {'id': 'tipo_trattativa', 'label': 'Tipo Trattativa', 'tipo': 'text'},
    {'id': 'num_pezzi', 'label': 'N. Pezzi', 'tipo': 'number'},
    {'id': 'stato', 'label': 'Stato', 'tipo': 'text'},
    {'id': 'data_inizio', 'label': 'Data Inizio', 'tipo': 'date'},
    {'id': 'data_chiusura', 'label': 'Data Chiusura', 'tipo': 'date'},
    {'id': 'provvigione', 'label': 'Provvigione', 'tipo': 'euro'},
    {'id': 'q_percentuale', 'label': 'Q%', 'tipo': 'number'},
    {'id': 'mesi', 'label': 'Mesi', 'tipo': 'number'},
    {'id': 'km_totali', 'label': 'Km Totali', 'tipo': 'number'},
    {'id': 'note', 'label': 'Note', 'tipo': 'text'},
]

CAMPI_TRATTATIVE_DICT = {c['id']: c for c in CAMPI_TRATTATIVE}


# ==============================================================================
# FUNZIONI ESTRAZIONE TOP PROSPECT
# ==============================================================================

def get_dati_top_prospect(db_path, user_id=None):
    """
    Estrae i dati dei Top Prospect confermati per l'export.
    Rispetta la visibilita gerarchica.
    
    Args:
        db_path: Path al database SQLite
        user_id: ID utente per filtro gerarchico (None = tutti)
    
    Returns:
        Lista di dict con i dati
    """
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Determina IDs visibili
    ids_visibili = None
    if user_id:
        ids_visibili = get_subordinati(conn, user_id)
    
    # Query principale
    query = '''
        SELECT tp.*, 
               c.nome_cliente, c.ragione_sociale, c.provincia, c.dipendenti,
               c.p_iva, c.cod_fiscale, c.veicoli_rilevati,
               c.commerciale_id,
               (SELECT COUNT(*) FROM veicoli v WHERE v.cliente_id = c.id) as num_veicoli,
               (SELECT data_appuntamento FROM top_prospect_appuntamenti 
                WHERE top_prospect_id = tp.id AND completato = 1 
                ORDER BY data_appuntamento DESC LIMIT 1) as ultimo_appuntamento,
               (SELECT data_appuntamento FROM top_prospect_appuntamenti 
                WHERE top_prospect_id = tp.id AND completato = 0 AND data_appuntamento >= date('now')
                ORDER BY data_appuntamento ASC LIMIT 1) as prossimo_appuntamento,
               (SELECT CASE WHEN COUNT(*) > 0 THEN 1 ELSE 0 END 
                FROM car_policy_meta WHERE cliente_id = c.id) as ha_car_policy,
               (SELECT testo FROM top_prospect_note 
                WHERE top_prospect_id = tp.id AND eliminato = 0 
                ORDER BY data_creazione DESC LIMIT 1) as ultima_nota,
               (SELECT testo FROM top_prospect_note 
                WHERE top_prospect_id = tp.id AND fissata = 1 AND eliminato = 0 
                LIMIT 1) as nota_fissata,
               u.nome || ' ' || u.cognome AS commerciale_nome_display
        FROM top_prospect tp
        JOIN clienti c ON tp.cliente_id = c.id
        LEFT JOIN utenti u ON c.commerciale_id = u.id
        WHERE tp.stato = 'confermato'
    '''
    
    params = []
    if ids_visibili:
        placeholders = ','.join('?' * len(ids_visibili))
        query += f' AND c.commerciale_id IN ({placeholders})'
        params.extend(ids_visibili)
    
    query += ' ORDER BY tp.priorita ASC, c.nome_cliente ASC'
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    # Costruisci risultato
    risultati = []
    for row in rows:
        veicoli_db = row['num_veicoli'] or 0
        veicoli_rilevati = row['veicoli_rilevati'] or 0
        flotta = max(veicoli_db, veicoli_rilevati)
        
        record = {
            'nome_cliente': row['nome_cliente'] or row['ragione_sociale'] or '',
            'p_iva': row['p_iva'] or '',
            'commerciale': row['commerciale_nome_display'] or '',
            'flotta': flotta,
            'dipendenti': row['dipendenti'] or 0,
            'provincia': row['provincia'] or '',
            'ultimo_appuntamento': row['ultimo_appuntamento'] or '',
            'prossimo_appuntamento': row['prossimo_appuntamento'] or '',
            'car_policy': 'Si' if row['ha_car_policy'] else 'No',
            'ultima_nota': (row['ultima_nota'] or '')[:200],  # Tronca a 200 caratteri
            'nota_fissata': (row['nota_fissata'] or '')[:200],
        }
        risultati.append(record)
    
    return risultati


# ==============================================================================
# FUNZIONI ESTRAZIONE TRATTATIVE
# ==============================================================================

def get_dati_trattative(db_path, user_id=None, filtri=None):
    """
    Estrae i dati delle trattative per l'export.
    Rispetta la visibilita gerarchica.
    
    Args:
        db_path: Path al database SQLite
        user_id: ID utente per filtro gerarchico (None = tutti)
        filtri: Dict con filtri opzionali:
            - tipo_trattativa: str
            - stato: str o lista
            - noleggiatore: str
            - data_da: str (YYYY-MM-DD)
            - data_a: str (YYYY-MM-DD)
            - solo_aperte: bool
            - solo_chiuse: bool
            - commerciale_id: int (specifico commerciale)
    
    Returns:
        Lista di dict con i dati
    """
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    filtri = filtri or {}
    
    # Determina IDs visibili
    ids_visibili = None
    if user_id:
        ids_visibili = get_subordinati(conn, user_id)
    
    # Query principale
    query = '''
        SELECT 
            t.*,
            c.ragione_sociale,
            c.p_iva,
            u.nome || ' ' || u.cognome AS commerciale_nome
        FROM trattative t
        LEFT JOIN clienti c ON t.cliente_id = c.id
        LEFT JOIN utenti u ON t.commerciale_id = u.id
        WHERE (t.cancellata IS NULL OR t.cancellata = 0)
    '''
    
    params = []
    
    # Filtro visibilita gerarchica
    if ids_visibili:
        placeholders = ','.join('?' * len(ids_visibili))
        query += f' AND t.commerciale_id IN ({placeholders})'
        params.extend(ids_visibili)
    
    # Filtro commerciale specifico
    if filtri.get('commerciale_id'):
        query += ' AND t.commerciale_id = ?'
        params.append(filtri['commerciale_id'])
    
    # Filtro tipo trattativa
    if filtri.get('tipo_trattativa'):
        query += ' AND t.tipo_trattativa = ?'
        params.append(filtri['tipo_trattativa'])
    
    # Filtro noleggiatore
    if filtri.get('noleggiatore'):
        query += ' AND t.noleggiatore = ?'
        params.append(filtri['noleggiatore'])
    
    # Filtro stato
    if filtri.get('stato'):
        if isinstance(filtri['stato'], list):
            placeholders = ','.join('?' * len(filtri['stato']))
            query += f' AND t.stato IN ({placeholders})'
            params.extend(filtri['stato'])
        else:
            query += ' AND t.stato = ?'
            params.append(filtri['stato'])
    
    # Filtro solo aperte (stati non chiusi)
    if filtri.get('solo_aperte'):
        # Importa stati chiusi
        try:
            from app.config_trattative import get_stati_chiusi
            stati_chiusi = get_stati_chiusi()
        except:
            stati_chiusi = ['Approvato', 'Bocciato', 'Perso', 'Approvato con riserve']
        
        if stati_chiusi:
            placeholders = ','.join('?' * len(stati_chiusi))
            query += f' AND t.stato NOT IN ({placeholders})'
            params.extend(stati_chiusi)
    
    # Filtro solo chiuse
    if filtri.get('solo_chiuse'):
        try:
            from app.config_trattative import get_stati_chiusi
            stati_chiusi = get_stati_chiusi()
        except:
            stati_chiusi = ['Approvato', 'Bocciato', 'Perso', 'Approvato con riserve']
        
        if stati_chiusi:
            placeholders = ','.join('?' * len(stati_chiusi))
            query += f' AND t.stato IN ({placeholders})'
            params.extend(stati_chiusi)
    
    # Filtro date
    if filtri.get('data_da'):
        query += ' AND t.data_inizio >= ?'
        params.append(filtri['data_da'])
    
    if filtri.get('data_a'):
        query += ' AND t.data_inizio <= ?'
        params.append(filtri['data_a'])
    
    query += ' ORDER BY t.data_inizio DESC'
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    # Costruisci risultato
    risultati = []
    for row in rows:
        record = {
            'ragione_sociale': row['ragione_sociale'] or '',
            'p_iva': row['p_iva'] or '',
            'commerciale_nome': row['commerciale_nome'] or '',
            'noleggiatore': row['noleggiatore'] or '',
            'marca': row['marca'] or '',
            'descrizione_veicolo': row['descrizione_veicolo'] or '',
            'tipologia_veicolo': row['tipologia_veicolo'] or '',
            'tipo_trattativa': row['tipo_trattativa'] or '',
            'num_pezzi': row['num_pezzi'] or 1,
            'stato': row['stato'] or '',
            'data_inizio': row['data_inizio'] or '',
            'data_chiusura': row['data_chiusura'] or '',
            'provvigione': row['provvigione'] or 0,
            'q_percentuale': row['q_percentuale'] or 0,
            'mesi': row['mesi'] or 0,
            'km_totali': row['km_totali'] or 0,
            'note': (row['note'] or '')[:200],
        }
        risultati.append(record)
    
    return risultati


# ==============================================================================
# GENERAZIONE EXCEL/CSV GENERICA
# ==============================================================================

def genera_excel_generico(dati, campi_config, output_path, titolo="Export"):
    """
    Genera file Excel da dati generici.
    
    Args:
        dati: Lista di dict con i dati
        campi_config: Dict {campo_id: {'label': ..., 'tipo': ...}}
        output_path: Path file output
        titolo: Titolo del foglio
    
    Returns:
        Path del file generato
    """
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = titolo[:31]  # Excel limita a 31 caratteri
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if not dati:
        ws.cell(row=1, column=1, value="Nessun dato da esportare")
        wb.save(str(output_path))
        return output_path
    
    # Ottieni campi dal primo record
    campi = list(dati[0].keys())
    
    # Intestazioni
    for col_idx, campo_id in enumerate(campi, 1):
        campo_info = campi_config.get(campo_id, {})
        cell = ws.cell(row=1, column=col_idx)
        cell.value = campo_info.get('label', campo_id)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Dati
    for row_idx, record in enumerate(dati, 2):
        for col_idx, campo_id in enumerate(campi, 1):
            campo_info = campi_config.get(campo_id, {})
            valore = record.get(campo_id)
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            # Formattazione per tipo
            if valore is None:
                cell.value = ""
            elif campo_info.get('tipo') == 'euro':
                cell.value = valore if valore else 0
                cell.number_format = '#,##0.00'
            elif campo_info.get('tipo') == 'number':
                cell.value = valore if valore else 0
                cell.number_format = '#,##0'
            else:
                cell.value = str(valore) if valore else ""
    
    # Applica filtro
    if dati:
        last_col = get_column_letter(len(campi))
        last_row = len(dati) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    
    # Ridimensiona colonne
    for col_idx, campo_id in enumerate(campi, 1):
        campo_info = campi_config.get(campo_id, {})
        col_letter = get_column_letter(col_idx)
        
        # Calcola larghezza ottimale
        max_length = len(campo_info.get('label', campo_id))
        for row_idx in range(2, min(len(dati) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 50))
        
        ws.column_dimensions[col_letter].width = min(max_length + 5, 50)
    
    # Blocca prima riga
    ws.freeze_panes = 'A2'
    
    # Salva
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    
    return output_path


def genera_csv_generico(dati, campi_config, output_path):
    """
    Genera file CSV da dati generici.
    
    Args:
        dati: Lista di dict con i dati
        campi_config: Dict {campo_id: {'label': ..., 'tipo': ...}}
        output_path: Path file output
    
    Returns:
        Path del file generato
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not dati:
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            f.write("Nessun dato da esportare\n")
        return output_path
    
    # Ottieni campi dal primo record
    campi = list(dati[0].keys())
    
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, delimiter=';')
        
        # Intestazioni
        headers = [campi_config.get(c, {}).get('label', c) for c in campi]
        writer.writerow(headers)
        
        # Dati
        for record in dati:
            row = []
            for campo_id in campi:
                valore = record.get(campo_id)
                if valore is None:
                    row.append('')
                else:
                    row.append(str(valore))
            writer.writerow(row)
    
    return output_path


# ==============================================================================
# FUNZIONI HELPER PER EXPORT TOP PROSPECT E TRATTATIVE
# ==============================================================================

def genera_export_top_prospect(db_path, output_path, user_id=None, formato='xlsx'):
    """
    Genera export Top Prospect confermati.
    
    Args:
        db_path: Path database
        output_path: Path file output
        user_id: ID utente per filtro gerarchico
        formato: 'xlsx' o 'csv'
    
    Returns:
        Path del file generato
    """
    dati = get_dati_top_prospect(db_path, user_id)
    
    if formato == 'csv':
        return genera_csv_generico(dati, CAMPI_TOP_PROSPECT_DICT, output_path)
    else:
        return genera_excel_generico(dati, CAMPI_TOP_PROSPECT_DICT, output_path, "Top Prospect")


def genera_export_trattative(db_path, output_path, user_id=None, filtri=None, formato='xlsx'):
    """
    Genera export Trattative.
    
    Args:
        db_path: Path database
        output_path: Path file output
        user_id: ID utente per filtro gerarchico
        filtri: Dict con filtri
        formato: 'xlsx' o 'csv'
    
    Returns:
        Path del file generato
    """
    dati = get_dati_trattative(db_path, user_id, filtri)
    
    if formato == 'csv':
        return genera_csv_generico(dati, CAMPI_TRATTATIVE_DICT, output_path)
    else:
        return genera_excel_generico(dati, CAMPI_TRATTATIVE_DICT, output_path, "Trattative")
