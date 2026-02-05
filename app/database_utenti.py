#!/usr/bin/env python3
# ==============================================================================
# GESTIONE FLOTTA - Modulo Database Utenti
# ==============================================================================
# Versione: 1.0.0
# Data: 2025-01-20
# Descrizione: Gestione utenti, permessi, supervisioni e autenticazione
# ==============================================================================

import sqlite3
import secrets
import string
from datetime import datetime
from pathlib import Path
from werkzeug.security import generate_password_hash, check_password_hash

# ==============================================================================
# IMPORT CONFIGURAZIONE
# ==============================================================================
try:
    from .config import DB_FILE, DB_DIR
except ImportError:
    # Fallback per esecuzione diretta
    from config import DB_FILE, DB_DIR


# ==============================================================================
# COSTANTI
# ==============================================================================

# Lunghezza password temporanea
PWD_TEMP_LENGTH = 12

# Tentativi massimi prima del blocco
MAX_TENTATIVI_LOGIN = 5

# Ruoli base disponibili
RUOLI_BASE = ['admin', 'commerciale', 'operatore', 'viewer']


# ==============================================================================
# INIZIALIZZAZIONE TABELLE UTENTI
# ==============================================================================

def init_tabelle_utenti(conn):
    """
    Crea le tabelle per il sistema utenti se non esistono.
    Da chiamare dopo init_database() principale.
    """
    cursor = conn.cursor()
    
    # =========================================================================
    # TABELLA UTENTI
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS utenti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            -- Identificativi
            codice_utente TEXT UNIQUE NOT NULL,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            
            -- Anagrafica (compilata dall'utente al primo accesso)
            nome TEXT,
            cognome TEXT,
            data_nascita TEXT,
            email TEXT,
            cellulare TEXT,
            
            -- Ruolo base
            ruolo_base TEXT DEFAULT 'operatore',
            
            -- Stato
            attivo INTEGER DEFAULT 1,
            pwd_temporanea INTEGER DEFAULT 1,
            profilo_completo INTEGER DEFAULT 0,
            bloccato INTEGER DEFAULT 0,
            tentativi_falliti INTEGER DEFAULT 0,
            
            -- Protezione (per utente admin di sistema)
            non_cancellabile INTEGER DEFAULT 0,
            non_modificabile INTEGER DEFAULT 0,
            
            -- Metadati
            data_creazione TEXT,
            data_ultimo_accesso TEXT,
            data_ultimo_cambio_pwd TEXT,
            creato_da INTEGER,
            
            -- Collegamento con flotta
            nome_commerciale_flotta TEXT,
            
            FOREIGN KEY (creato_da) REFERENCES utenti(id)
        )
    ''')
    
    # =========================================================================
    # TABELLA SUPERVISIONI
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS supervisioni (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            supervisore_id INTEGER NOT NULL,
            subordinato_id INTEGER NOT NULL,
            
            data_inizio TEXT NOT NULL,
            data_fine TEXT,
            
            FOREIGN KEY (supervisore_id) REFERENCES utenti(id),
            FOREIGN KEY (subordinato_id) REFERENCES utenti(id),
            UNIQUE(supervisore_id, subordinato_id, data_inizio)
        )
    ''')
    
    # =========================================================================
    # TABELLA CATALOGO PERMESSI
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS permessi_catalogo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            codice TEXT UNIQUE NOT NULL,
            descrizione TEXT NOT NULL,
            categoria TEXT NOT NULL,
            ordine INTEGER DEFAULT 0,
            
            attivo INTEGER DEFAULT 1
        )
    ''')
    
    # =========================================================================
    # TABELLA PERMESSI UTENTE
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS utenti_permessi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            utente_id INTEGER NOT NULL,
            permesso_id INTEGER NOT NULL,
            abilitato INTEGER DEFAULT 1,
            
            data_assegnazione TEXT,
            assegnato_da INTEGER,
            
            FOREIGN KEY (utente_id) REFERENCES utenti(id) ON DELETE CASCADE,
            FOREIGN KEY (permesso_id) REFERENCES permessi_catalogo(id),
            FOREIGN KEY (assegnato_da) REFERENCES utenti(id),
            UNIQUE(utente_id, permesso_id)
        )
    ''')
    
    # =========================================================================
    # TABELLA LOG ACCESSI
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS log_accessi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            utente_id INTEGER,
            username_tentativo TEXT,
            
            azione TEXT NOT NULL,
            ip_address TEXT,
            user_agent TEXT,
            
            dettaglio TEXT,
            data_ora TEXT NOT NULL,
            
            FOREIGN KEY (utente_id) REFERENCES utenti(id)
        )
    ''')
    
    # =========================================================================
    # TABELLA LOG ATTIVITA
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS log_attivita (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            utente_id INTEGER NOT NULL,
            
            azione TEXT NOT NULL,
            entita TEXT NOT NULL,
            entita_id INTEGER,
            
            dettaglio TEXT,
            ip_address TEXT,
            data_ora TEXT NOT NULL,
            
            FOREIGN KEY (utente_id) REFERENCES utenti(id)
        )
    ''')
    
    # =========================================================================
    # TABELLA STORICO ASSEGNAZIONI COMMERCIALI
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS storico_assegnazioni (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            -- Dati assegnazione
            cliente_nome TEXT NOT NULL,
            cliente_piva TEXT,
            commerciale_precedente TEXT,
            commerciale_nuovo TEXT NOT NULL,
            
            -- Chi ha fatto l'assegnazione
            utente_id INTEGER NOT NULL,
            
            -- Quando
            data_ora TEXT NOT NULL,
            
            -- Note opzionali
            note TEXT,
            
            FOREIGN KEY (utente_id) REFERENCES utenti(id)
        )
    ''')
    
    # =========================================================================
    # TABELLA STORICO EXPORT
    # =========================================================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS storico_export (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            -- Chi ha esportato
            utente_id INTEGER NOT NULL,
            
            -- Cosa ha esportato
            tipo_export TEXT NOT NULL,
            filtri_applicati TEXT,
            num_record INTEGER,
            nome_file TEXT,
            
            -- Quando
            data_ora TEXT NOT NULL,
            
            FOREIGN KEY (utente_id) REFERENCES utenti(id)
        )
    ''')
    
    # =========================================================================
    # INDICI
    # =========================================================================
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_utenti_username ON utenti(username)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_utenti_codice ON utenti(codice_utente)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_utenti_attivo ON utenti(attivo)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_supervisioni_sup ON supervisioni(supervisore_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_supervisioni_sub ON supervisioni(subordinato_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_supervisioni_attive ON supervisioni(data_fine)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_permessi_codice ON permessi_catalogo(codice)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_permessi_categoria ON permessi_catalogo(categoria)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_utenti_permessi_utente ON utenti_permessi(utente_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_log_accessi_utente ON log_accessi(utente_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_log_accessi_data ON log_accessi(data_ora)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_log_attivita_utente ON log_attivita(utente_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_log_attivita_data ON log_attivita(data_ora)')
    
    # Indici storico_assegnazioni
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_assegnazioni_cliente ON storico_assegnazioni(cliente_nome)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_assegnazioni_data ON storico_assegnazioni(data_ora)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_assegnazioni_utente ON storico_assegnazioni(utente_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_assegnazioni_commerciale ON storico_assegnazioni(commerciale_nuovo)')
    
    # Indici storico_export
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_export_utente ON storico_export(utente_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_export_data ON storico_export(data_ora)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_storico_export_tipo ON storico_export(tipo_export)')
    
    conn.commit()
    
    # Inserisce permessi catalogo se non esistono
    _inserisci_permessi_catalogo(conn)
    
    # Crea utente admin di sistema se non esiste
    _crea_admin_sistema(conn)
    
    return True


def _inserisci_permessi_catalogo(conn):
    """Inserisce i permessi nel catalogo se non esistono."""
    cursor = conn.cursor()
    
    permessi = [
        # CLIENTI
        ('clienti_visualizza', 'Visualizzare lista clienti', 'clienti', 10),
        ('clienti_modifica', 'Modificare anagrafica clienti', 'clienti', 20),
        ('clienti_note_visualizza', 'Visualizzare note clienti', 'clienti', 30),
        ('clienti_note_modifica', 'Aggiungere/modificare note', 'clienti', 40),
        
        # DOCUMENTI
        ('documenti_visualizza', 'Visualizzare documenti', 'documenti', 50),
        ('documenti_carica', 'Caricare nuovi documenti', 'documenti', 60),
        ('documenti_elimina', 'Eliminare documenti', 'documenti', 70),
        
        # VEICOLI
        ('veicoli_visualizza', 'Visualizzare veicoli', 'veicoli', 80),
        ('veicoli_modifica', 'Modificare dati veicoli', 'veicoli', 90),
        
        # STRUMENTI
        ('export_excel', 'Esportare dati in Excel', 'strumenti', 100),
        ('import_creditsafe', 'Importare PDF Creditsafe', 'strumenti', 110),
        
        # STATISTICHE
        ('statistiche_proprie', 'Vedere statistiche propri dati', 'statistiche', 120),
        ('statistiche_globali', 'Vedere statistiche globali', 'statistiche', 130),
        
        # FLOTTA
        ('flotta_assegnazioni', 'Gestire assegnazioni commerciali', 'flotta', 140),
        
        # ADMIN
        ('admin_utenti', 'Creare utenti e reset password', 'admin', 200),
        ('admin_permessi', 'Modificare permessi utenti', 'admin', 210),
        ('admin_sistema', 'Configurazioni di sistema', 'admin', 220),
    ]
    
    for codice, descrizione, categoria, ordine in permessi:
        cursor.execute('''
            INSERT OR IGNORE INTO permessi_catalogo (codice, descrizione, categoria, ordine)
            VALUES (?, ?, ?, ?)
        ''', (codice, descrizione, categoria, ordine))
    
    conn.commit()


def _crea_admin_sistema(conn):
    """Crea l'utente admin di sistema se non esiste."""
    cursor = conn.cursor()
    
    # Verifica se esiste gia
    cursor.execute("SELECT id FROM utenti WHERE username = 'admin'")
    if cursor.fetchone():
        return  # Gia esiste
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Genera password temporanea
    pwd_temp = genera_password_temporanea()
    pwd_hash = generate_password_hash(pwd_temp)
    
    cursor.execute('''
        INSERT INTO utenti (
            codice_utente, username, password_hash,
            nome, cognome, ruolo_base,
            attivo, pwd_temporanea, profilo_completo,
            non_cancellabile, non_modificabile,
            data_creazione
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        'ADM001', 'admin', pwd_hash,
        'Amministratore', 'Sistema', 'admin',
        1, 1, 1,  # attivo, pwd_temporanea, profilo_completo (admin non deve compilare)
        1, 1,     # non_cancellabile, non_modificabile
        now
    ))
    
    conn.commit()
    
    # Log della password temporanea (da rimuovere in produzione o salvare sicuro)
    print(f"")
    print(f"=" * 60)
    print(f"  UTENTE ADMIN CREATO")
    print(f"=" * 60)
    print(f"  Username: admin")
    print(f"  Password temporanea: {pwd_temp}")
    print(f"  IMPORTANTE: Cambiare la password al primo accesso!")
    print(f"=" * 60)
    print(f"")
    
    return pwd_temp


# ==============================================================================
# FUNZIONI UTILITÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ PASSWORD
# ==============================================================================

def genera_password_temporanea(lunghezza=PWD_TEMP_LENGTH):
    """Genera una password temporanea sicura."""
    # Caratteri: lettere maiuscole, minuscole, numeri (no simboli ambigui)
    caratteri = string.ascii_letters + string.digits
    # Rimuove caratteri ambigui (0, O, l, 1, I)
    caratteri = caratteri.replace('0', '').replace('O', '').replace('l', '').replace('1', '').replace('I', '')
    return ''.join(secrets.choice(caratteri) for _ in range(lunghezza))


def verifica_password(password_hash, password):
    """Verifica una password contro il suo hash."""
    return check_password_hash(password_hash, password)


def hash_password(password):
    """Genera l'hash di una password."""
    return generate_password_hash(password)


# ==============================================================================
# FUNZIONI GESTIONE UTENTI
# ==============================================================================

def get_connection():
    """Ottiene una connessione al database."""
    conn = sqlite3.connect(str(DB_FILE))
    conn.row_factory = sqlite3.Row
    return conn


def crea_utente(conn, username, creato_da_id=None):
    """
    Crea un nuovo utente con solo username e password temporanea.
    L'utente completera il profilo al primo accesso.
    
    Codici utente:
    - 000000-000004: Utenti iniziali (admin + commerciali)
    - 000005-999997: Nuovi utenti (assegnazione progressiva)
    - 999998-999999: Riservati per test
    
    Returns:
        dict: {'id': int, 'username': str, 'password_temporanea': str} o None se errore
    """
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Verifica username non esista
    cursor.execute("SELECT id FROM utenti WHERE username = ?", (username,))
    if cursor.fetchone():
        return None  # Username gia esistente
    
    # Genera codice utente progressivo
    # Cerca il MAX tra i codici validi (esclusi 999998 e 999999 riservati per test)
    cursor.execute("""
        SELECT MAX(CAST(codice_utente AS INTEGER)) 
        FROM utenti 
        WHERE codice_utente GLOB '[0-9][0-9][0-9][0-9][0-9][0-9]' 
        AND CAST(codice_utente AS INTEGER) < 999998
    """)
    result = cursor.fetchone()[0]
    
    # Se non ci sono utenti o il MAX e' < 4, partiamo da 5
    if result is None or result < 4:
        nuovo_numero = 5
    else:
        nuovo_numero = result + 1
    
    # Verifica limite massimo (999997)
    if nuovo_numero > 999997:
        print("ERRORE: Raggiunto limite massimo utenti (999997)")
        return None
    
    codice_utente = f"{nuovo_numero:06d}"
    
    # Genera password temporanea
    pwd_temp = genera_password_temporanea()
    pwd_hash = generate_password_hash(pwd_temp)
    
    cursor.execute('''
        INSERT INTO utenti (
            codice_utente, username, password_hash,
            ruolo_base, attivo, pwd_temporanea, profilo_completo,
            data_creazione, creato_da
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        codice_utente, username, pwd_hash,
        'operatore', 1, 1, 0,  # pwd_temporanea=1, profilo_completo=0
        now, creato_da_id
    ))
    
    conn.commit()
    
    return {
        'id': cursor.lastrowid,
        'codice_utente': codice_utente,
        'username': username,
        'password_temporanea': pwd_temp
    }


def verifica_unicita_contatti(conn, utente_id, email=None, cellulare=None):
    """
    Verifica che email e cellulare non siano gia usati da altri utenti.
    
    Args:
        conn: Connessione database
        utente_id: ID utente da escludere dal controllo (puo essere None per nuovi utenti)
        email: Email da verificare (opzionale)
        cellulare: Cellulare da verificare (opzionale)
    
    Returns:
        tuple: (valido, messaggio_errore)
               valido = True se i dati sono unici, False altrimenti
    """
    cursor = conn.cursor()
    
    # Verifica email
    if email:
        if utente_id:
            cursor.execute("SELECT id, username FROM utenti WHERE email = ? AND id != ?", (email, utente_id))
        else:
            cursor.execute("SELECT id, username FROM utenti WHERE email = ?", (email,))
        
        esistente = cursor.fetchone()
        if esistente:
            return (False, f"Email gia in uso dall'utente {esistente['username']}")
    
    # Verifica cellulare
    if cellulare:
        if utente_id:
            cursor.execute("SELECT id, username FROM utenti WHERE cellulare = ? AND id != ?", (cellulare, utente_id))
        else:
            cursor.execute("SELECT id, username FROM utenti WHERE cellulare = ?", (cellulare,))
        
        esistente = cursor.fetchone()
        if esistente:
            return (False, f"Numero di cellulare gia in uso dall'utente {esistente['username']}")
    
    return (True, None)


def completa_profilo_utente(conn, utente_id, nome, cognome, cellulare, email):
    """
    Completa il profilo utente con i dati obbligatori.
    Chiamata al primo accesso dopo il cambio password.
    
    Returns:
        tuple: (successo, messaggio_errore)
               successo = True se completato, False se errore
    """
    # Verifica unicita email e cellulare
    valido, errore = verifica_unicita_contatti(conn, utente_id, email, cellulare)
    if not valido:
        return (False, errore)
    
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        UPDATE utenti SET
            nome = ?,
            cognome = ?,
            cellulare = ?,
            email = ?,
            profilo_completo = 1
        WHERE id = ?
    ''', (nome, cognome, cellulare, email, utente_id))
    
    conn.commit()
    return (cursor.rowcount > 0, None)


def cambia_password(conn, utente_id, nuova_password):
    """
    Cambia la password di un utente.
    
    Returns:
        bool: True se cambiata, False se errore
    """
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    pwd_hash = generate_password_hash(nuova_password)
    
    cursor.execute('''
        UPDATE utenti SET
            password_hash = ?,
            pwd_temporanea = 0,
            data_ultimo_cambio_pwd = ?
        WHERE id = ?
    ''', (pwd_hash, now, utente_id))
    
    conn.commit()
    return cursor.rowcount > 0


def reset_password(conn, utente_id, reset_da_id=None):
    """
    Resetta la password di un utente generando una nuova temporanea.
    
    Returns:
        str: Nuova password temporanea o None se errore
    """
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    pwd_temp = genera_password_temporanea()
    pwd_hash = generate_password_hash(pwd_temp)
    
    cursor.execute('''
        UPDATE utenti SET
            password_hash = ?,
            pwd_temporanea = 1,
            tentativi_falliti = 0,
            bloccato = 0
        WHERE id = ? AND non_modificabile = 0
    ''', (pwd_hash, utente_id))
    
    if cursor.rowcount > 0:
        conn.commit()
        
        # Log dell'operazione
        log_accesso(conn, utente_id, 'reset_pwd', None, None, 
                   f'Reset da utente ID {reset_da_id}')
        
        return pwd_temp
    
    return None


def get_utente_by_username(conn, username):
    """Recupera un utente per username."""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM utenti WHERE username = ? AND attivo = 1", (username,))
    return cursor.fetchone()


def get_utente_by_id(conn, utente_id):
    """Recupera un utente per ID."""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM utenti WHERE id = ?", (utente_id,))
    return cursor.fetchone()


def get_tutti_utenti(conn, solo_attivi=True):
    """Recupera tutti gli utenti."""
    cursor = conn.cursor()
    if solo_attivi:
        cursor.execute("SELECT * FROM utenti WHERE attivo = 1 ORDER BY cognome, nome")
    else:
        cursor.execute("SELECT * FROM utenti ORDER BY cognome, nome")
    return cursor.fetchall()


def aggiorna_ultimo_accesso(conn, utente_id):
    """Aggiorna la data di ultimo accesso."""
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute("UPDATE utenti SET data_ultimo_accesso = ? WHERE id = ?", (now, utente_id))
    conn.commit()


def incrementa_tentativi_falliti(conn, username):
    """Incrementa i tentativi falliti e blocca se necessario."""
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE utenti SET 
            tentativi_falliti = tentativi_falliti + 1,
            bloccato = CASE WHEN tentativi_falliti + 1 >= ? THEN 1 ELSE 0 END
        WHERE username = ?
    ''', (MAX_TENTATIVI_LOGIN, username))
    
    conn.commit()


def resetta_tentativi_falliti(conn, utente_id):
    """Resetta i tentativi falliti dopo login riuscito."""
    cursor = conn.cursor()
    cursor.execute("UPDATE utenti SET tentativi_falliti = 0 WHERE id = ?", (utente_id,))
    conn.commit()


def sblocca_utente(conn, utente_id):
    """Sblocca un utente bloccato (solo admin)."""
    cursor = conn.cursor()
    cursor.execute("UPDATE utenti SET bloccato = 0, tentativi_falliti = 0 WHERE id = ?", (utente_id,))
    conn.commit()
    return cursor.rowcount > 0


# ==============================================================================
# FUNZIONI SUPERVISIONI
# ==============================================================================

def aggiungi_supervisione(conn, supervisore_id, subordinato_id):
    """Aggiunge una relazione di supervisione."""
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Verifica che non esista gia una supervisione attiva
    cursor.execute('''
        SELECT id FROM supervisioni 
        WHERE supervisore_id = ? AND subordinato_id = ? AND data_fine IS NULL
    ''', (supervisore_id, subordinato_id))
    
    if cursor.fetchone():
        return False  # Gia esiste
    
    cursor.execute('''
        INSERT INTO supervisioni (supervisore_id, subordinato_id, data_inizio)
        VALUES (?, ?, ?)
    ''', (supervisore_id, subordinato_id, now))
    
    conn.commit()
    return True


def rimuovi_supervisione(conn, supervisore_id, subordinato_id):
    """Termina una relazione di supervisione (non cancella, mette data_fine)."""
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        UPDATE supervisioni SET data_fine = ?
        WHERE supervisore_id = ? AND subordinato_id = ? AND data_fine IS NULL
    ''', (now, supervisore_id, subordinato_id))
    
    conn.commit()
    return cursor.rowcount > 0


def get_subordinati(conn, utente_id):
    """
    Restituisce tutti gli ID degli utenti visibili (se stesso + subordinati ricorsivi).
    Admin vede TUTTI gli utenti attivi.
    Usa CTE ricorsiva per la gerarchia a cascata.
    """
    cursor = conn.cursor()
    
    # Verifica se utente e' admin
    cursor.execute("SELECT ruolo_base FROM utenti WHERE id = ?", (utente_id,))
    row = cursor.fetchone()
    if row and row['ruolo_base'] == 'admin':
        # Admin vede tutti gli utenti attivi
        cursor.execute("SELECT id FROM utenti WHERE attivo = 1")
        return [r['id'] for r in cursor.fetchall()]
    
    # Altri utenti: gerarchia normale con CTE ricorsiva
    cursor.execute('''
        WITH RECURSIVE subordinati AS (
            SELECT ? as utente_id
            UNION ALL
            SELECT s.subordinato_id
            FROM supervisioni s
            INNER JOIN subordinati sub ON s.supervisore_id = sub.utente_id
            WHERE s.data_fine IS NULL
        )
        SELECT DISTINCT utente_id FROM subordinati
    ''', (utente_id,))
    
    return [row['utente_id'] for row in cursor.fetchall()]


def get_supervisori_diretti(conn, utente_id):
    """Restituisce i supervisori diretti di un utente."""
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.* FROM utenti u
        JOIN supervisioni s ON u.id = s.supervisore_id
        WHERE s.subordinato_id = ? AND s.data_fine IS NULL
    ''', (utente_id,))
    return cursor.fetchall()


def get_subordinati_diretti(conn, utente_id):
    """Restituisce i subordinati diretti di un utente."""
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.* FROM utenti u
        JOIN supervisioni s ON u.id = s.subordinato_id
        WHERE s.supervisore_id = ? AND s.data_fine IS NULL
    ''', (utente_id,))
    return cursor.fetchall()


# ==============================================================================
# FUNZIONI PERMESSI
# ==============================================================================

def get_permessi_catalogo(conn):
    """Restituisce tutti i permessi del catalogo."""
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM permessi_catalogo WHERE attivo = 1 ORDER BY ordine")
    return cursor.fetchall()


def get_permessi_per_categoria(conn):
    """Restituisce i permessi raggruppati per categoria."""
    cursor = conn.cursor()
    cursor.execute('''
        SELECT categoria, codice, descrizione, id
        FROM permessi_catalogo 
        WHERE attivo = 1 
        ORDER BY ordine
    ''')
    
    permessi_per_cat = {}
    for row in cursor.fetchall():
        cat = row['categoria']
        if cat not in permessi_per_cat:
            permessi_per_cat[cat] = []
        permessi_per_cat[cat].append(dict(row))
    
    return permessi_per_cat


def get_permessi_utente(conn, utente_id):
    """Restituisce i permessi assegnati a un utente."""
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pc.codice, pc.descrizione, pc.categoria, up.abilitato
        FROM permessi_catalogo pc
        LEFT JOIN utenti_permessi up ON pc.id = up.permesso_id AND up.utente_id = ?
        WHERE pc.attivo = 1
        ORDER BY pc.ordine
    ''', (utente_id,))
    return cursor.fetchall()


def ha_permesso(conn, utente_id, codice_permesso):
    """Verifica se un utente ha un permesso specifico."""
    cursor = conn.cursor()
    
    # Prima verifica se e admin (ha tutti i permessi)
    cursor.execute("SELECT ruolo_base FROM utenti WHERE id = ?", (utente_id,))
    utente = cursor.fetchone()
    if utente and utente['ruolo_base'] == 'admin':
        return True
    
    # Verifica permesso specifico
    cursor.execute('''
        SELECT up.abilitato 
        FROM utenti_permessi up
        JOIN permessi_catalogo pc ON up.permesso_id = pc.id
        WHERE up.utente_id = ? AND pc.codice = ?
    ''', (utente_id, codice_permesso))
    
    result = cursor.fetchone()
    return result and result['abilitato'] == 1


def assegna_permesso(conn, utente_id, codice_permesso, abilitato=True, assegnato_da_id=None):
    """Assegna o revoca un permesso a un utente."""
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Trova ID permesso
    cursor.execute("SELECT id FROM permessi_catalogo WHERE codice = ?", (codice_permesso,))
    perm = cursor.fetchone()
    if not perm:
        return False
    
    permesso_id = perm['id']
    
    # Inserisci o aggiorna
    cursor.execute('''
        INSERT INTO utenti_permessi (utente_id, permesso_id, abilitato, data_assegnazione, assegnato_da)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(utente_id, permesso_id) DO UPDATE SET
            abilitato = excluded.abilitato,
            data_assegnazione = excluded.data_assegnazione,
            assegnato_da = excluded.assegnato_da
    ''', (utente_id, permesso_id, 1 if abilitato else 0, now, assegnato_da_id))
    
    conn.commit()
    return True


def assegna_permessi_multipli(conn, utente_id, permessi_dict, assegnato_da_id=None):
    """
    Assegna piu permessi contemporaneamente.
    
    Args:
        permessi_dict: {'codice_permesso': True/False, ...}
    """
    for codice, abilitato in permessi_dict.items():
        assegna_permesso(conn, utente_id, codice, abilitato, assegnato_da_id)


def get_permessi_default_ruolo(ruolo):
    """
    Restituisce i permessi default per un ruolo.
    
    Args:
        ruolo: 'admin', 'commerciale', 'operatore', 'viewer'
    
    Returns:
        dict: {'codice_permesso': True/False, ...}
    """
    # Permessi base per tutti
    permessi_base = {
        'clienti_visualizza': True,
        'clienti_note_visualizza': True,
        'documenti_visualizza': True,
        'veicoli_visualizza': True,
        'statistiche_proprie': True,
    }
    
    if ruolo == 'admin':
        # Admin ha TUTTI i permessi
        return {
            'clienti_visualizza': True,
            'clienti_modifica': True,
            'clienti_note_visualizza': True,
            'clienti_note_modifica': True,
            'documenti_visualizza': True,
            'documenti_carica': True,
            'documenti_elimina': True,
            'veicoli_visualizza': True,
            'veicoli_modifica': True,
            'export_excel': True,
            'import_creditsafe': True,
            'statistiche_proprie': True,
            'statistiche_globali': True,
            'admin_utenti': True,
            'admin_permessi': True,
            'admin_sistema': True,
        }
    
    elif ruolo == 'commerciale':
        # Commerciale: tutto tranne admin
        return {
            'clienti_visualizza': True,
            'clienti_modifica': True,
            'clienti_note_visualizza': True,
            'clienti_note_modifica': True,
            'documenti_visualizza': True,
            'documenti_carica': True,
            'documenti_elimina': True,
            'veicoli_visualizza': True,
            'veicoli_modifica': True,
            'export_excel': True,
            'import_creditsafe': True,
            'statistiche_proprie': True,
            'statistiche_globali': True,
            'admin_utenti': False,
            'admin_permessi': False,
            'admin_sistema': False,
        }
    
    elif ruolo == 'operatore':
        # Operatore: visualizza + modifica base, no admin
        return {
            'clienti_visualizza': True,
            'clienti_modifica': True,
            'clienti_note_visualizza': True,
            'clienti_note_modifica': True,
            'documenti_visualizza': True,
            'documenti_carica': True,
            'documenti_elimina': False,
            'veicoli_visualizza': True,
            'veicoli_modifica': False,
            'export_excel': True,
            'import_creditsafe': False,
            'statistiche_proprie': True,
            'statistiche_globali': False,
            'admin_utenti': False,
            'admin_permessi': False,
            'admin_sistema': False,
        }
    
    else:  # viewer
        # Viewer: solo visualizzazione
        return {
            'clienti_visualizza': True,
            'clienti_modifica': False,
            'clienti_note_visualizza': True,
            'clienti_note_modifica': False,
            'documenti_visualizza': True,
            'documenti_carica': False,
            'documenti_elimina': False,
            'veicoli_visualizza': True,
            'veicoli_modifica': False,
            'export_excel': False,
            'import_creditsafe': False,
            'statistiche_proprie': True,
            'statistiche_globali': False,
            'admin_utenti': False,
            'admin_permessi': False,
            'admin_sistema': False,
        }


def assegna_permessi_default_ruolo(conn, utente_id, ruolo, assegnato_da_id=None):
    """
    Assegna i permessi default in base al ruolo.
    
    Args:
        conn: Connessione database
        utente_id: ID utente
        ruolo: 'admin', 'commerciale', 'operatore', 'viewer'
        assegnato_da_id: ID di chi assegna i permessi
    
    Returns:
        int: Numero permessi assegnati
    """
    permessi = get_permessi_default_ruolo(ruolo)
    assegna_permessi_multipli(conn, utente_id, permessi, assegnato_da_id)
    return len(permessi)


# ==============================================================================
# FUNZIONI LOG
# ==============================================================================

def log_accesso(conn, utente_id, azione, ip_address=None, user_agent=None, dettaglio=None, username_tentativo=None):
    """Registra un evento di accesso."""
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        INSERT INTO log_accessi (utente_id, username_tentativo, azione, ip_address, user_agent, dettaglio, data_ora)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (utente_id, username_tentativo, azione, ip_address, user_agent, dettaglio, now))
    
    conn.commit()


def log_attivita(conn, utente_id, azione, entita, entita_id=None, dettaglio=None, ip_address=None):
    """Registra un'attivita utente."""
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        INSERT INTO log_attivita (utente_id, azione, entita, entita_id, dettaglio, ip_address, data_ora)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (utente_id, azione, entita, entita_id, dettaglio, ip_address, now))
    
    conn.commit()


# ==============================================================================
# FUNZIONI VALIDAZIONE EMAIL
# ==============================================================================

def valida_email_dominio(email, config_path=None):
    """
    Valida l'email secondo la configurazione del dominio.
    
    Returns:
        tuple: (valida: bool, messaggio: str)
    """
    import configparser
    from pathlib import Path
    
    if not email or '@' not in email:
        return False, "Email non valida"
    
    # Carica configurazione
    if config_path is None:
        # Percorso default
        base_path = Path(__file__).parent.parent
        config_path = base_path / 'impostazioni' / 'email_config.conf'
    
    if not Path(config_path).exists():
        # Se non esiste config, accetta qualsiasi email
        return True, "OK"
    
    config = configparser.ConfigParser()
    config.read(config_path)
    
    validazione_attiva = config.getboolean('validazione', 'validazione_dominio_attiva', fallback=False)
    
    if not validazione_attiva:
        return True, "OK"
    
    dominio_consentito = config.get('validazione', 'dominio_consentito', fallback='')
    messaggio_errore = config.get('validazione', 'messaggio_errore', 
                                   fallback=f"L'email deve essere del dominio @{dominio_consentito}")
    
    # Estrai dominio dall'email
    dominio_email = email.split('@')[1].lower()
    
    if dominio_email == dominio_consentito.lower():
        return True, "OK"
    else:
        return False, messaggio_errore


# ==============================================================================
# MIGRAZIONE DATI ESISTENTI
# ==============================================================================

def migra_commerciali_esistenti(conn):
    """
    Migra i commerciali esistenti (da stringa a utenti).
    Crea utenti per PELUCCHI, PERANI, ZUBANI e aggiorna i riferimenti.
    
    ATTENZIONE: Eseguire una sola volta!
    """
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Mapping commerciali esistenti
    commerciali = [
        ('COM001', 'p.ciotti', 'Paolo', 'Ciotti', 'commerciale'),
        ('COM002', 'm.perani', 'Michele', 'Perani', 'commerciale'),
        ('COM003', 'f.zubani', 'Fausto', 'Zubani', 'commerciale'),
        ('COM004', 'c.pelucchi', 'Cristian', 'Pelucchi', 'commerciale'),
    ]
    
    utenti_creati = {}
    
    for codice, username, nome, cognome, ruolo in commerciali:
        # Verifica se esiste gia
        cursor.execute("SELECT id FROM utenti WHERE username = ?", (username,))
        esistente = cursor.fetchone()
        
        if esistente:
            utenti_creati[cognome.upper()] = esistente['id']
            print(f"  Utente {username} gia esistente (ID: {esistente['id']})")
            continue
        
        # Genera password temporanea
        pwd_temp = genera_password_temporanea()
        pwd_hash = generate_password_hash(pwd_temp)
        
        cursor.execute('''
            INSERT INTO utenti (
                codice_utente, username, password_hash,
                nome, cognome, ruolo_base,
                attivo, pwd_temporanea, profilo_completo,
                data_creazione
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            codice, username, pwd_hash,
            nome, cognome, ruolo,
            1, 1, 0,  # pwd_temporanea=1, profilo_completo=0
            now
        ))
        
        utente_id = cursor.lastrowid
        utenti_creati[cognome.upper()] = utente_id
        
        print(f"  Creato utente: {username} (ID: {utente_id}) - Password: {pwd_temp}")
    
    conn.commit()
    
    # Crea supervisioni (Paolo supervisiona tutti)
    paolo_id = utenti_creati.get('CIOTTI')
    if paolo_id:
        for cognome, utente_id in utenti_creati.items():
            if cognome != 'CIOTTI' and utente_id != paolo_id:
                cursor.execute('''
                    INSERT OR IGNORE INTO supervisioni (supervisore_id, subordinato_id, data_inizio)
                    VALUES (?, ?, ?)
                ''', (paolo_id, utente_id, now))
                print(f"  Supervisione: Paolo Ciotti &euro; {cognome}")
    
    conn.commit()
    
    # Aggiorna tabelle clienti e veicoli (aggiungi colonna se non esiste)
    try:
        cursor.execute("ALTER TABLE clienti ADD COLUMN commerciale_id INTEGER REFERENCES utenti(id)")
        print("  Aggiunta colonna commerciale_id a clienti")
    except sqlite3.OperationalError:
        print("  Colonna commerciale_id gia presente in clienti")
    
    try:
        cursor.execute("ALTER TABLE veicoli ADD COLUMN commerciale_id INTEGER REFERENCES utenti(id)")
        print("  Aggiunta colonna commerciale_id a veicoli")
    except sqlite3.OperationalError:
        print("  Colonna commerciale_id gia presente in veicoli")
    
    # Aggiorna riferimenti
    mapping_nomi = {
        'PERANI': utenti_creati.get('PERANI'),
        'ZUBANI': utenti_creati.get('ZUBANI'),
        'PELUCCHI': utenti_creati.get('PELUCCHI'),
    }
    
    for nome_commerciale, utente_id in mapping_nomi.items():
        if utente_id:
            cursor.execute("UPDATE clienti SET commerciale_id = ? WHERE UPPER(commerciale) = ?", 
                          (utente_id, nome_commerciale))
            n_clienti = cursor.rowcount
            
            cursor.execute("UPDATE veicoli SET commerciale_id = ? WHERE UPPER(commerciale) = ?", 
                          (utente_id, nome_commerciale))
            n_veicoli = cursor.rowcount
            
            print(f"  Migrato {nome_commerciale}: {n_clienti} clienti, {n_veicoli} veicoli")
    
    conn.commit()
    
    return utenti_creati


# ==============================================================================
# MAIN (per test)
# ==============================================================================

if __name__ == '__main__':
    print("Test modulo database_utenti...")
    
    conn = get_connection()
    
    print("\n1. Inizializzazione tabelle...")
    init_tabelle_utenti(conn)
    
    print("\n2. Verifica permessi catalogo...")
    permessi = get_permessi_catalogo(conn)
    print(f"   Trovati {len(permessi)} permessi nel catalogo")
    
    print("\n3. Test creazione utente...")
    risultato = crea_utente(conn, 'test.user')
    if risultato:
        print(f"   Creato: {risultato['username']} - PWD: {risultato['password_temporanea']}")
    
    conn.close()
    print("\nTest completato!")


def carica_permessi_utente(conn, utente_id):
    """
    Carica i codici dei permessi abilitati per un utente.
    Da usare dopo il login per popolare la sessione.
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pc.codice
        FROM utenti_permessi up
        JOIN permessi_catalogo pc ON up.permesso_id = pc.id
        WHERE up.utente_id = ? AND up.abilitato = 1
    ''', (utente_id,))
    return [row['codice'] for row in cursor.fetchall()]
# ==============================================================================
# FUNZIONI DA AGGIUNGERE A database_utenti.py
# ==============================================================================
# Copia questo contenuto alla fine del file app/database_utenti.py
# ==============================================================================


def get_tutti_utenti(conn, solo_attivi=True):
    """
    Restituisce tutti gli utenti.
    
    Args:
        conn: Connessione database
        solo_attivi: Se True, restituisce solo utenti attivi
    
    Returns:
        list: Lista di utenti (dict)
    """
    cursor = conn.cursor()
    if solo_attivi:
        cursor.execute('''
            SELECT * FROM utenti WHERE attivo = 1 ORDER BY cognome, nome
        ''')
    else:
        cursor.execute('''
            SELECT * FROM utenti ORDER BY cognome, nome
        ''')
    return [dict(row) for row in cursor.fetchall()]


def get_utente_by_id(conn, utente_id):
    """
    Restituisce un utente per ID.
    
    Returns:
        dict o None
    """
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM utenti WHERE id = ?', (utente_id,))
    row = cursor.fetchone()
    return dict(row) if row else None


def sblocca_utente(conn, utente_id):
    """
    Sblocca un utente bloccato.
    
    Returns:
        bool: True se sbloccato con successo
    """
    try:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE utenti SET bloccato = 0, tentativi_falliti = 0 WHERE id = ?
        ''', (utente_id,))
        conn.commit()
        return True
    except Exception as e:
        print(f"Errore sblocco utente: {e}")
        return False


def get_permessi_per_categoria(conn):
    """
    Restituisce i permessi raggruppati per categoria.
    
    Returns:
        dict: {categoria: [permessi]}
    """
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM permessi_catalogo ORDER BY categoria, codice')
    
    risultato = {}
    for row in cursor.fetchall():
        perm = dict(row)
        cat = perm['categoria']
        if cat not in risultato:
            risultato[cat] = []
        risultato[cat].append(perm)
    
    return risultato


def get_permessi_utente(conn, utente_id):
    """
    Restituisce i permessi di un utente con stato abilitazione.
    
    Returns:
        list: Lista di permessi con campo 'abilitato'
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pc.*, COALESCE(up.abilitato, 0) as abilitato
        FROM permessi_catalogo pc
        LEFT JOIN utenti_permessi up ON pc.id = up.permesso_id AND up.utente_id = ?
        ORDER BY pc.categoria, pc.codice
    ''', (utente_id,))
    return [dict(row) for row in cursor.fetchall()]


def assegna_permesso(conn, utente_id, codice_permesso, abilitato, assegnato_da=None):
    """
    Assegna o rimuove un permesso a un utente.
    
    Args:
        utente_id: ID utente
        codice_permesso: Codice del permesso
        abilitato: True per abilitare, False per disabilitare
        assegnato_da: ID utente che assegna
    
    Returns:
        bool: True se operazione riuscita
    """
    try:
        cursor = conn.cursor()
        
        # Trova ID permesso
        cursor.execute('SELECT id FROM permessi_catalogo WHERE codice = ?', (codice_permesso,))
        row = cursor.fetchone()
        if not row:
            return False
        permesso_id = row['id']
        
        # Inserisci o aggiorna
        cursor.execute('''
            INSERT INTO utenti_permessi (utente_id, permesso_id, abilitato, assegnato_da, data_assegnazione)
            VALUES (?, ?, ?, ?, datetime('now'))
            ON CONFLICT(utente_id, permesso_id) DO UPDATE SET 
                abilitato = excluded.abilitato,
                assegnato_da = excluded.assegnato_da,
                data_assegnazione = excluded.data_assegnazione
        ''', (utente_id, permesso_id, 1 if abilitato else 0, assegnato_da))
        
        conn.commit()
        return True
    except Exception as e:
        print(f"Errore assegnazione permesso: {e}")
        return False


def get_supervisori_diretti(conn, utente_id):
    """
    Restituisce i supervisori diretti di un utente.
    
    Returns:
        list: Lista di utenti supervisori
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.* FROM utenti u
        JOIN supervisioni s ON u.id = s.supervisore_id
        WHERE s.subordinato_id = ? AND s.data_fine IS NULL
        ORDER BY u.cognome, u.nome
    ''', (utente_id,))
    return [dict(row) for row in cursor.fetchall()]


def get_subordinati_diretti(conn, utente_id):
    """
    Restituisce i subordinati diretti di un utente.
    
    Returns:
        list: Lista di utenti subordinati
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.* FROM utenti u
        JOIN supervisioni s ON u.id = s.subordinato_id
        WHERE s.supervisore_id = ? AND s.data_fine IS NULL
        ORDER BY u.cognome, u.nome
    ''', (utente_id,))
    return [dict(row) for row in cursor.fetchall()]


def aggiungi_supervisione(conn, supervisore_id, subordinato_id):
    """
    Aggiunge una supervisione tra due utenti.
    
    Returns:
        bool: True se aggiunta con successo
    """
    try:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR IGNORE INTO supervisioni (supervisore_id, subordinato_id, data_inizio)
            VALUES (?, ?, datetime('now'))
        ''', (supervisore_id, subordinato_id))
        conn.commit()
        return True
    except Exception as e:
        print(f"Errore aggiunta supervisione: {e}")
        return False


def rimuovi_supervisione(conn, supervisore_id, subordinato_id):
    """
    Rimuove una supervisione (imposta data_fine).
    
    Returns:
        bool: True se rimossa con successo
    """
    try:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE supervisioni SET data_fine = datetime('now')
            WHERE supervisore_id = ? AND subordinato_id = ? AND data_fine IS NULL
        ''', (supervisore_id, subordinato_id))
        conn.commit()
        return True
    except Exception as e:
        print(f"Errore rimozione supervisione: {e}")
        return False


# ==============================================================================
# FUNZIONI LOG RETENTION
# ==============================================================================

import configparser
from pathlib import Path


def get_config_retention():
    """
    Legge la configurazione retention dal file.
    
    Returns:
        dict: Configurazione con valori default se file non esiste
    """
    config_default = {
        'log_accessi_giorni': 90,
        'log_attivita_giorni': 180,
        'pulizia_automatica': True,
        'righe_per_pagina': 50,
        'mostra_ip': True,
        'mostra_user_agent': False
    }
    
    try:
        base_path = Path(__file__).parent.parent
        config_path = base_path / 'impostazioni' / 'log_config.conf'
        
        if not config_path.exists():
            return config_default
        
        config = configparser.ConfigParser()
        config.read(config_path)
        
        return {
            'log_accessi_giorni': config.getint('retention', 'log_accessi_giorni', fallback=90),
            'log_attivita_giorni': config.getint('retention', 'log_attivita_giorni', fallback=180),
            'pulizia_automatica': config.getboolean('retention', 'pulizia_automatica', fallback=True),
            'righe_per_pagina': config.getint('visualizzazione', 'righe_per_pagina', fallback=50),
            'mostra_ip': config.getboolean('visualizzazione', 'mostra_ip', fallback=True),
            'mostra_user_agent': config.getboolean('visualizzazione', 'mostra_user_agent', fallback=False)
        }
    except Exception as e:
        print(f"Errore lettura config retention: {e}")
        return config_default


def get_log_accessi_utente(conn, utente_id, limite=None):
    """
    Restituisce i log accessi di un utente rispettando la retention.
    
    Args:
        conn: Connessione database
        utente_id: ID utente (None per tutti)
        limite: Limite righe (None usa config)
    
    Returns:
        list: Lista log accessi
    """
    config = get_config_retention()
    giorni = config['log_accessi_giorni']
    limite = limite or config['righe_per_pagina']
    
    cursor = conn.cursor()
    
    if utente_id:
        cursor.execute('''
            SELECT la.*, u.username, u.nome, u.cognome
            FROM log_accessi la
            LEFT JOIN utenti u ON la.utente_id = u.id
            WHERE la.utente_id = ?
            AND la.data_ora >= datetime('now', ?)
            ORDER BY la.data_ora DESC
            LIMIT ?
        ''', (utente_id, f'-{giorni} days', limite))
    else:
        cursor.execute('''
            SELECT la.*, u.username, u.nome, u.cognome
            FROM log_accessi la
            LEFT JOIN utenti u ON la.utente_id = u.id
            WHERE la.data_ora >= datetime('now', ?)
            ORDER BY la.data_ora DESC
            LIMIT ?
        ''', (f'-{giorni} days', limite))
    
    return [dict(row) for row in cursor.fetchall()]


def get_log_accessi_tutti(conn, pagina=1, filtro_tipo=None):
    """
    Restituisce tutti i log accessi con paginazione.
    
    Args:
        conn: Connessione database
        pagina: Numero pagina (1-based)
        filtro_tipo: Filtra per tipo evento (es: 'login_ok')
    
    Returns:
        dict: {'log': lista, 'totale': count, 'pagine': num_pagine}
    """
    config = get_config_retention()
    giorni = config['log_accessi_giorni']
    per_pagina = config['righe_per_pagina']
    offset = (pagina - 1) * per_pagina
    
    cursor = conn.cursor()
    
    # Conta totale
    if filtro_tipo:
        cursor.execute('''
            SELECT COUNT(*) FROM log_accessi
            WHERE data_ora >= datetime('now', ?)
            AND azione = ?
        ''', (f'-{giorni} days', filtro_tipo))
    else:
        cursor.execute('''
            SELECT COUNT(*) FROM log_accessi
            WHERE data_ora >= datetime('now', ?)
        ''', (f'-{giorni} days',))
    
    totale = cursor.fetchone()[0]
    pagine = (totale + per_pagina - 1) // per_pagina
    
    # Ottieni log
    if filtro_tipo:
        cursor.execute('''
            SELECT la.*, u.username, u.nome, u.cognome
            FROM log_accessi la
            LEFT JOIN utenti u ON la.utente_id = u.id
            WHERE la.data_ora >= datetime('now', ?)
            AND la.azione = ?
            ORDER BY la.data_ora DESC
            LIMIT ? OFFSET ?
        ''', (f'-{giorni} days', filtro_tipo, per_pagina, offset))
    else:
        cursor.execute('''
            SELECT la.*, u.username, u.nome, u.cognome
            FROM log_accessi la
            LEFT JOIN utenti u ON la.utente_id = u.id
            WHERE la.data_ora >= datetime('now', ?)
            ORDER BY la.data_ora DESC
            LIMIT ? OFFSET ?
        ''', (f'-{giorni} days', per_pagina, offset))
    
    return {
        'log': [dict(row) for row in cursor.fetchall()],
        'totale': totale,
        'pagine': pagine,
        'pagina_corrente': pagina,
        'per_pagina': per_pagina
    }


def pulisci_log_vecchi(conn):
    """
    Elimina i log piu vecchi della retention configurata.
    
    Returns:
        dict: {'accessi_eliminati': n, 'attivita_eliminate': n}
    """
    config = get_config_retention()
    
    if not config['pulizia_automatica']:
        return {'accessi_eliminati': 0, 'attivita_eliminate': 0}
    
    cursor = conn.cursor()
    
    # Pulisci log_accessi
    cursor.execute('''
        DELETE FROM log_accessi
        WHERE data_ora < datetime('now', ?)
    ''', (f'-{config["log_accessi_giorni"]} days',))
    accessi = cursor.rowcount
    
    # Pulisci log_attivita
    cursor.execute('''
        DELETE FROM log_attivita
        WHERE data_ora < datetime('now', ?)
    ''', (f'-{config["log_attivita_giorni"]} days',))
    attivita = cursor.rowcount
    
    conn.commit()
    
    return {
        'accessi_eliminati': accessi,
        'attivita_eliminate': attivita
    }


def get_statistiche_accessi_utente(conn, utente_id):
    """
    Restituisce statistiche accessi per un utente.
    
    Returns:
        dict: Statistiche varie
    """
    config = get_config_retention()
    giorni = config['log_accessi_giorni']
    
    cursor = conn.cursor()
    
    # Conteggio per tipo
    cursor.execute('''
        SELECT azione, COUNT(*) as conteggio
        FROM log_accessi
        WHERE utente_id = ?
        AND data_ora >= datetime('now', ?)
        GROUP BY azione
    ''', (utente_id, f'-{giorni} days'))
    
    per_tipo = {row['azione']: row['conteggio'] for row in cursor.fetchall()}
    
    # Ultimo accesso
    cursor.execute('''
        SELECT data_ora, ip_address
        FROM log_accessi
        WHERE utente_id = ? AND azione = 'login_ok'
        ORDER BY data_ora DESC LIMIT 1
    ''', (utente_id,))
    ultimo = cursor.fetchone()
    
    return {
        'login_ok': per_tipo.get('login_ok', 0),
        'login_fallito': per_tipo.get('login_fallito', 0),
        'logout': per_tipo.get('logout', 0),
        'cambio_pwd': per_tipo.get('cambio_pwd', 0),
        'ultimo_accesso': dict(ultimo) if ultimo else None,
        'retention_giorni': giorni
    }
# ==============================================================================
# FUNZIONI MAPPATURA IP - Da aggiungere a database_utenti.py
# ==============================================================================

def get_mappatura_ip():
    """
    Legge la mappatura IP dal file Excel.
    
    Returns:
        list: Lista di dict {'prefisso': '192.168.1.', 'nome': 'Rete Locale', 'note': '...'}
    """
    try:
        from openpyxl import load_workbook
        from pathlib import Path
        
        base_path = Path(__file__).parent.parent
        file_path = base_path / 'impostazioni' / 'mappatura_ip.xlsx'
        
        if not file_path.exists():
            return []
        
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        mappatura = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Salta header
            if row[0]:  # Se c'e un IP/prefisso
                mappatura.append({
                    'prefisso': str(row[0]).strip(),
                    'nome': str(row[1]).strip() if row[1] else 'Sconosciuto',
                    'note': str(row[2]).strip() if len(row) > 2 and row[2] else ''
                })
        
        wb.close()
        return mappatura
    
    except Exception as e:
        print(f"Errore lettura mappatura IP: {e}")
        return []


def riconosci_ip(ip_address):
    """
    Riconosce un IP dalla mappatura.
    
    Args:
        ip_address: Indirizzo IP da cercare
    
    Returns:
        dict: {'nome': 'Rete Locale', 'note': '...', 'riconosciuto': True/False}
    """
    if not ip_address:
        return {'nome': 'Sconosciuto', 'note': '', 'riconosciuto': False}
    
    mappatura = get_mappatura_ip()
    
    for entry in mappatura:
        prefisso = entry['prefisso']
        # Match per prefisso (es: 192.168.1. matcha 192.168.1.100)
        if ip_address.startswith(prefisso):
            return {
                'nome': entry['nome'],
                'note': entry['note'],
                'riconosciuto': True
            }
    
    return {'nome': 'Non riconosciuto', 'note': '', 'riconosciuto': False}


def get_log_accessi_con_rete(conn, utente_id=None, limite=50):
    """
    Restituisce log accessi con info rete riconosciuta.
    
    Args:
        conn: Connessione database
        utente_id: ID utente (None per tutti)
        limite: Numero max righe
    
    Returns:
        list: Log con campo 'rete' aggiunto
    """
    cursor = conn.cursor()
    
    if utente_id:
        cursor.execute('''
            SELECT la.*, u.username, u.nome, u.cognome
            FROM log_accessi la
            LEFT JOIN utenti u ON la.utente_id = u.id
            WHERE la.utente_id = ?
            ORDER BY la.data_ora DESC
            LIMIT ?
        ''', (utente_id, limite))
    else:
        cursor.execute('''
            SELECT la.*, u.username, u.nome, u.cognome
            FROM log_accessi la
            LEFT JOIN utenti u ON la.utente_id = u.id
            ORDER BY la.data_ora DESC
            LIMIT ?
        ''', (limite,))
    
    logs = []
    for row in cursor.fetchall():
        log = dict(row)
        log['rete'] = riconosci_ip(log.get('ip_address'))
        logs.append(log)
    
    return logs


def get_ultimo_accesso_con_rete(conn, utente_id):
    """
    Restituisce l'ultimo accesso di un utente con info rete.
    
    Returns:
        dict o None
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT data_ora, ip_address
        FROM log_accessi
        WHERE utente_id = ? AND azione = 'login_ok'
        ORDER BY data_ora DESC
        LIMIT 1
    ''', (utente_id,))
    
    row = cursor.fetchone()
    if not row:
        return None
    
    return {
        'data_ora': row['data_ora'],
        'ip': row['ip_address'],
        'rete': riconosci_ip(row['ip_address'])
    }


def get_statistiche_accessi_per_rete(conn, utente_id=None):
    """
    Restituisce statistiche accessi raggruppate per rete.
    
    Returns:
        dict: {'Rete Locale': 15, 'VPN': 3, ...}
    """
    cursor = conn.cursor()
    
    if utente_id:
        cursor.execute('''
            SELECT ip_address, COUNT(*) as cnt
            FROM log_accessi
            WHERE utente_id = ? AND azione = 'login_ok'
            GROUP BY ip_address
        ''', (utente_id,))
    else:
        cursor.execute('''
            SELECT ip_address, COUNT(*) as cnt
            FROM log_accessi
            WHERE azione = 'login_ok'
            GROUP BY ip_address
        ''')
    
    per_rete = {}
    for row in cursor.fetchall():
        rete = riconosci_ip(row['ip_address'])
        nome = rete['nome']
        per_rete[nome] = per_rete.get(nome, 0) + row['cnt']
    
    return per_rete


# ==============================================================================
# FUNZIONI STORICO ASSEGNAZIONI COMMERCIALI
# ==============================================================================

def registra_assegnazione(conn, cliente_nome, cliente_piva, commerciale_precedente, 
                          commerciale_nuovo, utente_id, note=None):
    """
    Registra un'assegnazione commerciale nello storico.
    
    Args:
        conn: Connessione database
        cliente_nome: Nome del cliente
        cliente_piva: P.IVA del cliente (opzionale)
        commerciale_precedente: Commerciale precedente (None se prima assegnazione)
        commerciale_nuovo: Nuovo commerciale assegnato
        utente_id: ID dell'utente che ha fatto l'assegnazione
        note: Note opzionali
    
    Returns:
        int: ID del record creato
    """
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        INSERT INTO storico_assegnazioni 
        (cliente_nome, cliente_piva, commerciale_precedente, commerciale_nuovo, 
         utente_id, data_ora, note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (cliente_nome, cliente_piva, commerciale_precedente, commerciale_nuovo,
          utente_id, now, note))
    
    conn.commit()
    return cursor.lastrowid


def get_storico_assegnazioni(conn, cliente_nome=None, commerciale=None, 
                              utente_id=None, limite=100):
    """
    Recupera lo storico assegnazioni con filtri opzionali.
    
    Args:
        conn: Connessione database
        cliente_nome: Filtra per cliente (opzionale)
        commerciale: Filtra per commerciale nuovo (opzionale)
        utente_id: Filtra per utente che ha assegnato (opzionale)
        limite: Numero massimo di record
    
    Returns:
        list: Lista di dict con storico assegnazioni
    """
    cursor = conn.cursor()
    
    query = '''
        SELECT sa.*, u.username, u.nome as operatore_nome, u.cognome as operatore_cognome
        FROM storico_assegnazioni sa
        LEFT JOIN utenti u ON sa.utente_id = u.id
        WHERE 1=1
    '''
    params = []
    
    if cliente_nome:
        query += ' AND sa.cliente_nome LIKE ?'
        params.append(f'%{cliente_nome}%')
    
    if commerciale:
        query += ' AND (sa.commerciale_nuovo = ? OR sa.commerciale_precedente = ?)'
        params.extend([commerciale, commerciale])
    
    if utente_id:
        query += ' AND sa.utente_id = ?'
        params.append(utente_id)
    
    query += ' ORDER BY sa.data_ora DESC LIMIT ?'
    params.append(limite)
    
    cursor.execute(query, params)
    return [dict(row) for row in cursor.fetchall()]


def get_storico_cliente(conn, cliente_nome):
    """
    Recupera tutto lo storico assegnazioni per un cliente specifico.
    
    Args:
        conn: Connessione database
        cliente_nome: Nome esatto del cliente
    
    Returns:
        list: Lista di dict con storico assegnazioni
    """
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT sa.*, u.username, u.nome as operatore_nome, u.cognome as operatore_cognome
        FROM storico_assegnazioni sa
        LEFT JOIN utenti u ON sa.utente_id = u.id
        WHERE sa.cliente_nome = ?
        ORDER BY sa.data_ora DESC
    ''', (cliente_nome,))
    
    return [dict(row) for row in cursor.fetchall()]


# ==============================================================================
# FUNZIONI STORICO EXPORT
# ==============================================================================

def registra_export(conn, utente_id, tipo_export, filtri_applicati=None, 
                    num_record=None, nome_file=None):
    """
    Registra un export nello storico.
    
    Args:
        conn: Connessione database
        utente_id: ID dell'utente che ha esportato
        tipo_export: Tipo di export (es: 'clienti', 'flotta', 'veicoli')
        filtri_applicati: JSON string dei filtri usati (opzionale)
        num_record: Numero di record esportati (opzionale)
        nome_file: Nome del file generato (opzionale)
    
    Returns:
        int: ID del record creato
    """
    import json
    
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Converti filtri in JSON se ÃƒÂ¨ un dict
    if isinstance(filtri_applicati, dict):
        filtri_applicati = json.dumps(filtri_applicati, ensure_ascii=False)
    
    cursor.execute('''
        INSERT INTO storico_export 
        (utente_id, tipo_export, filtri_applicati, num_record, nome_file, data_ora)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (utente_id, tipo_export, filtri_applicati, num_record, nome_file, now))
    
    conn.commit()
    return cursor.lastrowid


def get_storico_export(conn, utente_id=None, tipo_export=None, limite=100):
    """
    Recupera lo storico export con filtri opzionali.
    
    Args:
        conn: Connessione database
        utente_id: Filtra per utente (opzionale)
        tipo_export: Filtra per tipo export (opzionale)
        limite: Numero massimo di record
    
    Returns:
        list: Lista di dict con storico export
    """
    cursor = conn.cursor()
    
    query = '''
        SELECT se.*, u.username, u.nome, u.cognome
        FROM storico_export se
        LEFT JOIN utenti u ON se.utente_id = u.id
        WHERE 1=1
    '''
    params = []
    
    if utente_id:
        query += ' AND se.utente_id = ?'
        params.append(utente_id)
    
    if tipo_export:
        query += ' AND se.tipo_export = ?'
        params.append(tipo_export)
    
    query += ' ORDER BY se.data_ora DESC LIMIT ?'
    params.append(limite)
    
    cursor.execute(query, params)
    return [dict(row) for row in cursor.fetchall()]


def get_statistiche_export(conn, utente_id=None):
    """
    Restituisce statistiche sugli export.
    
    Args:
        conn: Connessione database
        utente_id: Filtra per utente (opzionale)
    
    Returns:
        dict: Statistiche per tipo export
    """
    cursor = conn.cursor()
    
    if utente_id:
        cursor.execute('''
            SELECT tipo_export, COUNT(*) as conteggio, SUM(num_record) as totale_record
            FROM storico_export
            WHERE utente_id = ?
            GROUP BY tipo_export
            ORDER BY conteggio DESC
        ''', (utente_id,))
    else:
        cursor.execute('''
            SELECT tipo_export, COUNT(*) as conteggio, SUM(num_record) as totale_record
            FROM storico_export
            GROUP BY tipo_export
            ORDER BY conteggio DESC
        ''')
    
    return [dict(row) for row in cursor.fetchall()]


# ==============================================================================
# FUNZIONI HELPER NOME COMMERCIALE
# ==============================================================================

def get_nome_commerciale_utente(conn, utente_id):
    """
    Restituisce il nome commerciale flotta associato all'utente.
    
    Args:
        conn: Connessione database
        utente_id: ID utente
    
    Returns:
        str o None: Nome commerciale flotta
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT nome_commerciale_flotta FROM utenti WHERE id = ?
    ''', (utente_id,))
    
    row = cursor.fetchone()
    return row['nome_commerciale_flotta'] if row else None


def set_nome_commerciale_utente(conn, utente_id, nome_commerciale):
    """
    Imposta il nome commerciale flotta per un utente.
    
    Args:
        conn: Connessione database
        utente_id: ID utente
        nome_commerciale: Nome commerciale da assegnare
    
    Returns:
        bool: True se aggiornato
    """
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE utenti SET nome_commerciale_flotta = ? WHERE id = ?
    ''', (nome_commerciale, utente_id))
    
    conn.commit()
    return cursor.rowcount > 0


def get_utenti_commerciali(conn):
    """
    Restituisce lista utenti con nome_commerciale_flotta impostato.
    
    Returns:
        list: Lista di dict con id, username, nome, cognome, nome_commerciale_flotta
    """
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, username, nome, cognome, nome_commerciale_flotta
        FROM utenti
        WHERE nome_commerciale_flotta IS NOT NULL 
        AND nome_commerciale_flotta != ''
        AND attivo = 1
        ORDER BY nome_commerciale_flotta
    ''')
    
    return [dict(row) for row in cursor.fetchall()]
