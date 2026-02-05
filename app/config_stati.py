# ==============================================================================
# CONFIG_STATI.PY - Gestione Stati da File Excel
# ==============================================================================
# Versione: 1.0.0
# Data: 2026-01-26
# Descrizione: Modulo per leggere le configurazioni degli stati da file Excel
#              nella cartella impostazioni/
# ==============================================================================

import openpyxl
from pathlib import Path
from functools import lru_cache
import logging

logger = logging.getLogger(__name__)

# ==============================================================================
# CONFIGURAZIONE PERCORSI
# ==============================================================================

BASE_DIR = Path(__file__).resolve().parent.parent
IMPOSTAZIONI_DIR = BASE_DIR / "impostazioni"

# File Excel configurazione stati
STATI_CLIENTE_FILE = IMPOSTAZIONI_DIR / "stati_cliente.xlsx"
STATI_CRM_FILE = IMPOSTAZIONI_DIR / "stati_crm.xlsx"
STATI_NOLEGGIATORE_FILE = IMPOSTAZIONI_DIR / "stati_noleggiatore.xlsx"
SCAGLIONI_FLOTTA_FILE = IMPOSTAZIONI_DIR / "scaglioni_flotta.xlsx"


# ==============================================================================
# FUNZIONI DI LETTURA EXCEL
# ==============================================================================

def _leggi_excel_stati(filepath, cache_key=None):
    """
    Legge un file Excel di configurazione stati.
    
    Struttura attesa del file Excel:
    - Colonna A: Codice
    - Colonna B: Etichetta
    - Colonna C: Colore (hex)
    - Colonna D: Ordine
    - Colonna E: Note
    
    Args:
        filepath: Path al file Excel
        cache_key: Chiave per invalidare cache (opzionale)
    
    Returns:
        list: Lista di dict con chiavi: codice, etichetta, colore, ordine, note
    """
    try:
        if not filepath.exists():
            logger.warning(f"File configurazione non trovato: {filepath}")
            return []
        
        wb = openpyxl.load_workbook(str(filepath), read_only=True)
        ws = wb.active
        
        stati = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Salta header
            if row[0]:  # Se c'e' un codice
                stato = {
                    'codice': str(row[0]).strip() if row[0] else '',
                    'etichetta': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                    'colore': str(row[2]).strip() if len(row) > 2 and row[2] else '#6c757d',
                    'ordine': int(row[3]) if len(row) > 3 and row[3] else 99,
                    'note': str(row[4]).strip() if len(row) > 4 and row[4] else ''
                }
                stati.append(stato)
        
        wb.close()
        
        # Ordina per campo ordine
        stati.sort(key=lambda x: x['ordine'])
        
        logger.debug(f"Caricati {len(stati)} stati da {filepath.name}")
        return stati
        
    except Exception as e:
        logger.error(f"Errore lettura {filepath}: {e}")
        return []


def _get_mappa_stati(stati_list):
    """
    Crea una mappa codice -> stato per lookup veloce.
    
    Args:
        stati_list: Lista di stati da _leggi_excel_stati()
    
    Returns:
        dict: {codice: {etichetta, colore, ordine, note}}
    """
    return {s['codice']: s for s in stati_list}


# ==============================================================================
# STATI CLIENTE
# ==============================================================================

def get_stati_cliente(force_reload=False):
    """
    Ritorna la lista degli stati cliente.
    
    Args:
        force_reload: Se True, ricarica dal file (ignora cache)
    
    Returns:
        list: Lista di stati cliente ordinati
    """
    if force_reload:
        get_stati_cliente_cached.cache_clear()
    return get_stati_cliente_cached()


@lru_cache(maxsize=1)
def get_stati_cliente_cached():
    """Versione cached di get_stati_cliente."""
    return _leggi_excel_stati(STATI_CLIENTE_FILE)


def get_stato_cliente_info(codice):
    """
    Ritorna le info di uno stato cliente specifico.
    
    Args:
        codice: Codice dello stato (es. "Prospetto", "Cliente")
    
    Returns:
        dict: {codice, etichetta, colore, ordine, note} o None
    """
    stati = get_stati_cliente()
    mappa = _get_mappa_stati(stati)
    return mappa.get(codice)


def get_stato_cliente_colore(codice):
    """
    Ritorna il colore hex di uno stato cliente.
    
    Args:
        codice: Codice dello stato
    
    Returns:
        str: Colore hex (es. "#28a745") o grigio di default
    """
    info = get_stato_cliente_info(codice)
    return info['colore'] if info else '#6c757d'


def get_stato_cliente_etichetta(codice):
    """
    Ritorna l'etichetta di uno stato cliente.
    
    Args:
        codice: Codice dello stato
    
    Returns:
        str: Etichetta o codice stesso se non trovato
    """
    info = get_stato_cliente_info(codice)
    return info['etichetta'] if info else codice


# ==============================================================================
# STATI CRM
# ==============================================================================

def get_stati_crm(force_reload=False):
    """
    Ritorna la lista degli stati CRM.
    
    Args:
        force_reload: Se True, ricarica dal file (ignora cache)
    
    Returns:
        list: Lista di stati CRM ordinati
    """
    if force_reload:
        get_stati_crm_cached.cache_clear()
    return get_stati_crm_cached()


@lru_cache(maxsize=1)
def get_stati_crm_cached():
    """Versione cached di get_stati_crm."""
    return _leggi_excel_stati(STATI_CRM_FILE)


def get_stato_crm_info(codice):
    """
    Ritorna le info di uno stato CRM specifico.
    
    Args:
        codice: Codice dello stato (es. "PROSPECT_NOSTRO")
    
    Returns:
        dict: {codice, etichetta, colore, ordine, note} o None
    """
    stati = get_stati_crm()
    mappa = _get_mappa_stati(stati)
    return mappa.get(codice)


def get_stato_crm_colore(codice):
    """
    Ritorna il colore hex di uno stato CRM.
    """
    info = get_stato_crm_info(codice)
    return info['colore'] if info else '#6c757d'


def get_stato_crm_etichetta(codice):
    """
    Ritorna l'etichetta di uno stato CRM.
    """
    info = get_stato_crm_info(codice)
    return info['etichetta'] if info else codice


# ==============================================================================
# STATI NOLEGGIATORE (Relazione cliente-noleggiatore)
# ==============================================================================

def get_stati_noleggiatore(force_reload=False):
    """
    Ritorna la lista degli stati relazione noleggiatore.
    
    Args:
        force_reload: Se True, ricarica dal file (ignora cache)
    
    Returns:
        list: Lista di stati noleggiatore ordinati
    """
    if force_reload:
        get_stati_noleggiatore_cached.cache_clear()
    return get_stati_noleggiatore_cached()


@lru_cache(maxsize=1)
def get_stati_noleggiatore_cached():
    """Versione cached di get_stati_noleggiatore."""
    return _leggi_excel_stati(STATI_NOLEGGIATORE_FILE)


def get_stato_noleggiatore_info(codice):
    """
    Ritorna le info di uno stato noleggiatore specifico.
    
    Args:
        codice: Codice dello stato (es. "NOSTRI", "ALTRO_BROKER")
    
    Returns:
        dict: {codice, etichetta, colore, ordine, note} o None
    """
    stati = get_stati_noleggiatore()
    mappa = _get_mappa_stati(stati)
    return mappa.get(codice)


def get_stato_noleggiatore_colore(codice):
    """
    Ritorna il colore hex di uno stato noleggiatore.
    """
    info = get_stato_noleggiatore_info(codice)
    return info['colore'] if info else '#6c757d'


def get_stato_noleggiatore_etichetta(codice):
    """
    Ritorna l'etichetta di uno stato noleggiatore.
    """
    info = get_stato_noleggiatore_info(codice)
    return info['etichetta'] if info else codice


# ==============================================================================
# SCAGLIONI FLOTTA (colori icona in base al numero veicoli)
# ==============================================================================

def get_scaglioni_flotta(force_reload=False):
    """
    Ritorna la lista degli scaglioni flotta dal file Excel.
    
    Struttura file Excel:
    - Colonna A: Min Veicoli (soglia minima)
    - Colonna B: Colore (hex)
    - Colonna C: Note
    
    Args:
        force_reload: Se True, ricarica dal file (ignora cache)
    
    Returns:
        list: Lista di dict {min_veicoli, colore, note} ordinata per min_veicoli DESC
    """
    if force_reload:
        get_scaglioni_flotta_cached.cache_clear()
    return get_scaglioni_flotta_cached()


@lru_cache(maxsize=1)
def get_scaglioni_flotta_cached():
    """Versione cached di get_scaglioni_flotta."""
    try:
        if not SCAGLIONI_FLOTTA_FILE.exists():
            logger.warning(f"File scaglioni flotta non trovato: {SCAGLIONI_FLOTTA_FILE}")
            return []
        
        wb = openpyxl.load_workbook(str(SCAGLIONI_FLOTTA_FILE), read_only=True)
        ws = wb.active
        
        scaglioni = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                scaglione = {
                    'min_veicoli': int(row[0]) if row[0] else 0,
                    'colore': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                    'note': str(row[2]).strip() if len(row) > 2 and row[2] else ''
                }
                scaglioni.append(scaglione)
        
        wb.close()
        
        # Ordina per min_veicoli DECRESCENTE (per trovare lo scaglione giusto)
        scaglioni.sort(key=lambda x: x['min_veicoli'], reverse=True)
        
        logger.debug(f"Caricati {len(scaglioni)} scaglioni flotta")
        return scaglioni
        
    except Exception as e:
        logger.error(f"Errore lettura scaglioni flotta: {e}")
        return []


def get_colore_flotta(num_veicoli):
    """
    Ritorna il colore per l'icona flotta in base al numero di veicoli.
    
    Args:
        num_veicoli: Numero di veicoli del cliente
    
    Returns:
        str: Colore hex o stringa vuota se nessuna icona
    """
    if num_veicoli is None:
        num_veicoli = 0
    
    scaglioni = get_scaglioni_flotta()
    
    # Trova lo scaglione corretto (lista ordinata DESC per min_veicoli)
    for scaglione in scaglioni:
        if num_veicoli >= scaglione['min_veicoli']:
            return scaglione['colore']
    
    return ''  # Nessun colore = nessuna icona


# ==============================================================================
# TIPI VEICOLO (Hardcoded - non da Excel)
# ==============================================================================

TIPI_VEICOLO = [
    {
        'codice': 'Installato',
        'etichetta': 'Installato',
        'colore': '#28a745',  # Verde
        'note': 'Veicoli gestiti da BR Car Service'
    },
    {
        'codice': 'Extra',
        'etichetta': 'Extra',
        'colore': '#fd7e14',  # Arancione
        'note': 'Veicoli esterni, gestiti da altro broker'
    }
]


def get_tipi_veicolo():
    """
    Ritorna la lista dei tipi veicolo.
    
    Returns:
        list: Lista di tipi veicolo
    """
    return TIPI_VEICOLO


def get_tipo_veicolo_info(codice):
    """
    Ritorna le info di un tipo veicolo.
    
    Args:
        codice: "Installato" o "Extra"
    
    Returns:
        dict o None
    """
    for t in TIPI_VEICOLO:
        if t['codice'] == codice:
            return t
    return None


def get_tipo_veicolo_colore(codice):
    """
    Ritorna il colore di un tipo veicolo.
    """
    info = get_tipo_veicolo_info(codice)
    return info['colore'] if info else '#6c757d'


# ==============================================================================
# FUNZIONI UTILITY
# ==============================================================================

def reload_all_cache():
    """
    Ricarica tutte le cache dei file Excel.
    Utile dopo modifiche ai file di configurazione.
    """
    get_stati_cliente_cached.cache_clear()
    get_stati_crm_cached.cache_clear()
    get_stati_noleggiatore_cached.cache_clear()
    get_scaglioni_flotta_cached.cache_clear()
    logger.info("Cache configurazioni stati ricaricata")


def get_all_config():
    """
    Ritorna tutte le configurazioni in un unico dict.
    Utile per debug o export.
    
    Returns:
        dict: {stati_cliente, stati_crm, stati_noleggiatore, tipi_veicolo}
    """
    return {
        'stati_cliente': get_stati_cliente(),
        'stati_crm': get_stati_crm(),
        'stati_noleggiatore': get_stati_noleggiatore(),
        'tipi_veicolo': get_tipi_veicolo()
    }


# ==============================================================================
# CONTEXT PROCESSOR PER JINJA2
# ==============================================================================

def stati_context_processor():
    """
    Context processor da registrare nell'app Flask.
    Inietta le funzioni di lookup stati nei template.
    
    Uso in template:
        {{ get_stato_cliente_colore(cliente.stato_cliente) }}
        {{ get_tipo_veicolo_colore(veicolo.tipo_veicolo) }}
    """
    return {
        'get_stato_cliente_colore': get_stato_cliente_colore,
        'get_stato_cliente_etichetta': get_stato_cliente_etichetta,
        'get_stato_crm_colore': get_stato_crm_colore,
        'get_stato_crm_etichetta': get_stato_crm_etichetta,
        'get_stato_noleggiatore_colore': get_stato_noleggiatore_colore,
        'get_stato_noleggiatore_etichetta': get_stato_noleggiatore_etichetta,
        'get_tipo_veicolo_colore': get_tipo_veicolo_colore,
        'get_colore_flotta': get_colore_flotta,
        'stati_cliente': get_stati_cliente(),
        'stati_crm': get_stati_crm(),
        'stati_noleggiatore': get_stati_noleggiatore(),
        'tipi_veicolo': get_tipi_veicolo(),
    }


# ==============================================================================
# TEST
# ==============================================================================

if __name__ == "__main__":
    # Test di lettura
    print("=== TEST CONFIG_STATI ===\n")
    
    print("STATI CLIENTE:")
    for s in get_stati_cliente():
        print(f"  {s['codice']:<45} | {s['colore']} | {s['ordine']}")
    
    print("\nSTATI CRM:")
    for s in get_stati_crm():
        print(f"  {s['codice']:<20} | {s['colore']} | {s['ordine']}")
    
    print("\nSTATI NOLEGGIATORE:")
    for s in get_stati_noleggiatore():
        print(f"  {s['codice']:<15} | {s['colore']} | {s['ordine']}")
    
    print("\nTIPI VEICOLO:")
    for t in get_tipi_veicolo():
        print(f"  {t['codice']:<12} | {t['colore']}")
    
    print("\n=== TEST LOOKUP ===")
    print(f"Colore 'Cliente': {get_stato_cliente_colore('Cliente')}")
    print(f"Colore 'Prospetto': {get_stato_cliente_colore('Prospetto')}")
    print(f"Colore 'NOSTRI': {get_stato_noleggiatore_colore('NOSTRI')}")
    print(f"Colore 'Extra': {get_tipo_veicolo_colore('Extra')}")
    print(f"Colore 'Installato': {get_tipo_veicolo_colore('Installato')}")
