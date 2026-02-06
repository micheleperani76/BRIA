#!/usr/bin/env python3
# ==============================================================================
# GESTIONE FLOTTA - Web Server
# ==============================================================================
# Versione: 2.0.0
# Data: 2025-01-19
# Descrizione: Server web Flask per interfaccia utente
# Novita v2.0: Nuova struttura cartelle clienti (clienti/CF/ e clienti/PIVA/)
# ==============================================================================

import os
import sys
import sqlite3
import psutil
import threading
import shutil
import uuid
from datetime import datetime
from pathlib import Path
import re
from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, Response, session, flash
from werkzeug.utils import secure_filename

# Aggiungi parent directory al path per import moduli
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.config import (
    BASE_DIR, DB_FILE, PDF_DIR, TEMPLATES_DIR, LOGS_DIR,
    WEB_HOST, WEB_PORT, WEB_DEBUG,
    PROVINCE_REGIONI, REGIONI, SCORE_COLORS, ALIMENTAZIONI_COLORS,
    CLIENTI_DIR, CLIENTI_CF_DIR, CLIENTI_PIVA_DIR,
    get_cliente_base_path, get_cliente_allegati_path, get_cliente_creditsafe_path,
    ensure_cliente_folders, pulisci_piva
)
from app.database import init_database, get_connection, get_statistiche_generali
from app.utils import (
    setup_logger, formatta_numero, formatta_euro, formatta_data,
    estrai_provincia, giorni_mancanti, pulisci_log_vecchi
)
from app.import_creditsafe import importa_tutti_pdf
# Import modulo identificativo
from app.utils_identificativo import (
    get_identificativo_cliente,
    get_identificativo_or_id,
    cerca_cliente_per_identificativo,
    get_cartella_allegati_cliente,
    get_cartella_nota_cliente,
    trova_cartella_nota_esistente,
    url_cliente,
    normalizza_piva,
    normalizza_cf
)

# Import modulo export Excel configurabile
from app.routes_export import register_export_routes
from app.routes_documenti_cliente import register_documenti_cliente_routes
from app.routes_documenti_strutturati import register_documenti_strutturati_routes
from app.routes_note_clienti import note_clienti_bp, register_note_clienti_legacy_routes
# Import modulo autenticazione
from app.routes_auth import auth_bp
from app.routes_admin_utenti import admin_utenti_bp
from app.routes_flotta_commerciali import flotta_commerciali_bp
from app.routes_collegamenti_clienti import collegamenti_bp
from app.routes_noleggiatori_cliente import noleggiatori_cliente_bp
from app.routes_sedi_cliente import sedi_bp
from app.routes_trattative import trattative_bp
from app.routes_top_prospect import top_prospect_bp
from app.routes_trascrizione import trascrizione_bp
from app.routes_notifiche import notifiche_bp
from app.routes_ticker import ticker_bp
from app.config_notifiche import POLLING_SECONDI as NOTIFICHE_POLLING
from app.auth import auth_context_processor, login_required
# Import funzioni database utenti per filtro supervisioni
from app.database_utenti import get_subordinati
# Import modulo gestione commerciali
from app.gestione_commerciali import (
    get_commerciali_assegnabili, get_commerciale_display,
    get_commerciali_con_clienti, assegna_cliente,
    format_nome_commerciale, TIPO_MANUALE
)
# Import connettore indicatori stato cliente
from app.connettori_stato_cliente import (
    get_indicatori_clienti_bulk,
    indicatori_context_processor
)
# Import modulo configurazione stati da Excel
from app.config_stati import (
    get_stati_cliente, get_stato_cliente_colore, get_stato_cliente_etichetta,
    get_stati_crm, get_stato_crm_colore, get_stato_crm_etichetta,
    get_stati_noleggiatore, get_stato_noleggiatore_colore, get_stato_noleggiatore_etichetta,
    get_tipi_veicolo, get_tipo_veicolo_colore,
    stati_context_processor
)

# ==============================================================================
# CONFIGURAZIONE ALLEGATI
# ==============================================================================

# Le cartelle clienti sono ora in clienti/CF/ e clienti/PIVA/ (da config.py)
# Assicura che esistano
CLIENTI_CF_DIR.mkdir(parents=True, exist_ok=True)
CLIENTI_PIVA_DIR.mkdir(parents=True, exist_ok=True)

# Estensioni file consentite
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'png', 'jpg', 'jpeg', 'gif', 'zip'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==============================================================================
# CONFIGURAZIONE IMPOSTAZIONI
# ==============================================================================

IMPOSTAZIONI_DIR = BASE_DIR / "impostazioni"
IMPOSTAZIONI_DIR.mkdir(parents=True, exist_ok=True)

# File configurazione noleggiatori assistenza
NOLEGGIATORI_ASSISTENZA_FILE = IMPOSTAZIONI_DIR / "noleggiatori_assistenza.xlsx"

# ==============================================================================
# INIZIALIZZAZIONE FLASK
# ==============================================================================

app = Flask(__name__, 
            template_folder=str(TEMPLATES_DIR),
            static_folder=str(TEMPLATES_DIR / 'static'))

# Configurazione upload
app.config['MAX_CONTENT_LENGTH'] = 256 * 1024 * 1024  # 256 MB max

# Inizializza database all'avvio
init_database()

# Secret key per sessioni (cambiare in produzione!)
app.secret_key = "BR-CAR-SERVICE-2025-chiave-segreta-cambiarla-in-produzione"

# Logger
logger = setup_logger('web_server')

# Registra route export Excel
register_export_routes(app)
register_documenti_cliente_routes(app)
register_documenti_strutturati_routes(app)


# Registra blueprint autenticazione
app.register_blueprint(auth_bp)
app.register_blueprint(admin_utenti_bp)
app.register_blueprint(flotta_commerciali_bp)
app.register_blueprint(note_clienti_bp)
app.register_blueprint(collegamenti_bp)
app.register_blueprint(noleggiatori_cliente_bp)
app.register_blueprint(sedi_bp)
app.register_blueprint(trattative_bp)
app.register_blueprint(top_prospect_bp)
app.register_blueprint(trascrizione_bp)
app.register_blueprint(notifiche_bp)
app.register_blueprint(ticker_bp)
register_note_clienti_legacy_routes(app)
app.context_processor(auth_context_processor)
app.context_processor(stati_context_processor)

# Context processor per notifiche (polling campanella)
@app.context_processor
def notifiche_context_processor():
    return {"config_notifiche_polling": NOTIFICHE_POLLING}
# ==============================================================================
# FILTRI JINJA2 PERSONALIZZATI
# ==============================================================================

@app.template_filter('format_numero')
def filter_format_numero(value, decimali=0):
    return formatta_numero(value, decimali)

@app.template_filter('format_euro')
def filter_format_euro(value):
    return formatta_euro(value)

@app.template_filter('format_data')
def filter_format_data(value):
    return formatta_data(value)

@app.template_filter('giorni_scadenza')
def filter_giorni_scadenza(value):
    giorni = giorni_mancanti(value)
    if giorni is None:
        return ""
    if giorni < 0:
        return f"Scaduto da {abs(giorni)} giorni"
    if giorni == 0:
        return "Scade oggi"
    return f"Tra {giorni} giorni"

# ==============================================================================
# CONTEXT PROCESSOR (variabili globali per template)
# ==============================================================================

@app.context_processor
def inject_globals():
    return {
        'now': datetime.now(),
        'score_colors': SCORE_COLORS,
        'alimentazioni_colors': ALIMENTAZIONI_COLORS,
        'province': sorted(PROVINCE_REGIONI.keys()),
        'regioni': REGIONI,
    }

@app.context_processor
def inject_identificativo_helpers():
    """Rende disponibili le funzioni identificativo nei template."""
    return {
        'get_identificativo': get_identificativo_cliente,
        'url_cliente': url_cliente,
    }

@app.context_processor
def inject_query_helpers():
    """Helper per generare query string nei template."""
    def query_string_with(key, value):
        args = dict(request.args)
        args[key] = str(value)
        return '&'.join(f'{k}={v}' for k, v in args.items() if v)
    
    def query_string_without(*keys):
        args = dict(request.args)
        for key in keys:
            args.pop(key, None)
        return '&'.join(f'{k}={v}' for k, v in args.items() if v)
    
    return {
        'query_string_with': query_string_with,
        'query_string_without': query_string_without,
    }

# ==============================================================================
# FUNZIONI HELPER
# ==============================================================================

def estrai_provincia(indirizzo):
    """Estrae la sigla provincia da un indirizzo italiano."""
    if not indirizzo:
        return None
    
    # Pattern: cerca 2 lettere maiuscole alla fine o dopo CAP
    # Esempio: "VIA ROMA 1, 25020 DELLO BS" -> BS
    import re
    
    # Cerca pattern CAP + citta + provincia
    m = re.search(r'\b\d{5}\s+\w+\s+([A-Z]{2})\b', indirizzo.upper())
    if m:
        return m.group(1)
    
    # Cerca 2 lettere maiuscole alla fine
    m = re.search(r'\b([A-Z]{2})\s*$', indirizzo.upper())
    if m:
        return m.group(1)
    
    return None

# ==============================================================================
# ROUTE: HOME / DASHBOARD
# ==============================================================================

# ==============================================================================
# DASHBOARD
# ==============================================================================

@app.route('/')
def home():
    """Route principale - redirect a dashboard (futuro: check login)"""
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principale - pagina iniziale con widget e notifiche"""
    return render_template('dashboard.html')



# ==============================================================================
# FUNZIONE: Trova dove e' stato trovato il termine di ricerca (SMART SEARCH)
# ==============================================================================
def get_search_matches_per_cliente(conn, search_term, cliente_ids):
    """
    Trova in quali campi e' stato trovato il termine di ricerca PER OGNI CLIENTE.
    Ritorna un dict {cliente_id: {categoria: [valori]}} per mostrare l'indicatore per riga.
    Ritorna anche un riepilogo globale per il box in alto.
    """
    if not search_term or not cliente_ids:
        return {}, {}
    
    matches_per_cliente = {cid: {} for cid in cliente_ids}
    matches_globali = {}
    search_param = f'%{search_term}%'
    cursor = conn.cursor()
    ids_list = list(cliente_ids)
    placeholders = ','.join('?' * len(ids_list))
    
    try:
        # Referenti
        cursor.execute(f"""
            SELECT r.cliente_id, r.nome || ' ' || COALESCE(r.cognome, '') as ref_nome
            FROM referenti_clienti r
            WHERE r.cliente_id IN ({placeholders})
            AND (r.nome LIKE ? OR r.cognome LIKE ? OR r.telefono LIKE ? 
                 OR r.cellulare LIKE ? OR r.email_principale LIKE ?)
        """, ids_list + [search_param]*5)
        for row in cursor.fetchall():
            cid, ref_nome = row[0], row[1].strip() if row[1] else ''
            if ref_nome:
                if 'referente' not in matches_per_cliente[cid]:
                    matches_per_cliente[cid]['referente'] = []
                matches_per_cliente[cid]['referente'].append(ref_nome)
                if 'referente' not in matches_globali:
                    matches_globali['referente'] = []
                if ref_nome not in matches_globali['referente']:
                    matches_globali['referente'].append(ref_nome)
        
        # Driver
        cursor.execute(f"""
            SELECT v.cliente_id, v.driver
            FROM veicoli v
            WHERE v.cliente_id IN ({placeholders})
            AND v.driver LIKE ? AND v.driver IS NOT NULL AND v.driver != ''
        """, ids_list + [search_param])
        for row in cursor.fetchall():
            cid, driver = row[0], row[1]
            if driver:
                if 'driver' not in matches_per_cliente[cid]:
                    matches_per_cliente[cid]['driver'] = []
                matches_per_cliente[cid]['driver'].append(driver)
                if 'driver' not in matches_globali:
                    matches_globali['driver'] = []
                if driver not in matches_globali['driver']:
                    matches_globali['driver'].append(driver)
        
        # Targhe
        cursor.execute(f"""
            SELECT v.cliente_id, UPPER(v.targa)
            FROM veicoli v
            WHERE v.cliente_id IN ({placeholders})
            AND v.targa LIKE ?
        """, ids_list + [search_param])
        for row in cursor.fetchall():
            cid, targa = row[0], row[1]
            if targa:
                if 'targa' not in matches_per_cliente[cid]:
                    matches_per_cliente[cid]['targa'] = []
                matches_per_cliente[cid]['targa'].append(targa)
                if 'targa' not in matches_globali:
                    matches_globali['targa'] = []
                if targa not in matches_globali['targa']:
                    matches_globali['targa'].append(targa)
        
        # Capogruppo
        cursor.execute(f"""
            SELECT id, capogruppo_nome
            FROM clienti
            WHERE id IN ({placeholders})
            AND (capogruppo_nome LIKE ? OR capogruppo_cf LIKE ?)
            AND capogruppo_nome IS NOT NULL AND capogruppo_nome != ''
        """, ids_list + [search_param, search_param])
        for row in cursor.fetchall():
            cid, capo = row[0], row[1]
            if capo:
                if 'capogruppo' not in matches_per_cliente[cid]:
                    matches_per_cliente[cid]['capogruppo'] = []
                matches_per_cliente[cid]['capogruppo'].append(capo)
                if 'capogruppo' not in matches_globali:
                    matches_globali['capogruppo'] = []
                if capo not in matches_globali['capogruppo']:
                    matches_globali['capogruppo'].append(capo)
        
        # Note (titolo + testo)
        cursor.execute(f"""
            SELECT n.cliente_id, n.titolo
            FROM note_clienti n
            WHERE n.cliente_id IN ({placeholders})
            AND (n.titolo LIKE ? OR n.testo LIKE ?)
        """, ids_list + [search_param, search_param])
        for row in cursor.fetchall():
            cid, titolo = row[0], row[1] or 'Nota'
            if 'note' not in matches_per_cliente[cid]:
                matches_per_cliente[cid]['note'] = []
            matches_per_cliente[cid]['note'].append(titolo[:20])
        # Conta per globale
        note_clienti = sum(1 for cid in matches_per_cliente if matches_per_cliente[cid].get('note'))
        if note_clienti > 0:
            matches_globali['note'] = [f"{note_clienti} cliente/i"]
        
        # Telefono/PEC cliente diretto
        cursor.execute(f"""
            SELECT id
            FROM clienti
            WHERE id IN ({placeholders})
            AND (telefono LIKE ? OR pec LIKE ?)
        """, ids_list + [search_param, search_param])
        for row in cursor.fetchall():
            cid = row[0]
            if 'contatto' not in matches_per_cliente[cid]:
                matches_per_cliente[cid]['contatto'] = ['Tel/PEC']
        # Conta per globale
        contatto_clienti = sum(1 for cid in matches_per_cliente if matches_per_cliente[cid].get('contatto'))
        if contatto_clienti > 0:
            matches_globali['contatto'] = [f"{contatto_clienti} cliente/i"]
            
    except Exception as e:
        print(f"Errore get_search_matches_per_cliente: {e}")
    
    # Limita a 3 elementi per categoria nel globale
    for k in matches_globali:
        if isinstance(matches_globali[k], list) and len(matches_globali[k]) > 3:
            matches_globali[k] = matches_globali[k][:3]
    
    return matches_per_cliente, matches_globali


@app.route('/clienti')
@login_required
def index():
    """Lista clienti con filtri e paginazione."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Parametri filtro
    search = request.args.get('search', '').strip()
    score = request.args.get('score', '')
    stato_cliente = request.args.get('stato_cliente', '')  # Filtro per stato cliente
    zona = request.args.get('zona', '')  # Nuovo: unione province/regioni
    forma = request.args.get('forma', '')
    commerciale = request.args.get('commerciale', '')
    con_flotta = request.args.get('con_flotta', '')
    credito_min = request.args.get('credito_min', '')
    credito_max = request.args.get('credito_max', '')
    prov_veicolo = request.args.get('prov_veicolo', '')
    order = request.args.get('order', 'nome_cliente')
    
    # Parsing zona: P:XX = provincia, R:NOME = regione
    provincia = ''
    regione = ''
    if zona.startswith('P:'):
        provincia = zona[2:]
    elif zona.startswith('R:'):
        regione = zona[2:]
    
    # Parametri paginazione e indice alfabetico
    lettera = request.args.get('lettera', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    if per_page not in [10, 25, 50, 75, 100]:
        per_page = 50
    
    # Query base con subquery per num_veicoli e provincia_veicolo (per ordinamento server-side)
    query = """SELECT c.*, 
               (SELECT COUNT(*) FROM veicoli WHERE cliente_id = c.id) as num_veicoli,
               (SELECT province_code FROM veicoli WHERE cliente_id = c.id AND province_code IS NOT NULL AND province_code != '' LIMIT 1) as provincia_veicolo
               FROM clienti c WHERE 1=1"""
    params = []
    
    # ==========================================================================
    # NOTA: Tutti gli utenti vedono TUTTI i clienti nella lista
    # Il controllo accesso viene fatto nel dettaglio cliente
    # Questo permette di sapere se un cliente esiste gia e chi lo gestisce
    # Filtro speciale: ricerca "top prospect" mostra solo clienti Top Prospect confermati
    is_top_prospect_search = search.lower().replace(" ", "").replace("-", "") == "topprospect"
    if is_top_prospect_search:
        query += " AND EXISTS (SELECT 1 FROM top_prospect tp WHERE tp.cliente_id = c.id AND tp.stato IN ('candidato', 'confermato'))"
        search = ""  # Evita la ricerca normale

    # ==========================================================================

    if search:
        query += """ AND (
            c.nome_cliente LIKE ? OR c.ragione_sociale LIKE ? 
            OR c.p_iva LIKE ? OR c.cod_fiscale LIKE ? OR c.numero_registrazione LIKE ?
            OR c.telefono LIKE ? OR c.pec LIKE ?
            OR c.capogruppo_nome LIKE ? OR c.capogruppo_cf LIKE ?
            OR EXISTS (SELECT 1 FROM referenti_clienti r WHERE r.cliente_id = c.id 
                       AND (r.nome LIKE ? OR r.cognome LIKE ? 
                            OR r.telefono LIKE ? OR r.cellulare LIKE ? 
                            OR r.email_principale LIKE ? OR r.email_secondarie LIKE ?
                            OR (r.nome || ' ' || r.cognome) LIKE ?
                            OR (r.cognome || ' ' || r.nome) LIKE ?))
            OR EXISTS (SELECT 1 FROM veicoli v WHERE v.cliente_id = c.id 
                       AND (v.targa LIKE ? OR v.driver LIKE ?))
            OR EXISTS (SELECT 1 FROM note_clienti n WHERE n.cliente_id = c.id 
                       AND (n.titolo LIKE ? OR n.testo LIKE ?))
        )"""
        search_param = f'%{search}%'
        params.extend([search_param] * 21)
    
    # Filtro score (incluso NS per senza score)
    if score:
        if score == 'NS':
            query += " AND (c.score IS NULL OR c.score = '')"
        else:
            query += " AND c.score = ?"
            params.append(score)
    
    # Filtro stato cliente (Prospetto, Cliente, N/D, ecc.)
    if stato_cliente:
        if stato_cliente == '__NULL__':
            query += " AND (c.stato_cliente IS NULL OR c.stato_cliente = '')"
        else:
            query += " AND c.stato_cliente = ?"
            params.append(stato_cliente)
            
    # Filtro lettera iniziale
    if lettera:
        if lettera == '0-9':
            query += " AND (c.nome_cliente GLOB '[0-9]*' OR c.ragione_sociale GLOB '[0-9]*')"
        else:
            query += " AND (UPPER(c.nome_cliente) LIKE ? OR UPPER(c.ragione_sociale) LIKE ?)"
            params.extend([f'{lettera.upper()}%', f'{lettera.upper()}%'])
    
    # Filtro provincia (estratta da indirizzo)
    if provincia:
        query += " AND c.indirizzo LIKE ?"
        params.append(f'%{provincia}%')
    
    # Filtro regione
    if regione:
        province_regione = [p for p, r in PROVINCE_REGIONI.items() if r == regione]
        if province_regione:
            placeholders = ','.join(['?' for _ in province_regione])
            conditions = ' OR '.join([f"c.indirizzo LIKE ?" for _ in province_regione])
            query += f" AND ({conditions})"
            params.extend([f'%{p}%' for p in province_regione])
    
    # Filtro forma giuridica
    if forma:
        query += " AND c.forma_giuridica = ?"
        params.append(forma)
    
    # Filtro commerciale (usa commerciale_id)
    if commerciale:
        if commerciale == '__NULL__':
            query += " AND (c.commerciale_id IS NULL OR c.commerciale_id = 0)"
        else:
            try:
                commerciale_id_filtro = int(commerciale)
                query += " AND c.commerciale_id = ?"
                params.append(commerciale_id_filtro)
            except (ValueError, TypeError):
                pass  # Ignora valori non numerici
    
    # Filtro credito
    if credito_min:
        query += " AND c.credito >= ?"
        params.append(float(credito_min))
    if credito_max:
        query += " AND c.credito <= ?"
        params.append(float(credito_max))
    
    # Filtro provincia veicoli
    if prov_veicolo:
        query += " AND c.id IN (SELECT DISTINCT cliente_id FROM veicoli WHERE province_code = ?)"
        params.append(prov_veicolo)
    
    # Ordinamento (server-side per tutte le colonne)
    # Nota: SQLite mette NULL prima per default, usiamo CASE WHEN per gestire l'ordine corretto
    order_map = {
        # Cliente: ordine alfabetico normale
        'nome_cliente': 'c.nome_cliente ASC',
        'nome_cliente_desc': 'c.nome_cliente DESC',
        
        # Stato Cliente: ordine alfabetico, vuoti/null in fondo per ASC, in cima per DESC
        'stato_cliente': "CASE WHEN c.stato_cliente IS NULL OR c.stato_cliente = '' THEN 1 ELSE 0 END, c.stato_cliente ASC",
        'stato_cliente_desc': "CASE WHEN c.stato_cliente IS NULL OR c.stato_cliente = '' THEN 0 ELSE 1 END, c.stato_cliente DESC",
        
        # P.IVA: ordine alfabetico, vuoti in fondo per ASC, in cima per DESC
        'p_iva': "CASE WHEN c.p_iva IS NULL OR c.p_iva = '' THEN 1 ELSE 0 END, c.p_iva ASC",
        'p_iva_desc': "CASE WHEN c.p_iva IS NULL OR c.p_iva = '' THEN 0 ELSE 1 END, c.p_iva DESC",
        
        # Score: ordine A, B, C, D, E, NS (vuoti/null in fondo per ASC, in cima per DESC)
        'score': "CASE c.score WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'D' THEN 4 WHEN 'E' THEN 5 ELSE 6 END ASC",
        'score_desc': "CASE c.score WHEN 'E' THEN 1 WHEN 'D' THEN 2 WHEN 'C' THEN 3 WHEN 'B' THEN 4 WHEN 'A' THEN 5 ELSE 0 END ASC",
        
        # Telefono: ordine alfabetico, vuoti in fondo per ASC, in cima per DESC
        'telefono': "CASE WHEN c.telefono IS NULL OR c.telefono = '' THEN 1 ELSE 0 END, c.telefono ASC",
        'telefono_desc': "CASE WHEN c.telefono IS NULL OR c.telefono = '' THEN 0 ELSE 1 END, c.telefono DESC",
        
        # Provincia: ordine alfabetico, vuoti in fondo per ASC, in cima per DESC
        'provincia': "CASE WHEN provincia_veicolo IS NULL OR provincia_veicolo = '' THEN 1 ELSE 0 END, provincia_veicolo ASC",
        'provincia_desc': "CASE WHEN provincia_veicolo IS NULL OR provincia_veicolo = '' THEN 0 ELSE 1 END, provincia_veicolo DESC",
        
        # Flotta: ordine numerico (0 in fondo per ASC, in cima per DESC)
        'flotta': "CASE WHEN num_veicoli = 0 OR num_veicoli IS NULL THEN 1 ELSE 0 END, num_veicoli ASC",
        'flotta_desc': "CASE WHEN num_veicoli = 0 OR num_veicoli IS NULL THEN 0 ELSE 1 END, num_veicoli DESC",
        
        # Commerciale: ordine per nome (null in fondo per ASC, in cima per DESC)
        'commerciale': "CASE WHEN c.commerciale_id IS NULL OR c.commerciale_id = 0 THEN 1 ELSE 0 END, c.commerciale_id ASC",
        'commerciale_desc': "CASE WHEN c.commerciale_id IS NULL OR c.commerciale_id = 0 THEN 0 ELSE 1 END, c.commerciale_id DESC",
        
        # Credito: ordine numerico
        'credito': 'c.credito ASC',
        'credito_desc': 'c.credito DESC',
        
        # Data: ordine cronologico
        'data': 'c.data_ultimo_aggiornamento ASC',
        'data_desc': 'c.data_ultimo_aggiornamento DESC',
    }
    
    query += f" ORDER BY {order_map.get(order, 'c.nome_cliente ASC')}"
    
    cursor.execute(query, params)
    clienti = [dict(row) for row in cursor.fetchall()]
    
    # Lookup bulk per nomi commerciali (efficiente)
    from app.gestione_commerciali import get_commerciale_display_bulk
    commerciale_ids = [c.get('commerciale_id') for c in clienti if c.get('commerciale_id')]
    commerciali_display = get_commerciale_display_bulk(conn, commerciale_ids)
    
    # Processa i clienti (num_veicoli e provincia_veicolo gia dalla query)
    for cliente in clienti:
        # Aggiungi nome commerciale display
        comm_id = cliente.get('commerciale_id')
        cliente['commerciale_display'] = commerciali_display.get(comm_id, 'Non assegnato') if comm_id else 'Non assegnato'
        
        # Usa provincia dalla subquery, con fallback a indirizzo
        if not cliente.get('provincia_veicolo'):
            cliente['provincia'] = estrai_provincia(cliente.get('indirizzo'))
        else:
            cliente['provincia'] = cliente['provincia_veicolo']
    
        # Verifica se ha referente principale
        cursor.execute('''
            SELECT COUNT(*) FROM referenti_clienti 
            WHERE cliente_id = ? AND principale = 1
        ''', (cliente['id'],))
        cliente['has_referente_principale'] = cursor.fetchone()[0] > 0
    
    # Filtro con_flotta (dopo aver contato i veicoli)
    if con_flotta == 'si':
        clienti = [c for c in clienti if c['num_veicoli'] > 0]
    elif con_flotta == 'no':
        clienti = [c for c in clienti if c['num_veicoli'] == 0]
    
    # Statistiche per cards
    stats = {
        'totale': len(clienti),
        'score_A': len([c for c in clienti if c.get('score') == 'A']),
        'score_B': len([c for c in clienti if c.get('score') == 'B']),
        'score_C': len([c for c in clienti if c.get('score') == 'C']),
        'score_D': len([c for c in clienti if c.get('score') == 'D']),
        'score_E': len([c for c in clienti if c.get('score') == 'E']),
        'score_NS': len([c for c in clienti if not c.get('score')]),
    }
    
    # Paginazione
    total_clienti = len(clienti)
    total_pages = (total_clienti + per_page - 1) // per_page if total_clienti > 0 else 1
    
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    clienti_paginati = clienti[start_idx:end_idx]
    
    # Lista commerciali per dropdown (solo chi ha ruolo commerciale)
    from app.gestione_commerciali import get_commerciali_tutti
    commerciali = get_commerciali_tutti(conn, solo_attivi=True)
    
    # Lista forme giuridiche per dropdown
    cursor.execute('''
        SELECT fg.codice, fg.descrizione FROM forme_giuridiche fg WHERE fg.descrizione IN (SELECT DISTINCT forma_giuridica FROM clienti WHERE forma_giuridica IS NOT NULL AND forma_giuridica != '') 
        ORDER BY ordine
        -- forme giuridiche ordinate
    ''')
    forme_giuridiche = [{'codice': row[0], 'descrizione': row[1]} for row in cursor.fetchall()]
    
    # Lista stati cliente effettivamente usati nel DB (per dropdown)
    cursor.execute('''
        SELECT DISTINCT stato_cliente FROM clienti 
        WHERE stato_cliente IS NOT NULL AND stato_cliente != ''
        ORDER BY stato_cliente
    ''')
    stati_cliente_usati = [row[0] for row in cursor.fetchall()]
    
    # Recupera indicatori stato per i clienti paginati (car policy, documenti scadenza, ecc.)
    indicatori_cliente = get_indicatori_clienti_bulk(conn, clienti_paginati)
    
    # Ricerca smart: trova dove e' stato trovato il termine (per cliente + globale)
    search_matches = {}
    search_matches_per_cliente = {}
    if search and clienti_paginati:
        cliente_ids = [c['id'] for c in clienti_paginati]
        search_matches_per_cliente, search_matches = get_search_matches_per_cliente(conn, search, cliente_ids)
    
    conn.close()
    
    # Controlla se ci sono filtri attivi
    filtri_attivi = any([search, score, stato_cliente, zona, forma, commerciale, 
                         con_flotta, credito_min, credito_max, prov_veicolo])
    
    return render_template('index.html',
                         clienti=clienti_paginati,
                         stats=stats,
                         commerciali=commerciali,
                         forme_giuridiche=forme_giuridiche,
                         stati_cliente_usati=stati_cliente_usati,
                         indicatori_cliente=indicatori_cliente,
                         filtri_attivi=filtri_attivi,
                         # Valori filtri correnti
                         search=search,
                         score=score,
                         stato_cliente=stato_cliente,
                         zona=zona,
                         forma=forma,
                         commerciale=commerciale,
                         con_flotta=con_flotta,
                         credito_min=credito_min,
                         credito_max=credito_max,
                         prov_veicolo=prov_veicolo,
                         order=order,
                         page=page,
                         per_page=per_page,
                         total_clienti=total_clienti,
                         total_pages=total_pages,
                         lettera=lettera,
                         search_matches=search_matches,
                         search_matches_per_cliente=search_matches_per_cliente)


# ==============================================================================
# ROUTE: DETTAGLIO CLIENTE
# ==============================================================================

def _render_dettaglio_cliente(cliente_id):
    """
    Funzione helper che prepara e renderizza il dettaglio cliente.
    Usata sia da /cliente/<id> che da /cerca/<identificativo>.
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    # Recupera cliente
    cursor.execute('SELECT * FROM clienti WHERE id = ?', (cliente_id,))
    cliente = cursor.fetchone()
    
    if not cliente:
        conn.close()
        return None
    
    cliente = dict(cliente)
    
    # ==========================================================================
    # CONTROLLO ACCESSO: verifica se l'utente puo' vedere questo cliente
    # ==========================================================================
    # Admin vede tutto
    # Se cliente non ha commerciale -> accesso consentito (per assegnarlo)
    # Se commerciale_id e' dell'utente o di un suo subordinato -> accesso consentito
    # Altrimenti -> blocca accesso
    from app.database_utenti import get_subordinati
    from app.gestione_commerciali import get_commerciale_display
    from app.database_utenti import ha_permesso
    
    commerciale_id_cliente = cliente.get('commerciale_id')
    user_ruolo = session.get('ruolo_base')
    user_id = session.get('user_id')
    
    puo_accedere = False
    commerciale_gestore = None
    
    # Calcola permesso per gestire assegnazioni
    puo_gestire_assegnazioni = user_ruolo == 'admin' or ha_permesso(conn, user_id, 'flotta_assegnazioni')
    
    if user_ruolo == 'admin':
        puo_accedere = True
    elif not commerciale_id_cliente or commerciale_id_cliente == 0:
        # Cliente senza commerciale -> tutti possono accedere
        puo_accedere = True
    else:
        # Verifica se e' il commerciale assegnato o un suo subordinato
        ids_visibili = get_subordinati(conn, user_id)
        if commerciale_id_cliente in ids_visibili:
            puo_accedere = True
        else:
            commerciale_gestore = get_commerciale_display(conn, commerciale_id_cliente)
    
    if not puo_accedere:
        conn.close()
        flash(f'Questo cliente e gestito da {commerciale_gestore}. Non hai i permessi per accedere.', 'warning')
        return redirect(url_for('index'))
    
    # Aggiungi identificativo e URL stabile al cliente
    cliente['_identificativo'] = get_identificativo_cliente(cliente)
    cliente['_url'] = url_cliente(cliente)
    
    # Recupera veicoli tramite cliente_id (con conteggio note)
    cursor.execute('''
        SELECT v.*, 
               (SELECT COUNT(*) FROM note_veicoli WHERE veicolo_id = v.id AND eliminato = 0) as num_note
        FROM veicoli v
        WHERE v.cliente_id = ?
        ORDER BY v.noleggiatore, v.scadenza
    ''', (cliente_id,))
    veicoli = [dict(row) for row in cursor.fetchall()]
    
    # Calcola statistiche veicoli con separazione Installato/Extra
    veicoli_installato = [v for v in veicoli if v.get('tipo_veicolo') == 'Installato']
    veicoli_extra = [v for v in veicoli if v.get('tipo_veicolo') != 'Installato']
    
    stats_veicoli = {
        'totale': len(veicoli),
        'installato': len(veicoli_installato),
        'extra': len(veicoli_extra),
        'canone_totale': sum(v.get('canone', 0) or 0 for v in veicoli),
        'canone_installato': sum(v.get('canone', 0) or 0 for v in veicoli_installato),
        'canone_extra': sum(v.get('canone', 0) or 0 for v in veicoli_extra),
        'noleggiatori': list(set(v.get('noleggiatore') for v in veicoli if v.get('noleggiatore'))),
    }
    
    # Raggruppa veicoli per noleggiatore
    veicoli_per_noleggiatore = {}
    for v in veicoli:
        nol = v.get('noleggiatore', 'Non specificato') or 'Non specificato'
        if nol not in veicoli_per_noleggiatore:
            veicoli_per_noleggiatore[nol] = {'veicoli': [], 'canone': 0}
        veicoli_per_noleggiatore[nol]['veicoli'].append(v)
        veicoli_per_noleggiatore[nol]['canone'] += v.get('canone', 0) or 0
    
    # Storico modifiche
    cursor.execute('''
        SELECT * FROM storico_modifiche 
        WHERE tabella = 'clienti' AND record_id = ?
        ORDER BY data_modifica DESC LIMIT 20
    ''', (cliente_id,))
    storico = [dict(row) for row in cursor.fetchall()]
    
    # Referenti cliente
    cursor.execute('''
        SELECT * FROM referenti_clienti 
        WHERE cliente_id = ?
        ORDER BY principale DESC, cognome, nome
    ''', (cliente_id,))
    referenti_rows = cursor.fetchall()
    
    # Per ogni referente, recupera le sedi collegate
    referenti = []
    for ref_row in referenti_rows:
        ref = dict(ref_row)
        cursor.execute('''
            SELECT id, tipo_sede, denominazione, citta 
            FROM sedi_cliente 
            WHERE referente_id = ?
        ''', (ref['id'],))
        ref['sedi_collegate'] = [dict(s) for s in cursor.fetchall()]
        referenti.append(ref)
    
    # Note cliente (solo non eliminate)
    cursor.execute('''
        SELECT * FROM note_clienti 
        WHERE cliente_id = ? AND (eliminato = 0 OR eliminato IS NULL)
        ORDER BY fissata DESC, data_creazione DESC
    ''', (cliente_id,))
    note_cliente_rows = cursor.fetchall()
    
    # Carica allegati per ogni nota
    note_cliente = []
    for nota_row in note_cliente_rows:
        nota = dict(nota_row)
        cursor.execute('''
            SELECT * FROM allegati_note 
            WHERE nota_cliente_id = ?
            ORDER BY data_upload
        ''', (nota['id'],))
        nota['allegati'] = [dict(a) for a in cursor.fetchall()]
        note_cliente.append(nota)
    
    # =========================================
    # STORICO COMMERCIALI
    # =========================================
    from app.gestione_commerciali import get_storico_cliente, get_commerciale_display
    
    # Commerciale attuale (da commerciale_id)
    commerciale_attuale = None
    commerciale_attuale_display = 'Non assegnato'
    if cliente.get('commerciale_id'):
        commerciale_attuale = cliente['commerciale_id']
        commerciale_attuale_display = get_commerciale_display(conn, commerciale_attuale)
    
    # Storico assegnazioni per questo cliente
    storico_commerciali = get_storico_cliente(conn, cliente.get('p_iva'), limite=20)
    
    # Ex-commerciali manuali (dalla nuova tabella)
    ex_commerciali = []
    try:
        cursor.execute("""
            SELECT nome_commerciale, data_inizio, data_fine, note
            FROM ex_commerciali_cliente
            WHERE cliente_piva = ?
            ORDER BY data_fine DESC, data_inizio DESC
        """, (cliente.get('p_iva'),))
        for row in cursor.fetchall():
            ex_commerciali.append({
                'nome': row['nome_commerciale'],
                'data_inizio': row['data_inizio'],
                'data_fine': row['data_fine'],
                'note': row['note']
            })
    except:
        pass  # Tabella potrebbe non esistere ancora
    
    # Commerciale precedente: prima cerca negli ex-commerciali, poi nello storico automatico
    commerciale_precedente_display = None
    if ex_commerciali:
        commerciale_precedente_display = ex_commerciali[0]['nome']
    else:
        for s in storico_commerciali:
            if s['da'] != 'Non assegnato' and s['da'] != commerciale_attuale_display:
                commerciale_precedente_display = s['da']
                break
    
    
    # =========================================
    # COLLEGAMENTI CLIENTI
    # =========================================
    from app.routes_collegamenti_clienti import get_collegamenti_cliente, get_descrizione_relazione
    from app.database_utenti import get_subordinati
    collegamenti_raw = get_collegamenti_cliente(conn, cliente_id)
    subordinati_vis = get_subordinati(conn, user_id) if user_id else []
    collegamenti = []
    for col in collegamenti_raw:
        col['gestito'] = col.get('commerciale_id') in subordinati_vis
        col['tipo_relazione_desc'] = col.get('descrizione_relazione', col.get('tipo_relazione', ''))
        col['url'] = url_for('cliente_per_identificativo', identificativo=get_identificativo_cliente({'p_iva': col.get('altro_cliente_piva'), 'cod_fiscale': col.get('altro_cliente_cf')}))
        collegamenti.append(col)
    # TOP PROSPECT
    cursor.execute("SELECT stato, priorita FROM top_prospect WHERE cliente_id = ? AND stato IN ('candidato', 'confermato')", (cliente_id,))
    tp_row = cursor.fetchone()
    top_prospect_info = {"presente": bool(tp_row), "stato": tp_row["stato"] if tp_row else None, "priorita": tp_row["priorita"] if tp_row else None} if tp_row else {"presente": False, "stato": None, "priorita": None}

    conn.close()
    
    return render_template('dettaglio.html',
                         cliente=cliente,
                         veicoli=veicoli,
                         veicoli_per_noleggiatore=veicoli_per_noleggiatore,
                         stats_veicoli=stats_veicoli,
                         storico=storico,
                         note_cliente=note_cliente,
                         referenti=referenti,
                         commerciale_attuale_display=commerciale_attuale_display,
                         commerciale_precedente_display=commerciale_precedente_display,
                         storico_commerciali=storico_commerciali,
                         ex_commerciali=ex_commerciali,
                         puo_gestire_assegnazioni=puo_gestire_assegnazioni,
                         collegamenti=collegamenti,
                         top_prospect_info=top_prospect_info)


@app.route('/cliente/<int:cliente_id>')
def dettaglio_cliente(cliente_id):
    """Visualizza dettaglio completo di un cliente (retrocompatibilita con ID numerico)."""
    result = _render_dettaglio_cliente(cliente_id)
    if result is None:
        return "Cliente non trovato", 404
    return result


@app.route('/cliente/<int:cliente_id>/evernote')
def export_evernote(cliente_id):
    """Esporta i dati cliente in formato Evernote (.txt)."""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM clienti WHERE id = ?', (cliente_id,))
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return "Cliente non trovato", 404
    
    cliente = dict(row)
    
    # Recupera veicoli
    cursor.execute('''
        SELECT * FROM veicoli WHERE cliente_id = ? ORDER BY noleggiatore, targa
    ''', (cliente_id,))
    veicoli = [dict(row) for row in cursor.fetchall()]
    
    # Conta veicoli per noleggiatore
    veicoli_per_nol = {}
    for v in veicoli:
        nol = v.get('noleggiatore', 'N/D') or 'N/D'
        veicoli_per_nol[nol] = veicoli_per_nol.get(nol, 0) + 1
    
    conn.close()
    
    # Genera contenuto Evernote nel formato richiesto
    lines = []
    
    # Nome azienda
    lines.append(cliente.get('ragione_sociale') or cliente.get('nome_cliente') or '')
    
    # Commerciale
    commerciale = cliente.get('commerciale', '')
    lines.append(f"COLLEGATA ({commerciale})" if commerciale else "COLLEGATA ()")
    lines.append("")
    
    # Descrizione attivita
    lines.append(f"Descrizione attivita: {cliente.get('desc_attivita', '')}")
    lines.append("")
    
    # Data costituzione
    lines.append(f"Data di costituzione: {cliente.get('data_costituzione', '')}")
    lines.append("")
    
    # Indirizzo
    indirizzo = cliente.get('indirizzo', '')
    lines.append(indirizzo if indirizzo else "Via\ncap paese citta")
    lines.append("")
    
    # P.IVA e CF
    lines.append(f"P.IVA: {cliente.get('p_iva', '')}")
    lines.append(f"CF: {cliente.get('cod_fiscale', '')}")
    lines.append("")
    
    # Dati bancari (vuoti se non presenti)
    lines.append(f"IBAN: {cliente.get('iban', '')}")
    lines.append(f"BIC: {cliente.get('bic', '')}")
    lines.append(f"SDI: {cliente.get('sdi', '')}")
    lines.append("")
    
    # Legale rappresentante
    lines.append(f"LEGALE RAPPRESENTANTE: {cliente.get('legale_rappresentante', '')}")
    lines.append(f"C.F. LEGALE RAPPRESENTANTE: {cliente.get('cf_legale', '')}")
    lines.append(f"NATO/A IL  A ")
    lines.append("")
    
    # Contatti
    lines.append(f"REFERENTE: {cliente.get('referente', '')}")
    lines.append(f"CELLULARE: {cliente.get('cellulare', '')}")
    lines.append(f"TELEFONO: {cliente.get('telefono', '')}")
    lines.append(f"E-MAIL: {cliente.get('email', '')}")
    lines.append("")
    
    # PEC e altri dati
    lines.append(f"PEC: {cliente.get('pec', '')}")
    fatturato = cliente.get('valore_produzione', '')
    if fatturato:
        lines.append(f"FATTURATO: EUR {fatturato:,.0f}")
    else:
        lines.append("FATTURATO: ")
    lines.append(f"PARCO VEICOLI: {len(veicoli)}")
    lines.append(f"DIPENDENTI: {cliente.get('dipendenti', '')}")
    lines.append("")
    
    # Noleggiatori
    lines.append(f"CRM: ")
    lines.append(f"ARVAL: {veicoli_per_nol.get('ARVAL', '')}")
    lines.append(f"AYVENS: {veicoli_per_nol.get('AYVENS', '')}")
    lines.append(f"ALPHABET: {veicoli_per_nol.get('ALPHABET', '')}")
    lines.append(f"SANTANDER: {veicoli_per_nol.get('SANTANDER', '')}")
    lines.append(f"LEASYS: {veicoli_per_nol.get('LEASYS', '')}")
    lines.append("")
    
    content = "\n".join(lines)
    
    # Nome file pulito
    nome_file = re.sub(r'[^\w\s-]', '', cliente.get('nome_cliente', 'cliente')).strip()
    nome_file = re.sub(r'[-\s]+', '_', nome_file)
    
    response = Response(content, mimetype='text/plain; charset=utf-8')
    response.headers['Content-Disposition'] = f'attachment; filename="{nome_file}_evernote.txt"'
    return response


@app.route('/storico_pdf/<path:filename>')
def serve_storico_pdf(filename):
    """Serve i PDF - cerca prima nella nuova struttura clienti/, poi in storico_pdf/ (legacy)."""
    from flask import send_from_directory
    
    # CASO 1: Path completo (es: clienti/PIVA/00552060980/creditsafe/file.pdf)
    if filename.startswith("clienti/"):
        file_path = BASE_DIR / filename
        if file_path.exists():
            return send_from_directory(str(file_path.parent), file_path.name)
    
    # NUOVA STRUTTURA: cerca in clienti/PIVA/*/creditsafe/
    # Cerca in tutte le sottocartelle creditsafe
    for tipo_dir in [CLIENTI_PIVA_DIR, CLIENTI_CF_DIR]:
        if tipo_dir.exists():
            for cliente_dir in tipo_dir.iterdir():
                if cliente_dir.is_dir():
                    creditsafe_dir = cliente_dir / 'creditsafe'
                    if creditsafe_dir.exists() and (creditsafe_dir / filename).exists():
                        return send_from_directory(str(creditsafe_dir), filename)
    
    # LEGACY: cerca in storico_pdf (per retrocompatibilita)
    storico_dir = BASE_DIR / 'storico_pdf'
    
    if storico_dir.exists():
        # Cerca nella sottocartella della prima lettera
        if filename:
            prima_lettera = filename[0].upper()
            subdir = storico_dir / prima_lettera
            if subdir.exists() and (subdir / filename).exists():
                return send_from_directory(str(subdir), filename)
        
        # Cerca direttamente nella cartella storico_pdf
        if (storico_dir / filename).exists():
            return send_from_directory(str(storico_dir), filename)
    
    return "PDF non trovato", 404


@app.route('/clienti/<path:filepath>')
def serve_file_cliente(filepath):
    """Serve qualsiasi file dalla cartella clienti/."""
    from flask import send_from_directory
    
    file_path = CLIENTI_DIR / filepath
    
    if file_path.exists() and file_path.is_file():
        return send_from_directory(str(file_path.parent), file_path.name)
    
    return "File non trovato", 404


# ==============================================================================
# ROUTE: FLOTTA
# ==============================================================================

@app.route('/flotta')
@login_required
def flotta():
    """Dashboard flotta veicoli."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Calcola commerciali visibili (filtro subordinati)
    user_id = session.get('user_id')
    cursor.execute("SELECT ruolo_base FROM utenti WHERE id = ?", (user_id,))
    utente_row = cursor.fetchone()
    is_admin = utente_row and utente_row['ruolo_base'] == 'admin'
    
    if is_admin:
        commerciali_visibili = None  # Admin vede tutto
    else:
        commerciali_visibili = get_subordinati(conn, user_id)  # Propri + subordinati
    
    # Statistiche generali
    cursor.execute('SELECT COUNT(*) FROM veicoli')
    totale_veicoli = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(canone), 0) FROM veicoli')
    canone_totale = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(DISTINCT p_iva) FROM veicoli')
    totale_clienti = cursor.fetchone()[0]
    
    # Per noleggiatore
    cursor.execute('''
        SELECT noleggiatore, COUNT(*) as veicoli, COALESCE(SUM(canone), 0) as canone
        FROM veicoli
        WHERE noleggiatore IS NOT NULL AND noleggiatore != ''
        GROUP BY noleggiatore
        ORDER BY veicoli DESC
    ''')
    per_noleggiatore = [dict(row) for row in cursor.fetchall()]
    
    # Per commerciale
    cursor.execute('''
        SELECT COALESCE(commerciale, 'Non assegnato') as commerciale, 
               COUNT(*) as veicoli, 
               COALESCE(SUM(canone), 0) as canone
        FROM veicoli
        GROUP BY commerciale
        ORDER BY CASE WHEN commerciale IS NULL OR commerciale = '' THEN 1 ELSE 0 END, commerciale
    ''')
    per_commerciale = [dict(row) for row in cursor.fetchall()]
    
    # Filtri scadenze da parametri GET
    giorni_da = request.args.get('giorni_da', '0', type=str)
    giorni_a = request.args.get('giorni_a', '30', type=str)
    filtro_commerciale = request.args.get('commerciale_scad', '')
    
    # Converti in interi con default
    try:
        giorni_da = int(giorni_da) if giorni_da else 0
    except:
        giorni_da = 0
    try:
        giorni_a = int(giorni_a) if giorni_a else 30
    except:
        giorni_a = 30
    
    # Query scadenze con JOIN per dati cliente
    query_scadenze = '''
        SELECT v.*, 
               c.id as cliente_id_link,
               CAST(julianday(v.scadenza) - julianday('now') AS INTEGER) as giorni_rimasti
        FROM veicoli v
        LEFT JOIN clienti c ON v.cliente_id = c.id
        WHERE v.scadenza IS NOT NULL 
          AND julianday(v.scadenza) - julianday('now') BETWEEN ? AND ?
    '''
    params_scadenze = [giorni_da, giorni_a]
    
    # Filtro per subordinati (se non admin)
    if commerciali_visibili is not None:
        placeholders = ','.join('?' * len(commerciali_visibili))
        query_scadenze += f" AND v.commerciale_id IN ({placeholders})"
        params_scadenze.extend(commerciali_visibili)
    
    if filtro_commerciale:
        if filtro_commerciale == '__NULL__':
            query_scadenze += " AND (v.commerciale_id IS NULL OR v.commerciale_id = 0)"
        else:
            try:
                filtro_comm_id = int(filtro_commerciale)
                query_scadenze += " AND v.commerciale_id = ?"
                params_scadenze.append(filtro_comm_id)
            except (ValueError, TypeError):
                pass
    
    query_scadenze += " ORDER BY v.scadenza"
    
    cursor.execute(query_scadenze, params_scadenze)
    scadenze_prossime = [dict(row) for row in cursor.fetchall()]
    
    # Lista commerciali per dropdown filtro (solo chi ha ruolo commerciale)
    from app.gestione_commerciali import get_commerciali_tutti
    commerciali_lista = get_commerciali_tutti(conn, solo_attivi=True)
    
    # Filtri scadenze PASSATE da parametri GET
    giorni_pass_da = request.args.get('giorni_pass_da', '0', type=str)
    giorni_pass_a = request.args.get('giorni_pass_a', '-30', type=str)
    filtro_comm_pass = request.args.get('commerciale_pass', '')
    
    try:
        giorni_pass_da = int(giorni_pass_da) if giorni_pass_da else 0
    except:
        giorni_pass_da = 0
    try:
        giorni_pass_a = int(giorni_pass_a) if giorni_pass_a else -30
    except:
        giorni_pass_a = -30
    
    # Query scadenze PASSATE (giorni negativi)
    query_passate = '''
        SELECT v.*, 
               c.id as cliente_id_link,
               CAST(julianday(v.scadenza) - julianday('now') AS INTEGER) as giorni_rimasti
        FROM veicoli v
        LEFT JOIN clienti c ON v.cliente_id = c.id
        WHERE v.scadenza IS NOT NULL 
          AND julianday(v.scadenza) - julianday('now') BETWEEN ? AND ?
    '''
    # Inverti min/max per query corretta (es: da -30 a 0 diventa BETWEEN -30 AND 0)
    min_pass = min(giorni_pass_a, giorni_pass_da)
    max_pass = max(giorni_pass_a, giorni_pass_da)
    params_passate = [min_pass, max_pass]
    
    # Filtro per subordinati (se non admin)
    if commerciali_visibili is not None:
        placeholders = ','.join('?' * len(commerciali_visibili))
        query_passate += f" AND v.commerciale_id IN ({placeholders})"
        params_passate.extend(commerciali_visibili)
    
    if filtro_comm_pass:
        if filtro_comm_pass == '__NULL__':
            query_passate += " AND (v.commerciale_id IS NULL OR v.commerciale_id = 0)"
        else:
            try:
                filtro_comm_pass_id = int(filtro_comm_pass)
                query_passate += " AND v.commerciale_id = ?"
                params_passate.append(filtro_comm_pass_id)
            except (ValueError, TypeError):
                pass
    
    query_passate += " ORDER BY v.scadenza DESC"
    
    cursor.execute(query_passate, params_passate)
    scadenze_passate = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('flotta.html',
                         totale_veicoli=totale_veicoli,
                         canone_totale=canone_totale,
                         totale_clienti=totale_clienti,
                         per_noleggiatore=per_noleggiatore,
                         per_commerciale=per_commerciale,
                         scadenze_prossime=scadenze_prossime,
                         scadenze_passate=scadenze_passate,
                         commerciali_lista=commerciali_lista,
                         giorni_da=giorni_da,
                         giorni_a=giorni_a,
                         filtro_commerciale=filtro_commerciale,
                         giorni_pass_da=giorni_pass_da,
                         giorni_pass_a=giorni_pass_a,
                         filtro_comm_pass=filtro_comm_pass)


@app.route('/flotta/cerca')
def flotta_cerca():
    """Ricerca veicoli."""
    q = request.args.get('q', '').strip()
    
    if not q:
        return redirect(url_for('flotta'))
    
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT * FROM veicoli 
        WHERE targa LIKE ? OR marca LIKE ? OR modello LIKE ? 
              OR driver LIKE ? OR p_iva LIKE ?
        ORDER BY targa
    ''', tuple([f'%{q}%'] * 5))
    
    veicoli = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return render_template('flotta_risultati.html', veicoli=veicoli, query=q)


@app.route('/flotta/cliente/<path:nome_cliente>')
def flotta_cliente(nome_cliente):
    """Dettaglio cliente con veicoli."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Cerca veicoli del cliente (con conteggio note)
    cursor.execute('''
        SELECT v.*, 
               (SELECT COUNT(*) FROM note_veicoli WHERE veicolo_id = v.id AND eliminato = 0) as num_note
        FROM veicoli v
        WHERE v.nome_cliente = ? 
        ORDER BY v.scadenza
    ''', (nome_cliente,))
    veicoli = [dict(row) for row in cursor.fetchall()]
    
    if not veicoli:
        conn.close()
        return "Cliente non trovato", 404
    
    # Prendi P.IVA dal primo veicolo
    p_iva = veicoli[0].get('p_iva')
    
    # Cerca dati Creditsafe
    cliente_creditsafe = None
    if p_iva:
        cursor.execute('''
            SELECT * FROM clienti 
            WHERE REPLACE(REPLACE(UPPER(p_iva), 'IT', ''), ' ', '') LIKE ?
        ''', (f"%{p_iva}%",))
        row = cursor.fetchone()
        if row:
            cliente_creditsafe = dict(row)
    
    # Statistiche
    stats = {
        'totale_veicoli': len(veicoli),
        'canone_totale': sum(v.get('canone', 0) or 0 for v in veicoli),
        'noleggiatori': list(set(v.get('noleggiatore') for v in veicoli if v.get('noleggiatore'))),
        'commerciale': veicoli[0].get('commerciale'),
    }
    
    # Raggruppa per noleggiatore
    per_noleggiatore = {}
    for v in veicoli:
        nol = v.get('noleggiatore', 'Non specificato')
        if nol not in per_noleggiatore:
            per_noleggiatore[nol] = []
        per_noleggiatore[nol].append(v)
    
    conn.close()
    
    return render_template('flotta_cliente.html',
                         nome_cliente=nome_cliente,
                         veicoli=veicoli,
                         per_noleggiatore=per_noleggiatore,
                         stats=stats,
                         creditsafe=cliente_creditsafe,
                         p_iva=p_iva)


# ==============================================================================
# ROUTE: REPORT PER NOLEGGIATORE
# ==============================================================================

@app.route('/flotta/per-noleggiatore')
def flotta_per_noleggiatore():
    """Report raggruppato per noleggiatore."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Calcola commerciali visibili (filtro subordinati)
    user_id = session.get('user_id')
    cursor.execute("SELECT ruolo_base FROM utenti WHERE id = ?", (user_id,))
    utente_row = cursor.fetchone()
    is_admin = utente_row and utente_row['ruolo_base'] == 'admin'
    
    if is_admin:
        commerciali_visibili = None  # Admin vede tutto
    else:
        commerciali_visibili = get_subordinati(conn, user_id)  # Propri + subordinati
    
    # Riepilogo generale
    cursor.execute('''
        SELECT noleggiatore, 
               COUNT(*) as veicoli,
               COUNT(DISTINCT p_iva) as clienti,
               COALESCE(SUM(canone), 0) as canone
        FROM veicoli
        WHERE noleggiatore IS NOT NULL AND noleggiatore != ''
        GROUP BY noleggiatore
        ORDER BY veicoli DESC
    ''')
    
    riepilogo = []
    totale_veicoli = 0
    totale_canone = 0
    
    for row in cursor.fetchall():
        r = dict(row)
        totale_veicoli += r['veicoli']
        totale_canone += r['canone']
        riepilogo.append(r)
    
    # Calcola percentuali
    for r in riepilogo:
        r['percentuale'] = (r['canone'] / totale_canone * 100) if totale_canone > 0 else 0
    
    # Dettaglio per noleggiatore
    dettaglio = {}
    for r in riepilogo:
        nol = r['noleggiatore']
        
        # Query base con commerciale_id
        query_det = '''
            SELECT p_iva, 
                   MAX(NOME_CLIENTE) as nome,
                   COUNT(*) as veicoli,
                   COALESCE(SUM(canone), 0) as canone,
                   commerciale_id
            FROM veicoli
            WHERE noleggiatore = ?
        '''
        params_det = [nol]
        
        # Filtro subordinati (se non admin)
        if commerciali_visibili is not None:
            placeholders = ','.join('?' * len(commerciali_visibili))
            query_det += f" AND commerciale_id IN ({placeholders})"
            params_det.extend(commerciali_visibili)
        
        query_det += " GROUP BY p_iva ORDER BY nome"
        
        cursor.execute(query_det, params_det)
        clienti_nol = []
        for row in cursor.fetchall():
            c = dict(row)
            c['commerciale_display'] = get_commerciale_display(conn, c.get('commerciale_id'))
            clienti_nol.append(c)
        
        dettaglio[nol] = {
            'clienti': clienti_nol,
            'veicoli': r['veicoli'],
            'canone': r['canone'],
        }
    
    conn.close()
    
    return render_template('flotta_noleggiatore.html',
                         riepilogo=riepilogo,
                         dettaglio=dettaglio,
                         totale_veicoli=totale_veicoli,
                         totale_canone=totale_canone)


# ==============================================================================

# ==============================================================================
# ROUTE: GESTIONE COMMERCIALI
# ==============================================================================

# @app.route('/flotta/gestione-commerciali')  # DISATTIVATO
# @login_required
def flotta_gestione_commerciali_OLD():  # DISATTIVATO - usa blueprint  # DISATTIVATO - usa blueprint
    """Pagina gestione massiva assegnazione commerciali."""
    from app.database_utenti import ha_permesso
    from flask import session
    
    # Verifica permesso
    conn_check = get_connection()
    user_id = session.get('user_id')
    if not ha_permesso(conn_check, user_id, 'flotta_assegnazioni'):
        conn_check.close()
        flash('Non hai il permesso per accedere a questa pagina.', 'danger')
        return redirect(url_for('flotta_per_commerciale'))
    conn_check.close()
    
    da = request.args.get('da', None)
    q = request.args.get('q', '')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Lista commerciali dal DB utenti (quelli abilitati in flotta)
    from app.database_utenti import get_utenti_commerciali, get_connection as get_conn_utenti
    conn_utenti = get_conn_utenti()
    utenti_commerciali = get_utenti_commerciali(conn_utenti)
    conn_utenti.close()
    commerciali = [u['nome_commerciale_flotta'] for u in utenti_commerciali]
    
    # Costruisci filtro
    where_parts = ["1=1"]
    params = []
    
    if da == '__NULL__' or da == '':
        where_parts.append("(commerciale IS NULL OR commerciale = '')")
    elif da:
        where_parts.append("commerciale = ?")
        params.append(da)
    
    if q:
        where_parts.append("NOME_CLIENTE LIKE ?")
        params.append(f'%{q}%')
    
    where_clause = " AND ".join(where_parts)
    
    # Lista clienti (con p_iva per link scheda)
    cursor.execute(f'''
        SELECT NOME_CLIENTE as nome,
               p_iva,
               COUNT(*) as veicoli, 
               COALESCE(SUM(canone), 0) as canone,
               commerciale,
               GROUP_CONCAT(DISTINCT noleggiatore) as noleggiatori
        FROM veicoli
        WHERE {where_clause}
        GROUP BY NOME_CLIENTE, p_iva
        ORDER BY NOME_CLIENTE
    ''', params)
    
    clienti = []
    for row in cursor.fetchall():
        clienti.append({
            'nome': row['nome'],
            'p_iva': row['p_iva'],
            'veicoli': row['veicoli'],
            'canone': row['canone'],
            'commerciale': row['commerciale'],
            'noleggiatori': row['noleggiatori'].split(',') if row['noleggiatori'] else []
        })
    
    conn.close()
    
    return render_template('flotta_gestione_commerciali.html',
                         commerciali=commerciali,
                         clienti=clienti,
                         da=da,
                         q=q)


@app.route('/flotta/assegna-commerciali', methods=['POST'])
@login_required
def flotta_assegna_commerciali():
    """Assegna commerciale a piu clienti con registrazione storico."""
    from app.database_utenti import registra_assegnazione, ha_permesso
    
    # Verifica permesso
    conn_utenti = get_connection()
    user_id = session.get('user_id')
    if not ha_permesso(conn_utenti, user_id, 'flotta_assegnazioni'):
        conn_utenti.close()
        flash('Non hai il permesso per gestire le assegnazioni.', 'danger')
        return redirect(url_for('flotta_per_commerciale'))
    conn_utenti.close()
    
    clienti_selezionati = request.form.getlist('clienti[]')
    nuovo_commerciale = request.form.get('nuovo_commerciale', '').strip().upper()
    
    if not clienti_selezionati:
        return redirect(url_for('flotta_gestione_commerciali', error='Nessun cliente selezionato'))
    
    if not nuovo_commerciale:
        return redirect(url_for('flotta_gestione_commerciali', error='Commerciale non specificato'))
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Aggiorna tutti i veicoli dei clienti selezionati e registra storico
    totale_aggiornati = 0
    for cliente in clienti_selezionati:
        # Recupera commerciale precedente e P.IVA
        cursor.execute('''
            SELECT DISTINCT commerciale, PIVA FROM veicoli 
            WHERE NOME_CLIENTE = ? LIMIT 1
        ''', (cliente,))
        row = cursor.fetchone()
        commerciale_precedente = row['commerciale'] if row else None
        cliente_piva = row['PIVA'] if row else None
        
        # Aggiorna veicoli
        cursor.execute('''
            UPDATE veicoli SET commerciale = ? WHERE NOME_CLIENTE = ?
        ''', (nuovo_commerciale, cliente))
        totale_aggiornati += cursor.rowcount
        
        # Registra nello storico (solo se c'e un cambiamento effettivo)
        if commerciale_precedente != nuovo_commerciale:
            registra_assegnazione(
                conn, 
                cliente_nome=cliente,
                cliente_piva=cliente_piva,
                commerciale_precedente=commerciale_precedente,
                commerciale_nuovo=nuovo_commerciale,
                utente_id=user_id,
                note=None
            )
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('flotta_gestione_commerciali', 
                          success=f'Aggiornati {totale_aggiornati} veicoli di {len(clienti_selezionati)} clienti a {nuovo_commerciale}'))


# ==============================================================================
# ROUTE: STATISTICHE
# ==============================================================================

@app.route('/statistiche')
@login_required
def statistiche():
    """Pagina statistiche generali."""
    conn = get_connection()
    cursor = conn.cursor()
    stats = get_statistiche_generali(conn)
    
    # Province con veicoli
    cursor.execute('''
        SELECT province_code, COUNT(*) as n 
        FROM veicoli 
        WHERE province_code IS NOT NULL AND province_code != ''
        GROUP BY province_code 
        ORDER BY n DESC
    ''')
    veicoli_per_prov = {r[0]: r[1] for r in cursor.fetchall()}
    
    # Aziende per provincia (clienti con P.IVA = aziende, esclusi privati con solo CF)
    cursor.execute('''
        SELECT v.province_code, COUNT(DISTINCT v.cliente_id) as n
        FROM veicoli v
        JOIN clienti c ON v.cliente_id = c.id
        WHERE v.province_code IS NOT NULL AND v.province_code != ''
          AND c.p_iva IS NOT NULL AND c.p_iva != ''
        GROUP BY v.province_code
    ''')
    aziende_per_prov = {r[0]: r[1] for r in cursor.fetchall()}
    
    # Combina province (tutte quelle con veicoli o aziende)
    tutte_province = set(veicoli_per_prov.keys()) | set(aziende_per_prov.keys())
    stats['province'] = sorted([
        {
            "sigla": p,
            "aziende": aziende_per_prov.get(p, 0),
            "veicoli": veicoli_per_prov.get(p, 0)
        }
        for p in tutte_province
    ], key=lambda x: (-x['aziende'], -x['veicoli']))
    
    conn.close()
    return render_template('statistiche.html', stats=stats)


# ==============================================================================
# ROUTE: AMMINISTRAZIONE
# ==============================================================================

@app.route('/admin')
@login_required
def admin():
    """Pannello amministrazione."""
    import subprocess
    
    # Conta file in cartelle
    pdf_da_elaborare = len(list(PDF_DIR.glob('*.pdf'))) if PDF_DIR.exists() else 0
    
    # Conta PDF nella nuova struttura clienti/ (creditsafe)
    pdf_in_storico = 0
    if CLIENTI_DIR.exists():
        pdf_in_storico = sum(1 for _ in CLIENTI_DIR.rglob('creditsafe/*.pdf'))
    # Fallback: conta anche in storico_pdf se esiste (legacy)
    storico_legacy = BASE_DIR / 'storico_pdf'
    if storico_legacy.exists():
        pdf_in_storico += sum(1 for _ in storico_legacy.rglob('*.pdf'))
    
    log_files = len(list(LOGS_DIR.glob('*.log'))) if LOGS_DIR.exists() else 0
    
    # PDF con errori
    PDF_ERRORI_DIR = BASE_DIR / 'pdf_errori'
    pdf_errori = len(list(PDF_ERRORI_DIR.glob('*.pdf'))) if PDF_ERRORI_DIR.exists() else 0
    
    # Dimensione DB
    db_size = DB_FILE.stat().st_size / 1024 / 1024 if DB_FILE.exists() else 0
    
    # Spazio disco totale programma (cartella gestione_flotta)
    try:
        result = subprocess.run(['du', '-sb', str(BASE_DIR)], capture_output=True, text=True)
        spazio_disco_bytes = int(result.stdout.split()[0])
        spazio_disco_mb = spazio_disco_bytes / 1024 / 1024
    except:
        spazio_disco_mb = 0
    
    # RAM processo Flask
    try:
        process = psutil.Process()
        ram_mb = process.memory_info().rss / 1024 / 1024
        pid = process.pid
    except:
        ram_mb = 0
        pid = 0
    
    return render_template('admin.html',
                         pdf_da_elaborare=pdf_da_elaborare,
                         pdf_in_storico=pdf_in_storico,
                         pdf_errori=pdf_errori,
                         log_files=log_files,
                         db_size=db_size,
                         spazio_disco_mb=spazio_disco_mb,
                         ram_mb=ram_mb,
                         pid=pid)


@app.route('/admin/import-pdf', methods=['POST'])
def admin_import_pdf():
    """Esegue import PDF manuale."""
    risultato = importa_tutti_pdf()
    return redirect(url_for('admin', 
                          msg=f"Import completato: {risultato['elaborati']} elaborati, {risultato['errori']} errori"))


@app.route('/admin/pulisci-log', methods=['POST'])
def admin_pulisci_log():
    """Pulisce log vecchi."""
    rimossi = pulisci_log_vecchi()
    return redirect(url_for('admin', msg=f"Rimossi {rimossi} file di log"))


# Variabile globale per stato import (per progress bar)
import_status = {
    'running': False,
    'current': 0,
    'total': 0,
    'current_file': '',
    'completed': [],
    'errors': []
}


@app.route('/admin/upload-pdf', methods=['POST'])
def admin_upload_pdf():
    """Upload PDF via drag&drop."""
    if 'files[]' not in request.files:
        return jsonify({'success': False, 'error': 'Nessun file ricevuto'}), 400
    
    files = request.files.getlist('files[]')
    uploaded = []
    errors = []
    
    for file in files:
        if file.filename == '':
            continue
        if not file.filename.lower().endswith('.pdf'):
            errors.append(f"{file.filename}: non e un PDF")
            continue
        
        try:
            # Salva in cartella pdf/
            filepath = PDF_DIR / file.filename
            # Se esiste gia, aggiungi suffisso
            if filepath.exists():
                base = filepath.stem
                i = 1
                while filepath.exists():
                    filepath = PDF_DIR / f"{base}_{i}.pdf"
                    i += 1
            
            file.save(str(filepath))
            uploaded.append(file.filename)
            logger.info(f"Upload PDF: {filepath.name}")
        except Exception as e:
            errors.append(f"{file.filename}: {str(e)}")
            logger.error(f"Errore upload {file.filename}: {e}")
    
    return jsonify({
        'success': True,
        'uploaded': uploaded,
        'errors': errors,
        'count': len(uploaded)
    })


@app.route('/admin/import-pdf-async', methods=['POST'])
def admin_import_pdf_async():
    """Avvia import PDF asincrono con progress."""
    global import_status
    
    if import_status['running']:
        return jsonify({'success': False, 'error': 'Import gia in corso'}), 400
    
    # Lista PDF da elaborare
    pdf_files = list(PDF_DIR.glob('*.pdf'))
    if not pdf_files:
        return jsonify({'success': False, 'error': 'Nessun PDF da elaborare'}), 400
    
    # Reset stato
    import_status = {
        'running': True,
        'current': 0,
        'total': len(pdf_files),
        'current_file': '',
        'completed': [],
        'errors': []
    }
    
    def run_import():
        global import_status
        try:
            from app.import_creditsafe import importa_pdf_singolo
            
            for i, pdf_path in enumerate(pdf_files):
                import_status['current'] = i + 1
                import_status['current_file'] = pdf_path.name
                
                try:
                    risultato = importa_pdf_singolo(str(pdf_path))
                    if risultato.get('success'):
                        import_status['completed'].append(pdf_path.name)
                    else:
                        import_status['errors'].append(f"{pdf_path.name}: {risultato.get('error', 'Errore sconosciuto')}")
                except Exception as e:
                    import_status['errors'].append(f"{pdf_path.name}: {str(e)}")
            
        except Exception as e:
            import_status['errors'].append(f"Errore generale: {str(e)}")
        finally:
            import_status['running'] = False
            import_status['current_file'] = ''
    
    # Avvia thread
    thread = threading.Thread(target=run_import)
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'success': True,
        'total': len(pdf_files),
        'message': 'Import avviato'
    })


@app.route('/admin/import-status')
def admin_import_status():
    """Restituisce stato import per polling."""
    global import_status
    return jsonify(import_status)


@app.route('/admin/crontab', methods=['GET', 'POST'])
def admin_crontab():
    """Legge o scrive crontab utente."""
    import subprocess
    import tempfile
    
    if request.method == 'GET':
        # Leggi crontab attuale
        try:
            result = subprocess.run(['crontab', '-l'], capture_output=True, text=True)
            if result.returncode == 0:
                return jsonify({'success': True, 'crontab': result.stdout})
            else:
                # Nessun crontab
                return jsonify({'success': True, 'crontab': ''})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})
    
    else:  # POST
        # Scrivi nuovo crontab
        new_crontab = request.json.get('crontab', '')
        
        try:
            # Scrivi in file temporaneo
            with tempfile.NamedTemporaryFile(mode='w', suffix='.cron', delete=False) as f:
                f.write(new_crontab)
                temp_path = f.name
            
            # Installa crontab
            result = subprocess.run(['crontab', temp_path], capture_output=True, text=True)
            
            # Rimuovi file temporaneo
            Path(temp_path).unlink()
            
            if result.returncode == 0:
                logger.info("Crontab aggiornato")
                return jsonify({'success': True, 'message': 'Crontab aggiornato'})
            else:
                return jsonify({'success': False, 'error': result.stderr})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})


# ==============================================================================
# ROUTE: EXPORT
# ==============================================================================

@app.route('/export/excel')
def export_excel():
    """Esporta lista clienti in Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl non installato", 500
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM clienti ORDER BY nome_cliente')
    clienti = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clienti"
    
    # Header
    headers = ['Nome', 'P.IVA', 'C.F.', 'Score', 'Credito', 'Indirizzo', 'Telefono', 'PEC', 'Commerciale']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    
    # Dati
    for row, c in enumerate(clienti, 2):
        ws.cell(row=row, column=1, value=c.get('nome_cliente'))
        ws.cell(row=row, column=2, value=c.get('p_iva'))
        ws.cell(row=row, column=3, value=c.get('cod_fiscale'))
        ws.cell(row=row, column=4, value=c.get('score'))
        ws.cell(row=row, column=5, value=c.get('credito'))
        ws.cell(row=row, column=6, value=c.get('indirizzo'))
        ws.cell(row=row, column=7, value=c.get('telefono'))
        ws.cell(row=row, column=8, value=c.get('pec'))
        ws.cell(row=row, column=9, value=c.get('commerciale'))
    
    # Salva
    output_path = LOGS_DIR / f"export_clienti_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(output_path))
    
    return send_file(str(output_path), as_attachment=True, download_name='clienti.xlsx')


# ==============================================================================
# ROUTE: GESTIONE VEICOLO SINGOLO
# ==============================================================================

@app.route('/veicolo/<int:veicolo_id>')
def gestione_veicolo(veicolo_id):
    """Pagina gestione veicolo singolo con simulazione km."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Recupera veicolo
    cursor.execute('SELECT * FROM veicoli WHERE id = ?', (veicolo_id,))
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return "Veicolo non trovato", 404
    
    veicolo = dict(row)
    
    # Calcola km totali contratto (logica per noleggiatore)
    km_base = veicolo.get('km') or 0
    km_franchigia = veicolo.get('km_franchigia') or 0
    
    if veicolo.get('noleggiatore') == 'LEASYS':
        km_totali_contratto = km_base + km_franchigia
    else:
        km_totali_contratto = km_base
    
    # Calcola km/mese
    durata = veicolo.get('durata') or 1
    km_mese = km_totali_contratto // durata if durata > 0 else 0
    
    # Giorni rimasti
    giorni_rimasti = None
    if veicolo.get('scadenza'):
        giorni_rimasti = giorni_mancanti(veicolo['scadenza'])
    
    # Recupera colore HEX, link assistenza e link denuncia dal file Excel
    noleggiatore = (veicolo.get('noleggiatore') or '').upper()
    noleggiatore_color = '#6c757d'  # default grigio chiaro HEX se non trovato
    link_assistenza = None
    link_denuncia = None
    link_restituzione = None
    try:
        import openpyxl
        if NOLEGGIATORI_ASSISTENZA_FILE.exists():
            wb = openpyxl.load_workbook(str(NOLEGGIATORI_ASSISTENZA_FILE), read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[0].strip().upper() == noleggiatore:
                    link_assistenza = row[1] if len(row) > 1 and row[1] else None
                    noleggiatore_color = row[2] if len(row) > 2 and row[2] else '#6c757d'
                    link_denuncia = row[3] if len(row) > 3 and row[3] else None
                    link_restituzione = row[4] if len(row) > 4 and row[4] else None
                    break
            wb.close()
    except Exception as e:
        logger.warning(f"Errore lettura file assistenza: {e}")
    
    # Storico km
    cursor.execute('''
        SELECT * FROM storico_km 
        WHERE veicolo_id = ? 
        ORDER BY data_rilevazione DESC
    ''', (veicolo_id,))
    storico_km = [dict(r) for r in cursor.fetchall()]
    
    # Note veicolo (solo non eliminate)
    cursor.execute('''
        SELECT * FROM note_veicoli 
        WHERE veicolo_id = ? AND eliminato = 0
        ORDER BY fissata DESC, data_creazione DESC
    ''', (veicolo_id,))
    note_veicolo = [dict(r) for r in cursor.fetchall()]
    
    conn.close()
    
    return render_template('veicolo.html',
                         veicolo=veicolo,
                         km_totali_contratto=km_totali_contratto,
                         km_mese=km_mese,
                         giorni_rimasti=giorni_rimasti,
                         noleggiatore_color=noleggiatore_color,
                         link_assistenza=link_assistenza,
                         link_denuncia=link_denuncia,
                         storico_km=storico_km,
                         note_veicolo=note_veicolo,
                         link_restituzione=link_restituzione,
                         oggi=datetime.now().strftime('%Y-%m-%d'))


@app.route('/veicolo/<int:veicolo_id>/costi-km', methods=['POST'])
def aggiorna_costi_km(veicolo_id):
    """Aggiorna i costi km di un veicolo."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Verifica veicolo esiste
    cursor.execute('SELECT id FROM veicoli WHERE id = ?', (veicolo_id,))
    if not cursor.fetchone():
        conn.close()
        return "Veicolo non trovato", 404
    
    # Recupera dati form
    tipo_sistema = request.form.get('tipo_sistema', 'lineare')
    
    if tipo_sistema == 'scaglioni':
        soglia = request.form.get('soglia_scaglione')
        soglia = float(soglia) if soglia else 15
    else:
        soglia = None
    
    costo_extra_1 = request.form.get('costo_km_extra_1')
    costo_extra_2 = request.form.get('costo_km_extra_2') if tipo_sistema == 'scaglioni' else None
    costo_rimb_1 = request.form.get('costo_km_rimborso_1')
    costo_rimb_2 = request.form.get('costo_km_rimborso_2') if tipo_sistema == 'scaglioni' else None
    
    # Converti in float o None
    def to_float(v):
        try:
            return float(v) if v else None
        except:
            return None
    
    cursor.execute('''
        UPDATE veicoli SET
            soglia_scaglione = ?,
            costo_km_extra_1 = ?,
            costo_km_extra_2 = ?,
            costo_km_rimborso_1 = ?,
            costo_km_rimborso_2 = ?
        WHERE id = ?
    ''', (
        soglia,
        to_float(costo_extra_1),
        to_float(costo_extra_2),
        to_float(costo_rimb_1),
        to_float(costo_rimb_2),
        veicolo_id
    ))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))


@app.route('/veicolo/<int:veicolo_id>/driver', methods=['POST'])
def aggiorna_driver(veicolo_id):
    """Aggiorna il driver di un veicolo (nome, telefono, email)."""
    conn = get_connection()
    cursor = conn.cursor()
    
    driver = request.form.get('driver', '').strip()
    driver_telefono = request.form.get('driver_telefono', '').strip()
    driver_email = request.form.get('driver_email', '').strip()
    
    cursor.execute('''
        UPDATE veicoli SET 
            driver = ?,
            driver_telefono = ?,
            driver_email = ?
        WHERE id = ?
    ''', (driver or None, driver_telefono or None, driver_email or None, veicolo_id))
    conn.commit()
    conn.close()
    
    return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))


# ==============================================================================
# ROUTE: GESTIONE NOTE VEICOLI
# ==============================================================================

@app.route('/veicolo/<int:veicolo_id>/nota/nuova', methods=['POST'])
@login_required
def nuova_nota_veicolo(veicolo_id):
    """Aggiunge una nuova nota al veicolo."""
    conn = get_connection()
    cursor = conn.cursor()
    
    testo = request.form.get('testo', '').strip()
    
    # Autore automatico da utente loggato
    user_id = session.get('user_id')
    autore = session.get('cognome', session.get('username', 'Sistema'))
    
    if not testo:
        conn.close()
        return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        INSERT INTO note_veicoli (veicolo_id, testo, autore, data_creazione, creato_da_id)
        VALUES (?, ?, ?, ?, ?)
    ''', (veicolo_id, testo, autore, now, user_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))


@app.route('/veicolo/<int:veicolo_id>/nota/modifica', methods=['POST'])
@login_required
def modifica_nota_veicolo(veicolo_id):
    """Modifica una nota esistente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    nota_id = request.form.get('nota_id')
    testo = request.form.get('testo', '').strip()
    
    # Chi modifica (utente loggato)
    user_id = session.get('user_id')
    
    if not nota_id or not testo:
        conn.close()
        return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Non modifica autore originale, solo modificato_da_id
    cursor.execute('''
        UPDATE note_veicoli 
        SET testo = ?, data_modifica = ?, modificato_da_id = ?
        WHERE id = ? AND veicolo_id = ? AND eliminato = 0
    ''', (testo, now, user_id, nota_id, veicolo_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))


@app.route('/veicolo/<int:veicolo_id>/nota/elimina', methods=['POST'])
def elimina_nota_veicolo(veicolo_id):
    """Elimina (soft delete) una nota."""
    conn = get_connection()
    cursor = conn.cursor()
    
    nota_id = request.form.get('nota_id')
    username = session.get('cognome', session.get('username', 'Sistema'))
    
    if not nota_id:
        conn.close()
        return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Soft delete: imposta flag eliminato e data
    cursor.execute('''
        UPDATE note_veicoli 
        SET eliminato = 1, data_eliminazione = ?, eliminato_da = ?
        WHERE id = ? AND veicolo_id = ?
    ''', (now, username, nota_id, veicolo_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))


@app.route('/veicolo/<int:veicolo_id>/salva-km', methods=['POST'])
def salva_rilevazione_km(veicolo_id):
    """Salva una rilevazione km (chiamata AJAX)."""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        data = request.get_json()
        km = data.get('km')
        data_rilevazione = data.get('data')
        note = data.get('note', '')
        
        if not km or not data_rilevazione:
            return jsonify({'success': False, 'error': 'Dati mancanti'})
        
        # Verifica veicolo esiste
        cursor.execute('SELECT id FROM veicoli WHERE id = ?', (veicolo_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Veicolo non trovato'})
        
        # Aggiorna km_attuali nel veicolo
        cursor.execute('''
            UPDATE veicoli SET 
                km_attuali = ?,
                data_rilevazione_km = ?
            WHERE id = ?
        ''', (km, data_rilevazione, veicolo_id))
        
        # Inserisci in storico
        cursor.execute('''
            INSERT INTO storico_km (veicolo_id, km_rilevati, data_rilevazione, note)
            VALUES (?, ?, ?, ?)
        ''', (veicolo_id, km, data_rilevazione, note))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/veicolo/<int:veicolo_id>/salva-targa', methods=['POST'])
def salva_targa_veicolo(veicolo_id):
    """Salva la targa di un veicolo (operazione una tantum, non modificabile)."""
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        data = request.get_json()
        targa = data.get('targa', '').strip().upper()
        
        if not targa:
            return jsonify({'success': False, 'error': 'Targa non valida'})
        
        # Verifica che il veicolo esista e non abbia gia una targa
        cursor.execute('SELECT id, targa FROM veicoli WHERE id = ?', (veicolo_id,))
        veicolo = cursor.fetchone()
        
        if not veicolo:
            return jsonify({'success': False, 'error': 'Veicolo non trovato'})
        
        if veicolo['targa']:
            return jsonify({'success': False, 'error': 'Questo veicolo ha gia una targa assegnata'})
        
        # Verifica che la targa non sia gia usata da un altro veicolo
        cursor.execute('SELECT id FROM veicoli WHERE targa = ? AND id != ?', (targa, veicolo_id))
        if cursor.fetchone():
            return jsonify({'success': False, 'error': 'Questa targa e gia assegnata a un altro veicolo'})
        
        # Salva la targa
        cursor.execute('UPDATE veicoli SET targa = ? WHERE id = ?', (targa, veicolo_id))
        conn.commit()
        conn.close()
        
        logger.info(f"Targa {targa} salvata per veicolo ID {veicolo_id}")
        return jsonify({'success': True})
        
    except Exception as e:
        conn.close()
        logger.error(f"Errore salvataggio targa: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/veicolo/<int:veicolo_id>/franchigia-km', methods=['POST'])
def aggiorna_franchigia_km(veicolo_id):
    """Aggiorna la franchigia km di un veicolo LEASYS."""
    conn = get_connection()
    cursor = conn.cursor()
    
    km_franchigia = request.form.get('km_franchigia', '0')
    
    try:
        km_franchigia = int(km_franchigia)
        if km_franchigia < 0:
            km_franchigia = 0
    except ValueError:
        km_franchigia = 0
    
    cursor.execute('''
        UPDATE veicoli SET km_franchigia = ?
        WHERE id = ?
    ''', (km_franchigia, veicolo_id))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Franchigia km aggiornata a {km_franchigia} per veicolo ID {veicolo_id}")
    return redirect(url_for('gestione_veicolo', veicolo_id=veicolo_id))


@app.route('/api/noleggiatore-assistenza/<noleggiatore>')
def get_link_assistenza(noleggiatore):
    """Ritorna il link assistenza per un noleggiatore (da file Excel)."""
    try:
        import openpyxl
        
        if not NOLEGGIATORI_ASSISTENZA_FILE.exists():
            return jsonify({'error': 'File configurazione non trovato', 'url': None})
        
        wb = openpyxl.load_workbook(str(NOLEGGIATORI_ASSISTENZA_FILE), read_only=True)
        ws = wb.active
        
        noleggiatore_upper = (noleggiatore or '').strip().upper()
        url_assistenza = None
        
        # Cerca il noleggiatore nel file (salta header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].strip().upper() == noleggiatore_upper:
                url_assistenza = row[1] if len(row) > 1 else None
                break
        
        wb.close()
        
        return jsonify({
            'noleggiatore': noleggiatore_upper,
            'url': url_assistenza if url_assistenza else None
        })
        
    except Exception as e:
        logger.error(f"Errore lettura link assistenza: {e}")
        return jsonify({'error': str(e), 'url': None})



@app.route('/cliente/<int:cliente_id>/note/cerca')
def cerca_note_cliente(cliente_id):
    """Cerca nelle note di un cliente."""
    q = request.args.get('q', '').strip()
    
    conn = get_connection()
    cursor = conn.cursor()
    
    if q:
        # Ricerca in titolo e testo
        search_param = f'%{q}%'
        cursor.execute('''
            SELECT * FROM note_clienti 
            WHERE cliente_id = ? AND eliminato = 0
              AND (titolo LIKE ? OR testo LIKE ?)
            ORDER BY fissata DESC, data_creazione DESC
        ''', (cliente_id, search_param, search_param))
    else:
        cursor.execute('''
            SELECT * FROM note_clienti 
            WHERE cliente_id = ? AND eliminato = 0
            ORDER BY fissata DESC, data_creazione DESC
        ''', (cliente_id,))
    
    note = [dict(row) for row in cursor.fetchall()]
    
    # Carica allegati per ogni nota
    for nota in note:
        cursor.execute('''
            SELECT * FROM allegati_note 
            WHERE nota_cliente_id = ?
            ORDER BY data_upload
        ''', (nota['id'],))
        nota['allegati'] = [dict(a) for a in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({'note': note, 'query': q, 'count': len(note)})


@app.route('/cliente/<int:cliente_id>/allegato/elimina', methods=['POST'])
def elimina_allegato_cliente(cliente_id):
    """Elimina un singolo allegato di una nota cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    allegato_id = request.form.get('allegato_id')
    
    if not allegato_id:
        conn.close()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'ID mancante'})
        return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))
    
    # Recupera percorso file
    cursor.execute('SELECT percorso FROM allegati_note WHERE id = ?', (allegato_id,))
    allegato = cursor.fetchone()
    
    if allegato:
        # Elimina file fisico
        try:
            percorso = Path(allegato['percorso'])
            if percorso.exists():
                percorso.unlink()
                cartella = percorso.parent
                if cartella.exists() and not any(cartella.iterdir()):
                    cartella.rmdir()
        except Exception as e:
            print(f"Errore eliminazione file: {e}")
        
        # Elimina record dal DB
        cursor.execute('DELETE FROM allegati_note WHERE id = ?', (allegato_id,))
        conn.commit()
    
    conn.close()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True})
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))


@app.route('/allegato/cliente/<int:allegato_id>')
def download_allegato_cliente(allegato_id):
    """Scarica un allegato di una nota cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM allegati_note WHERE id = ?', (allegato_id,))
    allegato = cursor.fetchone()
    conn.close()
    
    if not allegato:
        return "Allegato non trovato", 404
    
    allegato = dict(allegato)
    percorso = Path(allegato['percorso'])
    
    if not percorso.exists():
        return "File non trovato", 404
    
    return send_file(str(percorso), 
                     download_name=allegato['nome_originale'],
                     as_attachment=True)


# ==============================================================================
# MAIN


# ==============================================================================
# ROUTE: GESTIONE REFERENTI CLIENTI
# ==============================================================================

@app.route('/cliente/<int:cliente_id>/referente/nuovo', methods=['POST'])
def nuovo_referente(cliente_id):
    """Aggiunge un nuovo referente al cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    ruolo = request.form.get('ruolo', '').strip()
    nome = request.form.get('nome', '').strip()
    cognome = request.form.get('cognome', '').strip()
    telefono = request.form.get('telefono', '').strip()
    interno = request.form.get('interno', '').strip()
    cellulare = request.form.get('cellulare', '').strip()
    email_principale = request.form.get('email_principale', '').strip()
    email_secondarie = request.form.get('email_secondarie', '').strip()
    linkedin = request.form.get('linkedin', '').strip()
    note = request.form.get('note', '').strip()
    principale = 1 if request.form.get('principale') else 0
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Se questo e principale, rimuovi flag da altri
    if principale:
        cursor.execute('UPDATE referenti_clienti SET principale = 0 WHERE cliente_id = ?', (cliente_id,))
    
    cursor.execute('''
        INSERT INTO referenti_clienti 
        (cliente_id, ruolo, nome, cognome, telefono, interno, cellulare, 
         email_principale, email_secondarie, linkedin, note, principale, data_creazione)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (cliente_id, ruolo or None, nome or None, cognome or None, telefono or None,
           interno or None, cellulare or None, email_principale or None, 
           email_secondarie or None, linkedin or None, note or None, principale, now))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))


@app.route('/cliente/<int:cliente_id>/referente/<int:referente_id>/modifica', methods=['POST'])
def modifica_referente(cliente_id, referente_id):
    """Modifica un referente esistente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    ruolo = request.form.get('ruolo', '').strip()
    nome = request.form.get('nome', '').strip()
    cognome = request.form.get('cognome', '').strip()
    telefono = request.form.get('telefono', '').strip()
    interno = request.form.get('interno', '').strip()
    cellulare = request.form.get('cellulare', '').strip()
    email_principale = request.form.get('email_principale', '').strip()
    email_secondarie = request.form.get('email_secondarie', '').strip()
    linkedin = request.form.get('linkedin', '').strip()
    note = request.form.get('note', '').strip()
    principale = 1 if request.form.get('principale') else 0
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Se questo e principale, rimuovi flag da altri
    if principale:
        cursor.execute('UPDATE referenti_clienti SET principale = 0 WHERE cliente_id = ? AND id != ?', 
                      (cliente_id, referente_id))
    
    cursor.execute('''
        UPDATE referenti_clienti SET
            ruolo = ?, nome = ?, cognome = ?, telefono = ?, interno = ?,
            cellulare = ?, email_principale = ?, email_secondarie = ?,
            linkedin = ?, note = ?, principale = ?, data_modifica = ?
        WHERE id = ? AND cliente_id = ?
    ''', (ruolo or None, nome or None, cognome or None, telefono or None,
           interno or None, cellulare or None, email_principale or None,
           email_secondarie or None, linkedin or None, note or None, 
           principale, now, referente_id, cliente_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))


@app.route('/cliente/<int:cliente_id>/referente/<int:referente_id>/elimina', methods=['POST'])
def elimina_referente(cliente_id, referente_id):
    """Elimina un referente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('DELETE FROM referenti_clienti WHERE id = ? AND cliente_id = ?', 
                  (referente_id, cliente_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))


@app.route('/cliente/<int:cliente_id>/contatti/modifica', methods=['POST'])
def modifica_contatti_cliente(cliente_id):
    """Modifica i contatti generali del cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    telefono = request.form.get('telefono', '').strip()
    pec = request.form.get('pec', '').strip()
    email = request.form.get('email', '').strip()
    sito_web = request.form.get('sito_web', '').strip()
    indirizzo_alternativo = request.form.get('indirizzo_alternativo', '').strip()
    legale_rappresentante = request.form.get('legale_rappresentante', '').strip()
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        UPDATE clienti SET
            telefono = ?, pec = ?, email = ?, sito_web = ?, indirizzo_alternativo = ?, legale_rappresentante = ?,
            data_ultimo_aggiornamento = ?
        WHERE id = ?
    ''', (telefono or None, pec or None, email or None, sito_web or None,
           indirizzo_alternativo or None, legale_rappresentante or None, now, cliente_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))


@app.route('/cliente/<int:cliente_id>/indirizzo/modifica', methods=['POST'])
def modifica_indirizzo_cliente(cliente_id):
    """
    Modifica l'indirizzo del cliente.
    Se l'indirizzo viene modificato, attiva automaticamente la protezione.
    Se l'indirizzo viene svuotato, disattiva la protezione.
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    via = request.form.get('via', '').strip().upper()
    cap = request.form.get('cap', '').strip()
    citta = request.form.get('citta', '').strip().upper()
    provincia = request.form.get('provincia', '').strip().upper()
    
    # Costruisci indirizzo completo
    indirizzo_parti = []
    if via:
        indirizzo_parti.append(via)
    if cap or citta:
        parte_citta = []
        if cap:
            parte_citta.append(cap)
        if citta:
            parte_citta.append(citta)
        if indirizzo_parti:
            indirizzo_parti.append(', ' + ' '.join(parte_citta))
        else:
            indirizzo_parti.append(' '.join(parte_citta))
    if provincia:
        indirizzo_parti.append(' ' + provincia)
    
    indirizzo = ''.join(indirizzo_parti) if indirizzo_parti else None
    
    # Estrai civico dalla via se presente (es: "VIA ROMA 123" -> via="VIA ROMA", civico="123")
    civico = None
    if via:
        import re
        m = re.match(r'^(.+?)\s+(\d+[A-Za-z0-9/\-]*)\s*$', via)
        if m:
            via = m.group(1).strip()
            civico = m.group(2).strip()
    
    # Protezione: leggi dal checkbox (utente decide se proteggere)
    indirizzo_protetto = 1 if request.form.get('indirizzo_protetto') else 0
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        UPDATE clienti SET
            indirizzo = ?, via = ?, civico = ?, cap = ?, citta = ?, provincia = ?,
            indirizzo_protetto = ?, data_ultimo_aggiornamento = ?
        WHERE id = ?
    ''', (indirizzo, via or None, civico, cap or None, citta or None, provincia or None,
          indirizzo_protetto, now, cliente_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))



# ==============================================================================
# ROUTE: MODIFICA CAPOGRUPPO CLIENTE
# ==============================================================================

@app.route('/cliente/<int:cliente_id>/capogruppo/modifica', methods=['POST'])
def modifica_capogruppo_cliente(cliente_id):
    """
    Modifica i dati capogruppo del cliente.
    Protezione controllata da checkbox utente.
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    capogruppo_nome = request.form.get('capogruppo_nome', '').strip()
    capogruppo_cf = request.form.get('capogruppo_cf', '').strip().upper()
    
    # Protezione: leggi dal checkbox (utente decide se proteggere)
    capogruppo_protetto = 1 if request.form.get('capogruppo_protetto') else 0
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        UPDATE clienti SET
            capogruppo_nome = ?, capogruppo_cf = ?,
            capogruppo_protetto = ?, data_ultimo_aggiornamento = ?
        WHERE id = ?
    ''', (capogruppo_nome or None, capogruppo_cf or None,
          capogruppo_protetto, now, cliente_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))



# ==============================================================================
# ROUTE: MODIFICA SDI/BIC CLIENTE
# ==============================================================================

@app.route('/cliente/<int:cliente_id>/sdibic/modifica', methods=['POST'])
def modifica_sdibic_cliente(cliente_id):
    """
    Modifica i campi SDI e BIC del cliente.
    Questi sono dati inseriti manualmente.
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    sdi = request.form.get('sdi', '').strip().upper()
    bic = request.form.get('bic', '').strip().upper()
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        UPDATE clienti SET
            sdi = ?, bic = ?,
            data_ultimo_aggiornamento = ?
        WHERE id = ?
    ''', (sdi or None, bic or None, now, cliente_id))
    
    conn.commit()
    conn.close()
    
    return redirect(url_for('dettaglio_cliente', cliente_id=cliente_id))
# ==============================================================================
# ROUTE: CLIENTE PER IDENTIFICATIVO (P.IVA / CF)
# ==============================================================================

@app.route('/cerca/<identificativo>')
def cliente_per_identificativo(identificativo):
    """
    Visualizza dettaglio cliente usando P.IVA o CF.
    URL stabili: /cerca/IT00552060980 o /cerca/RSSMRA80A01F205X
    NON fa redirect, serve direttamente il template per mantenere l'URL stabile.
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    cliente = cerca_cliente_per_identificativo(cursor, identificativo)
    
    if not cliente:
        conn.close()
        html = """<!DOCTYPE html>
<html>
<head><title>Cliente non trovato</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container mt-5">
    <div class="alert alert-warning">
        <h4>Cliente non trovato</h4>
        <p>Nessun cliente con identificativo: <code>%s</code></p>
        <hr>
        <a href="/" class="btn btn-primary">Torna alla lista clienti</a>
    </div>
</div>
</body>
</html>""" % identificativo
        return html, 404
    
    conn.close()
    
    # Usa la funzione helper condivisa per renderizzare il dettaglio
    result = _render_dettaglio_cliente(cliente['id'])
    if result is None:
        return "Cliente non trovato", 404
    return result


@app.route('/api/cliente/<identificativo>')
def api_cliente(identificativo):
    """API JSON per ottenere dati cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    cliente = cerca_cliente_per_identificativo(cursor, identificativo)
    
    if not cliente:
        conn.close()
        return jsonify({
            'error': 'Cliente non trovato',
            'identificativo': identificativo
        }), 404
    
    cliente['_identificativo'] = get_identificativo_cliente(cliente)
    cliente['_url'] = url_cliente(cliente)
    
    cursor.execute('SELECT COUNT(*) FROM veicoli WHERE cliente_id = ?', (cliente['id'],))
    cliente['_num_veicoli'] = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(canone), 0) FROM veicoli WHERE cliente_id = ?', (cliente['id'],))
    cliente['_canone_totale'] = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM note_clienti WHERE cliente_id = ? AND eliminato = 0', (cliente['id'],))
    cliente['_num_note'] = cursor.fetchone()[0]
    
    conn.close()
    return jsonify(cliente)


@app.route('/api/cerca')
def api_cerca_cliente():
    """API ricerca clienti con ricerca smart fuzzy."""
    import re
    
    q = request.args.get('q', '').strip()
    limit = min(int(request.args.get('limit', 10)), 50)
    
    if not q or len(q) < 2:
        return jsonify({'error': 'Termine di ricerca troppo corto'}), 400
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Parametro normale per ricerca esatta
    search_param = '%' + q + '%'
    
    # Parametro normalizzato: rimuove punti, spazi, trattini, &, apostrofi
    # Cosi "at.ib" = "a t i b" = "a.t.i.b." = "atib"
    q_normalized = re.sub(r'[.\-&\'\s]+', '', q.lower())
    search_normalized = '%' + q_normalized + '%'
    
    # JOIN con referenti per prendere il referente principale
    # La ricerca normalizzata usa REPLACE multipli per pulire il campo DB
    # Nota: char(39) = apostrofo, evita conflitti con apici Python
    cursor.execute("""
        SELECT c.id, c.nome_cliente, c.ragione_sociale, c.p_iva, c.cod_fiscale, 
               c.score, c.credito, c.telefono, c.email,
               r.nome as ref_nome, r.cognome as ref_cognome, 
               r.telefono as ref_telefono, r.cellulare as ref_cellulare,
               r.email_principale as ref_email
        FROM clienti c
        LEFT JOIN referenti_clienti r ON r.cliente_id = c.id AND r.principale = 1
        WHERE 
           -- Ricerca normale (esatta)
           c.nome_cliente LIKE ? 
           OR c.ragione_sociale LIKE ? 
           OR c.p_iva LIKE ? 
           OR c.cod_fiscale LIKE ?
           -- Ricerca fuzzy: normalizza nome_cliente rimuovendo . - & ' e spazi
           OR LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(c.nome_cliente, '.', ''), ' ', ''), '-', ''), '&', ''), char(39), '')) LIKE ?
           -- Ricerca fuzzy: normalizza ragione_sociale
           OR LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(c.ragione_sociale, '.', ''), ' ', ''), '-', ''), '&', ''), char(39), '')) LIKE ?
        ORDER BY c.nome_cliente LIMIT ?
    """, (search_param, search_param, search_param, search_param, 
          search_normalized, search_normalized, limit))
    
    risultati = []
    for row in cursor.fetchall():
        cliente = dict(row)
        cliente['_identificativo'] = get_identificativo_cliente(cliente)
        cliente['_url'] = url_cliente(cliente)
        risultati.append(cliente)
    
    conn.close()
    return jsonify({'query': q, 'count': len(risultati), 'risultati': risultati})


# ==============================================================================

# ==============================================================================
# API: Referente principale
# ==============================================================================

@app.route('/api/cliente/<int:cliente_id>/referente-principale')
def api_referente_principale(cliente_id):
    """Ritorna i dati del referente principale di un cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT * FROM referenti_clienti 
        WHERE cliente_id = ? AND principale = 1
        LIMIT 1
    ''', (cliente_id,))
    
    row = cursor.fetchone()
    conn.close()
    
    if row:
        return jsonify({'success': True, 'referente': dict(row)})
    else:
        return jsonify({'success': False, 'error': 'Nessun referente principale'})


@app.route('/api/cliente/<int:cliente_id>/parco-potenziale', methods=['POST'])
@login_required
def api_parco_potenziale(cliente_id):
    """Salva il valore del parco potenziale per un cliente."""
    from flask import request
    data = request.get_json()
    valore = data.get('valore')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            UPDATE clienti SET parco_potenziale = ? WHERE id = ?
        ''', (valore, cliente_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/cliente/<int:cliente_id>/veicoli-rilevati', methods=['POST'])
@login_required
def api_veicoli_rilevati(cliente_id):
    """Salva il numero di veicoli rilevati per un cliente con data automatica."""
    from flask import request
    from datetime import datetime
    
    data = request.get_json()
    valore = data.get('valore')
    data_rilevazione = datetime.now().strftime('%d/%m/%Y') if valore else None
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            UPDATE clienti SET veicoli_rilevati = ?, data_rilevazione = ? WHERE id = ?
        ''', (valore, data_rilevazione, cliente_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'data_rilevazione': data_rilevazione})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})


# ==============================================================================
# API RICERCA COMMERCIALI (comando commbr)
# ==============================================================================

@app.route('/api/commerciali/cerca')
@app.route('/api/utenti/cerca')
@login_required
def api_cerca_utenti():
    """
    Cerca utenti per nome/cognome/username.
    Usato dai comandi @com e @ope nella ricerca smart.
    
    Parametri:
        q: termine di ricerca (opzionale)
        tipo: 'commerciale' o 'operatore' (default: commerciale)
    
    Regole:
        - Escludi sempre admin (ruolo_base='admin' o username='admin')
        - Escludi sempre utenti disattivati (attivo=0)
        - @com mostra solo ruolo_base='commerciale'
        - @ope mostra solo ruolo_base='operatore'
    """
    from app.database_utenti import get_connection as get_conn_utenti
    
    termine = request.args.get('q', '').strip().lower()
    tipo = request.args.get('tipo', 'commerciale').strip().lower()
    
    # Validazione tipo
    if tipo not in ('commerciale', 'operatore'):
        tipo = 'commerciale'
    
    conn = get_conn_utenti()
    cursor = conn.cursor()
    
    try:
        # Query base: escludi admin e utenti disattivati
        # Filtra per ruolo_base in base al tipo richiesto
        if termine:
            cursor.execute('''
                SELECT id, username, nome, cognome, email, cellulare, data_nascita, ruolo_base
                FROM utenti 
                WHERE ruolo_base = ?
                AND attivo = 1
                AND ruolo_base != 'admin'
                AND username != 'admin'
                AND (LOWER(nome) LIKE ? OR LOWER(cognome) LIKE ? OR LOWER(username) LIKE ?)
                ORDER BY cognome, nome
            ''', (tipo, f'%{termine}%', f'%{termine}%', f'%{termine}%'))
        else:
            cursor.execute('''
                SELECT id, username, nome, cognome, email, cellulare, data_nascita, ruolo_base
                FROM utenti 
                WHERE ruolo_base = ?
                AND attivo = 1
                AND ruolo_base != 'admin'
                AND username != 'admin'
                ORDER BY cognome, nome
            ''', (tipo,))
        
        risultati = []
        for row in cursor.fetchall():
            risultati.append({
                'id': row['id'],
                'username': row['username'],
                'nome': row['nome'],
                'cognome': row['cognome'],
                'email': row['email'],
                'cellulare': row['cellulare'],
                'data_nascita': row['data_nascita'],
                'ruolo': row['ruolo_base']
            })
        
        conn.close()
        # Retrocompatibilit&agrave;: mantieni 'commerciali' come chiave
        return jsonify({'success': True, 'commerciali': risultati, 'utenti': risultati, 'tipo': tipo})
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e), 'commerciali': [], 'utenti': []})


# ==============================================================================
# ROUTE NOTE FULLSCREEN
# ==============================================================================

@app.route('/cliente/<int:cliente_id>/note')
def note_fullscreen(cliente_id):
    """Vista fullscreen per gestione note cliente."""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM clienti WHERE id = ?', (cliente_id,))
    cliente = cursor.fetchone()
    
    if not cliente:
        conn.close()
        return "Cliente non trovato", 404
    
    cliente = dict(cliente)
    
    cursor.execute('''
        SELECT n.*, 
               (SELECT GROUP_CONCAT(a.id || '::' || a.nome_originale, '|||') 
                FROM allegati_note a WHERE a.nota_cliente_id = n.id) as allegati_raw
        FROM note_clienti n
        WHERE n.cliente_id = ? AND (n.eliminato = 0 OR n.eliminato IS NULL)
        ORDER BY n.fissata DESC, n.data_creazione DESC
    ''', (cliente_id,))
    
    note_rows = cursor.fetchall()
    note_cliente = []
    
    for row in note_rows:
        nota = dict(row)
        nota['allegati'] = []
        if nota.get('allegati_raw'):
            for a in nota['allegati_raw'].split('|||'):
                if '::' in a:
                    aid, nome = a.split('::', 1)
                    nota['allegati'].append({'id': int(aid), 'nome_originale': nome})
        if 'allegati_raw' in nota:
            del nota['allegati_raw']
        note_cliente.append(nota)
    
    nota_id = request.args.get('nota', type=int)
    nota_attiva = None
    nota_attiva_id = None
    
    if nota_id:
        for n in note_cliente:
            if n['id'] == nota_id:
                nota_attiva = n
                nota_attiva_id = nota_id
                break
    elif note_cliente:
        nota_attiva = note_cliente[0]
        nota_attiva_id = note_cliente[0]['id']
    
    cursor.execute('''
        SELECT DISTINCT commerciale FROM clienti 
        WHERE commerciale IS NOT NULL AND commerciale != ''
        ORDER BY commerciale
    ''')
    commerciali = [row[0] for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('note_fullscreen.html',
                         cliente=cliente,
                         note_cliente=note_cliente,
                         nota_attiva=nota_attiva,
                         nota_attiva_id=nota_attiva_id,
                         commerciali=commerciali)



def run_server(host=None, port=None, debug=None):

    """Avvia il server web."""
    host = host or WEB_HOST
    port = port or WEB_PORT
    debug = debug if debug is not None else WEB_DEBUG
    
    logger.info(f"Avvio server su http://{host}:{port}")
    
    # Notifica avvio sistema (campanella admin)
    try:
        from app.connettori_notifiche.sistema import notifica_avvio_sistema
        from app.database import get_connection
        conn = get_connection()
        notifica_avvio_sistema(conn)
        conn.close()
        logger.info("Notifica avvio sistema inviata")
    except Exception as e:
        logger.warning(f"Notifica avvio non inviata: {e}")
    app.run(host=host, port=port, debug=debug)


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Gestione Flotta - Web Server')
    parser.add_argument('--host', default=WEB_HOST, help='Host (default: 0.0.0.0)')
    parser.add_argument('--port', '-p', type=int, default=WEB_PORT, help='Porta (default: 5001)')
    parser.add_argument('--debug', action='store_true', help='Modalita debug')
    
    args = parser.parse_args()
    run_server(host=args.host, port=args.port, debug=args.debug)

